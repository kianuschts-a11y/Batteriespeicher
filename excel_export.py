import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break
from datetime import datetime
import io
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import base64
import tempfile
import os
from PIL import Image as PILImage

def clear_plot_cache():
    """
    Leert den Streamlit-Cache für Plot-Funktionen um sicherzustellen, 
    dass Grafiken neu generiert werden
    """
    try:
        # Leere den Streamlit-Cache für Plot-Funktionen
        if hasattr(st, 'cache_data'):
            st.cache_data.clear()
        if hasattr(st, 'cache_resource'):
            st.cache_resource.clear()
    except Exception as e:
        print(f"Cache konnte nicht geleert werden: {e}")

def plotly_fig_to_image(fig, width=None, height=None):
    """
    Konvertiert eine Plotly-Figur zu einem PIL-Image
    Verwendet größere Auflösung für bessere Lesbarkeit in Excel
    """
    try:
        # Methode 1: Verwende die eingebaute Plotly-Download-Funktionalität
        # Mit vergrößerter Größe und höherer Auflösung für bessere Lesbarkeit
        img_bytes = fig.to_image(format="png", width=1400, height=1000, scale=2.0)
        
        if img_bytes:
            # Konvertiere zu PIL Image
            img = PILImage.open(io.BytesIO(img_bytes))
            return img
        
        # Methode 2: Fallback - verwende mittlere Größe falls Methode 1 fehlschlägt
        img_bytes = fig.to_image(format="png", width=1200, height=800, scale=1.5)
        
        if img_bytes:
            # Konvertiere zu PIL Image
            img = PILImage.open(io.BytesIO(img_bytes))
            return img
        
        # Methode 3: Letzter Fallback - kleinere Größe
        img_bytes = fig.to_image(format="png", width=1000, height=700, scale=1.0)
        
        if img_bytes:
            # Konvertiere zu PIL Image
            img = PILImage.open(io.BytesIO(img_bytes))
            return img
        else:
            return None
            
    except Exception as e:
        st.error(f"Fehler beim Konvertieren der Grafik: {str(e)}")
        return None

def create_monthly_energy_flow_png(time_series_data, year, month, current_settings=None):
    """
    Erstellt PNG-Bilder für monatliche Energieflüsse - verwendet plot_energy_flows_for_period wie in der UI
    
    Args:
        time_series_data: DataFrame mit Zeitreihen
        year: Jahr
        month: Monatsnummer (1-12)
        current_settings: Settings-Dictionary
    
    Returns:
        PIL Image oder None
    """
    try:
        from analysis import plot_energy_flows_for_period
        import pandas as pd
        
        # Berechne Start- und Enddatum für den Monat
        start_date = f"{year}-{month:02d}-01"
        
        # Berechne letzten Tag des Monats
        if month == 12:
            end_date = f"{year}-{month:02d}-31"
        else:
            # Letzter Tag ist einen Tag vor dem ersten Tag des nächsten Monats
            next_month_start = pd.Timestamp(year=year, month=month+1, day=1)
            last_day = (next_month_start - pd.Timedelta(days=1)).day
            end_date = f"{year}-{month:02d}-{last_day:02d}"
        
        # Verwende die tatsächliche optimale Batteriekapazität
        battery_capacity = current_settings.get('optimal_battery_capacity_kwh', 10.0) if current_settings else 10.0
        
        # Generiere die GLEICHE Grafik wie in der UI
        fig = plot_energy_flows_for_period(time_series_data, start_date, end_date, battery_capacity)
        
        if fig:
            # Passe für Excel-Export an (weißer Hintergrund, schwarze Schrift)
            fig = prepare_figure_for_excel_export(fig)
            
            # WICHTIG: Deaktiviere Plotly's automatische Datenvereinfachung
            # um die volle Genauigkeit des Batterieladezustands zu erhalten
            for trace in fig.data:
                trace.update(simplify=False)  # Keine Datenvereinfachung
            
            # Konvertiere zu PNG mit hoher Auflösung (wie in der UI)
            # Verwende größere Auflösung für bessere Detailgenauigkeit
            img_bytes = fig.to_image(
                format="png", 
                width=1600,  # Erhöht von 1400 für mehr Details
                height=1200,  # Erhöht von 1000 für mehr Details
                scale=2.5,  # Erhöht von 2.0 für bessere Qualität
                engine="kaleido"
            )
            image = PILImage.open(io.BytesIO(img_bytes))
            
            return image
        else:
            print(f"Konnte keine Grafik für Monat {month} erstellen")
            return None
            
    except Exception as e:
        print(f"Fehler beim Erstellen des monatlichen Energiefluss-PNGs für Monat {month}: {e}")
        import traceback
        traceback.print_exc()
        return None

def create_cost_comparison_png(optimal_sim_result, current_settings, optimization_results=None, consumption_series=None, pv_generation_series=None, battery_cost_curve=None):
    """
    Erstellt PNG-Bild für Kostenvergleich - verwendet direkten PNG-Export wie manuell
    """
    cost_fig = None  # Initialisiere Variable für except-Block
    
    # Prüfe ob battery_cost_curve verfügbar ist
    if battery_cost_curve is None:
        print("WARNUNG: battery_cost_curve ist None in create_cost_comparison_png")
        return None
    
    try:
        from analysis import plot_cost_comparison_with_without_battery
        
        # Generiere die gleiche Grafik wie in der UI
        # Hole die optimale Batteriekapazität aus den Optimierungsergebnissen
        # Verwende das gleiche Optimierungskriterium wie in der UI
        if optimization_results:
            import pandas as pd
            df_results = pd.DataFrame(optimization_results)
            optimization_criterion = current_settings.get('optimization_criterion', 'Deckungsbeitrag III gesamt (Barwert)')
            # Nur DB Barwert und DB Nominal unterstützen
            if optimization_criterion == "Deckungsbeitrag III gesamt (Nominal)":
                best_idx = df_results['total_db3_nominal'].idxmax()
            else:  # Default: Deckungsbeitrag III gesamt (Barwert)
                best_idx = df_results['total_db3_present_value'].idxmax()
            optimal_capacity = df_results.loc[best_idx, 'battery_capacity_kwh']
        else:
            optimal_capacity = current_settings.get("battery_capacity_kwh", 10.0)  # Fallback
        
        cost_fig = plot_cost_comparison_with_without_battery(
            annual_energy_cost_with_battery=optimal_sim_result["kpis"]["annual_energy_cost"],
            total_consumption=optimal_sim_result["kpis"]["total_consumption_kwh"],
            total_pv_generation=optimal_sim_result["kpis"]["total_pv_generation_kwh"],
            battery_capacity_kwh=optimal_capacity,  # Verwende die optimale Kapazität
            battery_cost_curve=battery_cost_curve,
            project_lifetime_years=current_settings.get("project_lifetime_years", 20),
            discount_rate=current_settings.get("discount_rate", 0.05),
            price_grid_per_kwh=current_settings.get("price_grid_per_kwh", 0.30),
            price_feed_in_per_kwh=current_settings.get("price_feed_in_per_kwh", 0.12),
            annual_capacity_loss_percent=current_settings.get("annual_capacity_loss_percent", 2.0),
            consumption_series=consumption_series,  # Für tatsächliche Simulation ohne Batterie
            pv_generation_series=pv_generation_series,  # Für tatsächliche Simulation ohne Batterie
            battery_efficiency_charge=current_settings.get("battery_efficiency_charge", 0.95),
            battery_efficiency_discharge=current_settings.get("battery_efficiency_discharge", 0.95),
            battery_max_charge_kw=current_settings.get("battery_max_charge_kw", 10.0),
            battery_max_discharge_kw=current_settings.get("battery_max_discharge_kw", 10.0),
            initial_soc_percent=current_settings.get("initial_soc_percent", 50.0),
            min_soc_percent=current_settings.get("min_soc_percent", 10.0),
            max_soc_percent=current_settings.get("max_soc_percent", 90.0),
            grid_export_with_battery=optimal_sim_result["kpis"].get("total_grid_export_kwh", 0),
            grid_import_with_battery=optimal_sim_result["kpis"].get("total_grid_import_kwh", 0)
        )
        
        if cost_fig and 'figure' in cost_fig:
            # Passe für Excel-Export an (weißer Hintergrund, schwarze Schrift)
            fig = prepare_figure_for_excel_export(cost_fig['figure'])
            
            # Deaktiviere Datenvereinfachung für präzise Darstellung
            for trace in fig.data:
                trace.update(simplify=False)
            
            # Direkter PNG-Export wie beim manuellen Export
            img_bytes = fig.to_image(
                format="png", 
                width=1400, 
                height=1000, 
                scale=2.0,
                engine="kaleido"  # Verwende kaleido für bessere Kompatibilität
            )
            
            if img_bytes:
                image = PILImage.open(io.BytesIO(img_bytes))
                return image
        
        return None
    except Exception as e:
        print(f"Fehler beim direkten PNG-Export des Kostenvergleichs: {e}")
        # Fallback: Verwende die alte plotly_fig_to_image Funktion
        if cost_fig and 'figure' in cost_fig:
            fig = prepare_figure_for_excel_export(cost_fig['figure'])
            return plotly_fig_to_image(fig)
        return None

def create_optimization_curve_png(optimization_results, current_settings):
    """
    Erstellt PNG-Bild für Optimierungskurve - verwendet direkten PNG-Export wie manuell
    """
    try:
        from analysis import plot_optimization_curve
        
        # Generiere die gleiche Grafik wie in der UI - stelle sicher, dass sie neu generiert wird
        opt_fig = plot_optimization_curve(optimization_results)
        
        if opt_fig:
            # Passe für Excel-Export an (weißer Hintergrund, schwarze Schrift)
            opt_fig = prepare_figure_for_excel_export(opt_fig)
            
            # Deaktiviere Datenvereinfachung für präzise Darstellung
            for trace in opt_fig.data:
                trace.update(simplify=False)
            
            # Direkter PNG-Export wie beim manuellen Export
            img_bytes = opt_fig.to_image(
                format="png", 
                width=1400, 
                height=1000, 
                scale=2.0,
                engine="kaleido"  # Verwende kaleido für bessere Kompatibilität
            )
            
            if img_bytes:
                image = PILImage.open(io.BytesIO(img_bytes))
                return image
        
        return None
    except Exception as e:
        print(f"Fehler beim direkten PNG-Export der Optimierungskurve: {e}")
        # Fallback: Verwende die alte plotly_fig_to_image Funktion
        if opt_fig:
            opt_fig = prepare_figure_for_excel_export(opt_fig)
            return plotly_fig_to_image(opt_fig)
        return None

def prepare_figure_for_excel_export(fig):
    """
    Passt eine Plotly-Figure für Excel-Export an (weißer Hintergrund, schwarze Schrift)
    """
    fig.update_layout(
        paper_bgcolor='white',
        plot_bgcolor='white',
        font=dict(color='black'),
        title_font=dict(color='black'),
        legend=dict(
            bgcolor='white',
            bordercolor='gray',
            font=dict(color='black')
        )
    )
    
    # Prüfe ob es Subplots gibt und aktualisiere alle Achsen explizit
    if hasattr(fig, 'layout') and hasattr(fig.layout, 'xaxis'):
        # Zähle die Anzahl der Achsen
        axis_count = 1
        while hasattr(fig.layout, f'xaxis{axis_count + 1}'):
            axis_count += 1
        
        # Aktualisiere alle X- und Y-Achsen
        for i in range(1, axis_count + 1):
            if i == 1:
                xaxis_name = 'xaxis'
                yaxis_name = 'yaxis'
            else:
                xaxis_name = f'xaxis{i}'
                yaxis_name = f'yaxis{i}'
            
            # X-Achse
            if hasattr(fig.layout, xaxis_name):
                xaxis = getattr(fig.layout, xaxis_name)
                xaxis.update(
                    title_font=dict(color='black', size=12),
                    tickfont=dict(color='black', size=10),
                    gridcolor='lightgray',
                    showgrid=True,
                    linecolor='black',
                    showline=True
                )
            
            # Y-Achse
            if hasattr(fig.layout, yaxis_name):
                yaxis = getattr(fig.layout, yaxis_name)
                yaxis.update(
                    title_font=dict(color='black', size=12),
                    tickfont=dict(color='black', size=10),
                    gridcolor='lightgray',
                    showgrid=True,
                    linecolor='black',
                    showline=True
                )
    
    # Fallback: Verwende update_xaxes/update_yaxes für alle Achsen
    fig.update_xaxes(
        title_font=dict(color='black', size=12),
        tickfont=dict(color='black', size=10),
        gridcolor='lightgray',
        showgrid=True,
        linecolor='black',
        showline=True
    )
    fig.update_yaxes(
        title_font=dict(color='black', size=12),
        tickfont=dict(color='black', size=10),
        gridcolor='lightgray',
        showgrid=True,
        linecolor='black',
        showline=True
    )
    
    # Stelle sicher, dass Subplot-Annotations (Titel) auch schwarz sind
    if fig.layout.annotations:
        for annotation in fig.layout.annotations:
            annotation.font.color = 'black'
    
    return fig

def create_sankey_diagram_png(kpis):
    """
    Erstellt PNG-Bild für Sankey-Diagramm - verwendet direkten PNG-Export wie manuell
    """
    try:
        from analysis import plot_sankey_diagram
        
        # Generiere die gleiche Grafik wie in der UI
        sankey_result = plot_sankey_diagram(kpis)
        
        if sankey_result and 'figure' in sankey_result:
            # Passe für Excel-Export an (weißer Hintergrund, schwarze Schrift)
            fig = prepare_figure_for_excel_export(sankey_result['figure'])
            
            # Deaktiviere Datenvereinfachung für präzise Darstellung
            for trace in fig.data:
                trace.update(simplify=False)
            
            # Direkter PNG-Export wie beim manuellen Export
            img_bytes = fig.to_image(
                format="png", 
                width=1400, 
                height=1000, 
                scale=2.0,
                engine="kaleido"  # Verwende kaleido für bessere Kompatibilität
            )
            
            if img_bytes:
                image = PILImage.open(io.BytesIO(img_bytes))
                return image
        
        return None
    except Exception as e:
        print(f"Fehler beim direkten PNG-Export des Sankey-Diagramms: {e}")
        # Fallback: Verwende die alte plotly_fig_to_image Funktion
        if sankey_result and 'figure' in sankey_result:
            fig = prepare_figure_for_excel_export(sankey_result['figure'])
            return plotly_fig_to_image(fig)
        return None

def create_sankey_diagram_no_battery_png(kpis_no_battery, total_consumption, total_pv_generation):
    """
    Erstellt PNG-Bild für Sankey-Diagramm OHNE Batteriespeicher - vereinfachte Version
    """
    try:
        import plotly.graph_objects as go
        
        # Vereinfachte Knoten ohne Batterie
        labels = [
            "PV-Erzeugung", "Netzbezug",  # Quellen
            "Verbrauch",  # Senke
            "Direkter Eigenverbrauch", "Netzeinspeisung"  # Zwischenstationen
        ]
        
        # Indizes
        pv_gen_idx = 0
        grid_import_idx = 1
        consumption_idx = 2
        direct_self_cons_idx = 3
        grid_export_idx = 4
        
        # Werte aus KPIs (falls verfügbar) oder aus Parametern
        if isinstance(kpis_no_battery, dict):
            grid_import = kpis_no_battery.get("total_grid_import_kwh", 0)
            grid_export = kpis_no_battery.get("total_grid_export_kwh", 0)
            direct_self_consumption = kpis_no_battery.get("total_direct_self_consumption_kwh", 0)
        else:
            # Fallback: Berechne aus Gesamtwerten
            grid_import = max(0, total_consumption - min(total_pv_generation, total_consumption))
            grid_export = max(0, total_pv_generation - min(total_pv_generation, total_consumption))
            direct_self_consumption = min(total_pv_generation, total_consumption)
        
        # Flüsse definieren
        source = [
            pv_gen_idx, pv_gen_idx, pv_gen_idx,  # PV-Erzeugung -> 3 Ziele
            grid_import_idx,  # Netzbezug -> Verbrauch
            direct_self_cons_idx  # Direkter Eigenverbrauch -> Verbrauch
        ]
        target = [
            direct_self_cons_idx, grid_export_idx, consumption_idx,  # Von PV
            consumption_idx,  # Von Netzbezug
            consumption_idx  # Von direktem Eigenverbrauch
        ]
        value = [
            direct_self_consumption,  # PV -> Direkter Eigenverbrauch
            grid_export,  # PV -> Netzeinspeisung
            0,  # PV -> Verbrauch (direkt, wird über Eigenverbrauch abgedeckt)
            grid_import,  # Netzbezug -> Verbrauch
            direct_self_consumption  # Direkter Eigenverbrauch -> Verbrauch
        ]
        
        # Filtere Null-Werte
        filtered_source = []
        filtered_target = []
        filtered_value = []
        for s, t, v in zip(source, target, value):
            if v > 0.01:
                filtered_source.append(s)
                filtered_target.append(t)
                filtered_value.append(v)
        
        # Farben für Knoten
        node_colors = [
            "#2E8B57",  # PV-Erzeugung (Dunkelgrün)
            "#4169E1",  # Netzbezug (Königsblau)
            "#DC143C",  # Verbrauch (Karmesinrot)
            "#32CD32",  # Direkter Eigenverbrauch (Limettengrün)
            "#FF8C00"   # Netzeinspeisung (Dunkelorange)
        ]
        
        # Farben für Flüsse
        link_colors = []
        for s, t in zip(filtered_source, filtered_target):
            if s == pv_gen_idx:
                if t == direct_self_cons_idx:
                    link_colors.append("rgba(50, 205, 50, 0.6)")
                elif t == grid_export_idx:
                    link_colors.append("rgba(255, 140, 0, 0.6)")
                else:
                    link_colors.append("rgba(46, 139, 87, 0.4)")
            elif s == grid_import_idx:
                link_colors.append("rgba(65, 105, 225, 0.6)")
            elif s == direct_self_cons_idx:
                link_colors.append("rgba(50, 205, 50, 0.8)")
            else:
                link_colors.append("rgba(128, 128, 128, 0.4)")
        
        # Erstelle Diagramm
        fig = go.Figure(data=[go.Sankey(
            node=dict(
                label=labels,
                pad=25,
                thickness=30,
                line=dict(color="black", width=1.5),
                color=node_colors
            ),
            link=dict(
                source=filtered_source,
                target=filtered_target,
                value=filtered_value,
                color=link_colors,
                line=dict(width=0.8)
            ),
            arrangement="snap"
        )])
        
        fig.update_layout(
            title_text="Jährliche Energieflüsse ohne Batteriespeicher",
            title_font_size=16,
            title_font_color="black",
            title_x=0.5,
            font=dict(size=12, color="black"),
            height=1000,
            width=1400,
            paper_bgcolor="white",
            plot_bgcolor="white"
        )
        
        # Exportiere als PNG
        img_bytes = fig.to_image(
            format="png",
            width=1400,
            height=1000,
            scale=2.0,
            engine="kaleido"
        )
        
        if img_bytes:
            image = PILImage.open(io.BytesIO(img_bytes))
            return image
        
        return None
    except Exception as e:
        print(f"Fehler beim Erstellen des Sankey-Diagramms ohne Batterie: {e}")
        import traceback
        traceback.print_exc()
        return None

def create_capacity_aging_sheet(wb, battery_capacity_kwh, annual_capacity_loss_percent, project_lifetime_years):
    """
    Erstellt ein neues Sheet mit der Kapazitätsentwicklung über die Jahre
    """
    ws = wb.create_sheet("Kapazitätsalterung")
    
    # Überschrift
    ws['A1'] = "Kapazitätsalterung der Batterie über die Projektlaufzeit"
    ws['A1'].font = Font(size=16, bold=True, color="FF6B35")
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Parameter
    ws['A3'] = "Parameter:"
    ws['A3'].font = Font(size=12, bold=True)
    ws['A4'] = "Ursprüngliche Batteriekapazität:"
    ws['B4'] = f"{battery_capacity_kwh} kWh"
    ws['A5'] = "Jährlicher Kapazitätsverlust:"
    ws['B5'] = f"{annual_capacity_loss_percent:.2f} %"
    ws['A6'] = "Projektlaufzeit:"
    ws['B6'] = f"{project_lifetime_years} Jahre"
    
    # Tabellenkopf
    ws['A8'] = "Jahr"
    ws['B8'] = "Kapazitätsfaktor"
    ws['C8'] = "Verfügbare Kapazität (kWh)"
    ws['D8'] = "Kapazitätsverlust (%)"
    ws['E8'] = "Verbleibende Kapazität (%)"
    
    # Styling für Tabellenkopf
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
    header_alignment = Alignment(horizontal='center')
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}8'].font = header_font
        ws[f'{col}8'].fill = header_fill
        ws[f'{col}8'].alignment = header_alignment
    
    # Daten berechnen und eintragen
    row = 9
    for year in range(1, project_lifetime_years + 1):
        capacity_factor = (1.0 - annual_capacity_loss_percent / 100.0) ** (year - 1)
        current_capacity_kwh = battery_capacity_kwh * capacity_factor
        capacity_loss_percent = ((battery_capacity_kwh - current_capacity_kwh) / battery_capacity_kwh) * 100
        remaining_capacity_percent = (current_capacity_kwh / battery_capacity_kwh) * 100
        
        ws[f'A{row}'] = year
        ws[f'B{row}'] = f"{capacity_factor:.4f}"
        ws[f'C{row}'] = f"{current_capacity_kwh}"
        ws[f'D{row}'] = f"{capacity_loss_percent:.2f} %" if not np.isnan(capacity_loss_percent) else "0.00 %"
        ws[f'E{row}'] = f"{remaining_capacity_percent:.2f} %" if not np.isnan(remaining_capacity_percent) else "100.00 %"
        
        # Farbkodierung für verbleibende Kapazität
        if remaining_capacity_percent >= 80:
            cell_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Hellgrün
        elif remaining_capacity_percent >= 60:
            cell_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Hellgelb
        elif remaining_capacity_percent >= 40:
            cell_fill = PatternFill(start_color="FFB366", end_color="FFB366", fill_type="solid")  # Orange
        else:
            cell_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")  # Hellrot
        
        ws[f'E{row}'].fill = cell_fill
        row += 1
    
    # Spaltenbreiten anpassen
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 20
    
    # Zusammenfassung
    summary_row = row + 2
    ws[f'A{summary_row}'] = "Zusammenfassung:"
    ws[f'A{summary_row}'].font = Font(size=12, bold=True)
    
    final_capacity_kwh = battery_capacity_kwh * (1.0 - annual_capacity_loss_percent / 100.0) ** (project_lifetime_years - 1)
    total_capacity_loss_percent = ((battery_capacity_kwh - final_capacity_kwh) / battery_capacity_kwh) * 100
    
    ws[f'A{summary_row + 1}'] = "Kapazität nach Projektende:"
    ws[f'B{summary_row + 1}'] = f"{final_capacity_kwh} kWh"
    ws[f'A{summary_row + 2}'] = "Gesamter Kapazitätsverlust:"
    ws[f'B{summary_row + 2}'] = f"{total_capacity_loss_percent:.2f} %" if not np.isnan(total_capacity_loss_percent) else "0.00 %"
    
    return ws

def add_chart_to_excel(ws, chart_image, row, col, title, width_cm=None, height_cm=None, add_title=True):
    """
    Fügt ein Chart-Bild zu einem Excel-Worksheet hinzu
    Verwendet die Originalgröße der Grafik, falls nicht anders angegeben
    
    Args:
        ws: Worksheet
        chart_image: PIL Image
        row: Startzeile (1-basiert)
        col: Startspalte (0-basiert, 0=A, 1=B, etc.)
        title: Titel der Grafik
        width_cm: Breite in cm (optional)
        height_cm: Höhe in cm (optional)
        add_title: Ob ein Titel über der Grafik hinzugefügt werden soll (default: True)
    """
    try:
        # Konvertiere PIL Image zu Bytes
        img_buffer = io.BytesIO()
        chart_image.save(img_buffer, format='PNG')
        img_buffer.seek(0)
        
        current_row = row
        
        # Füge Titel hinzu (falls gewünscht)
        if add_title and title:
            col_letter = chr(65+col)
            # Zuerst Wert und Formatierung setzen, dann mergen
            cell_title = ws[f'{col_letter}{current_row}']
            cell_title.value = title
            cell_title.font = Font(size=12, bold=True, color="FF6B35")
            cell_title.alignment = Alignment(horizontal='center')
            # Jetzt mergen
            ws.merge_cells(f'{col_letter}{current_row}:{chr(65+col+3)}{current_row}')
            current_row += 1
        
        # Erstelle temporäre Datei für openpyxl
        with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp_file:
            tmp_file.write(img_buffer.getvalue())
            tmp_file.flush()
            tmp_file.close()  # Explizit schließen
            
            # Füge das Bild hinzu
            img = Image(tmp_file.name)
            
            # Verwende vergrößerte Darstellung für bessere Lesbarkeit
            if width_cm is None or height_cm is None:
                # Hole die Originalgröße der PIL-Image
                original_width, original_height = chart_image.size
                
                # Konvertiere Pixel zu Punkten mit vergrößerter Darstellung
                # 1 Pixel ≈ 0.75 Punkte bei 96 DPI, dann mit Faktor 0.7 für größere Darstellung
                width_pt = (original_width * 0.75) * 0.7  # Größere Darstellung als vorher
                height_pt = (original_height * 0.75) * 0.7
            else:
                # Verwende angegebene Größe in cm
                width_pt = width_cm * 28.35
                height_pt = height_cm * 28.35
            
            img.width = width_pt
            img.height = height_pt
            
            # Positioniere das Bild
            img.anchor = f'{chr(65+col)}{current_row}'
            
            ws.add_image(img)
            
            # Speichere den Dateipfad für späteres Löschen
            ws._temp_files = getattr(ws, '_temp_files', [])
            ws._temp_files.append(tmp_file.name)
        
        # Berechne die nächste Zeile basierend auf der Höhe des Bildes
        rows_needed = int(height_pt / 15) + 2  # 15 Punkte pro Zeile, plus etwas Puffer
        return max(current_row + rows_needed, row + int(height_pt / 15) + 3)
        
    except Exception as e:
        st.error(f"Fehler beim Hinzufügen der Grafik: {str(e)}")
        return row + 5

def add_two_charts_side_by_side(ws, left_image, right_image, row, left_title, right_title):
    """
    Fügt zwei Chart-Bilder nebeneinander in Excel ein
    
    Args:
        ws: Worksheet
        left_image: PIL Image für linkes Diagramm
        right_image: PIL Image für rechtes Diagramm
        row: Startzeile (1-basiert)
        left_title: Titel für linkes Diagramm
        right_title: Titel für rechtes Diagramm
    
    Returns:
        int: Nächste verfügbare Zeile
    """
    try:
        current_row = row
        
        # Titel hinzufügen
        ws[f'A{current_row}'] = left_title
        ws[f'A{current_row}'].font = Font(size=12, bold=True, color="FF6B35")
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        ws.merge_cells(f'A{current_row}:D{current_row}')
        
        ws[f'E{current_row}'] = right_title
        ws[f'E{current_row}'].font = Font(size=12, bold=True, color="FF6B35")
        ws[f'E{current_row}'].alignment = Alignment(horizontal='center')
        ws.merge_cells(f'E{current_row}:H{current_row}')
        
        current_row += 1
        
        # Linkes Bild
        left_buffer = io.BytesIO()
        left_image.save(left_buffer, format='PNG')
        left_buffer.seek(0)
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp_left:
            tmp_left.write(left_buffer.getvalue())
            tmp_left.flush()
            tmp_left.close()
            
            img_left = Image(tmp_left.name)
            left_width, left_height = left_image.size
            img_left.width = (left_width * 0.75) * 0.35  # Kleinere Darstellung für nebeneinander
            img_left.height = (left_height * 0.75) * 0.35
            img_left.anchor = f'A{current_row}'
            ws.add_image(img_left)
            
            ws._temp_files = getattr(ws, '_temp_files', [])
            ws._temp_files.append(tmp_left.name)
        
        # Rechtes Bild
        right_buffer = io.BytesIO()
        right_image.save(right_buffer, format='PNG')
        right_buffer.seek(0)
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp_right:
            tmp_right.write(right_buffer.getvalue())
            tmp_right.flush()
            tmp_right.close()
            
            img_right = Image(tmp_right.name)
            right_width, right_height = right_image.size
            img_right.width = (right_width * 0.75) * 0.35
            img_right.height = (right_height * 0.75) * 0.35
            img_right.anchor = f'E{current_row}'
            ws.add_image(img_right)
            
            ws._temp_files = getattr(ws, '_temp_files', [])
            ws._temp_files.append(tmp_right.name)
        
        # Berechne nächste Zeile
        max_height = max(img_left.height, img_right.height)
        rows_needed = int(max_height / 15) + 2
        return current_row + rows_needed
        
    except Exception as e:
        st.error(f"Fehler beim Hinzufügen der nebeneinander liegenden Grafiken: {str(e)}")
        return row + 5

def add_sankey_comparison_table(ws, row, kpis_no_battery, kpis_with_battery, total_consumption, total_pv_generation):
    """
    Fügt eine Vergleichstabelle unter den Sankey-Diagrammen hinzu
    
    Args:
        ws: Worksheet
        row: Startzeile
        kpis_no_battery: KPIs ohne Batterie
        kpis_with_battery: KPIs mit Batterie
        total_consumption: Gesamtverbrauch
        total_pv_generation: Gesamt-PV-Erzeugung
    
    Returns:
        int: Nächste verfügbare Zeile
    """
    try:
        current_row = row + 1
        
        # Titel
        ws[f'A{current_row}'] = "Vergleich der Energieflüsse"
        ws[f'A{current_row}'].font = Font(size=14, bold=True, color="FF6B35")
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        ws.merge_cells(f'A{current_row}:H{current_row}')
        current_row += 2
        
        # Header-Zeile
        headers = ['Kennzahl', 'Ohne Batterie', 'Mit Batterie', 'Differenz', 'Differenz %']
        header_cols = ['A', 'C', 'E', 'G', 'H']
        
        for col, header in zip(header_cols, headers):
            cell = ws[f'{col}{current_row}']
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Merge für breitere Spalten
        ws.merge_cells(f'A{current_row}:B{current_row}')
        ws.merge_cells(f'C{current_row}:D{current_row}')
        ws.merge_cells(f'E{current_row}:F{current_row}')
        
        current_row += 1
        
        # Berechne Werte
        grid_import_no_bat = kpis_no_battery.get("total_grid_import_kwh", 0)
        grid_import_with_bat = kpis_with_battery.get("total_grid_import_kwh", 0)
        
        grid_export_no_bat = kpis_no_battery.get("total_grid_export_kwh", 0)
        grid_export_with_bat = kpis_with_battery.get("total_grid_export_kwh", 0)
        
        self_cons_no_bat = kpis_no_battery.get("total_direct_self_consumption_kwh", 0)
        self_cons_with_bat = kpis_with_battery.get("total_direct_self_consumption_kwh", 0) + kpis_with_battery.get("total_battery_discharge_kwh", 0)
        
        autarky_no_bat = (1 - grid_import_no_bat / total_consumption) * 100 if total_consumption > 0 else 0
        autarky_with_bat = kpis_with_battery.get("autarky_rate", 0) * 100
        
        self_cons_rate_no_bat = (self_cons_no_bat / total_pv_generation) * 100 if total_pv_generation > 0 else 0
        self_cons_rate_with_bat = kpis_with_battery.get("self_consumption_rate", 0) * 100
        
        battery_charge = kpis_with_battery.get("total_battery_charge_kwh", 0)
        battery_discharge = kpis_with_battery.get("total_battery_discharge_kwh", 0)
        battery_losses = battery_charge - battery_discharge
        
        # Datenzeilen
        data_rows = [
            ("PV-Erzeugung (kWh)", total_pv_generation, total_pv_generation, 0, 0),
            ("Gesamtverbrauch (kWh)", total_consumption, total_consumption, 0, 0),
            ("Netzbezug (kWh)", grid_import_no_bat, grid_import_with_bat, 
             grid_import_with_bat - grid_import_no_bat, 
             ((grid_import_with_bat - grid_import_no_bat) / grid_import_no_bat * 100) if grid_import_no_bat > 0 else 0),
            ("Netzeinspeisung (kWh)", grid_export_no_bat, grid_export_with_bat,
             grid_export_with_bat - grid_export_no_bat,
             ((grid_export_with_bat - grid_export_no_bat) / grid_export_no_bat * 100) if grid_export_no_bat > 0 else 0),
            ("Eigenverbrauch (kWh)", self_cons_no_bat, self_cons_with_bat,
             self_cons_with_bat - self_cons_no_bat,
             ((self_cons_with_bat - self_cons_no_bat) / self_cons_no_bat * 100) if self_cons_no_bat > 0 else 0),
            ("Autarkiegrad (%)", autarky_no_bat, autarky_with_bat,
             autarky_with_bat - autarky_no_bat,
             ((autarky_with_bat - autarky_no_bat) / autarky_no_bat * 100) if autarky_no_bat > 0 else 0),
            ("Eigenverbrauchsquote (%)", self_cons_rate_no_bat, self_cons_rate_with_bat,
             self_cons_rate_with_bat - self_cons_rate_no_bat,
             ((self_cons_rate_with_bat - self_cons_rate_no_bat) / self_cons_rate_no_bat * 100) if self_cons_rate_no_bat > 0 else 0),
            ("Batterie Ladung (kWh)", "-", battery_charge, "-", "-"),
            ("Batterie Entladung (kWh)", "-", battery_discharge, "-", "-"),
            ("Batterieverluste (kWh)", "-", battery_losses, "-", "-"),
        ]
        
        # Füge Datenzeilen hinzu
        for label, val_no_bat, val_with_bat, diff, diff_pct in data_rows:
            # Label
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
            ws.merge_cells(f'A{current_row}:B{current_row}')
            
            # Ohne Batterie
            if val_no_bat == "-":
                ws[f'C{current_row}'] = val_no_bat
                ws[f'C{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            else:
                ws[f'C{current_row}'] = val_no_bat
                ws[f'C{current_row}'].number_format = '#,##0.00'
                ws[f'C{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
            ws.merge_cells(f'C{current_row}:D{current_row}')
            
            # Mit Batterie
            ws[f'E{current_row}'] = val_with_bat
            if val_with_bat != "-":
                ws[f'E{current_row}'].number_format = '#,##0.00'
            ws[f'E{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
            ws[f'E{current_row}'].fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            ws.merge_cells(f'E{current_row}:F{current_row}')
            
            # Differenz
            if diff == "-":
                ws[f'G{current_row}'] = diff
                ws[f'G{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            else:
                ws[f'G{current_row}'] = diff
                ws[f'G{current_row}'].number_format = '#,##0.00'
                ws[f'G{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
                # Färbe negative Werte grün (Verbesserung), positive rot (Verschlechterung)
                if diff < 0 and "Netzbezug" in label:
                    ws[f'G{current_row}'].font = Font(color="2E7D32", bold=True)
                elif diff > 0 and ("Autarkie" in label or "Eigenverbrauch" in label):
                    ws[f'G{current_row}'].font = Font(color="2E7D32", bold=True)
            
            # Differenz %
            if diff_pct == "-":
                ws[f'H{current_row}'] = diff_pct
                ws[f'H{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            else:
                ws[f'H{current_row}'] = diff_pct / 100  # Als Dezimalzahl für Prozentformat
                ws[f'H{current_row}'].number_format = '0.00%'
                ws[f'H{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
            
            current_row += 1
        
        # Spaltenbreiten anpassen
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 12
        
        return current_row + 2
        
    except Exception as e:
        print(f"Fehler beim Erstellen der Sankey-Vergleichstabelle: {e}")
        return row + 5

def calculate_daily_consumption_breakdown(consumption_series, annual_consumption_kwh, number_of_persons, bundesland_code, selected_year, file_path=None, profile_type=None):
    """
    Berechnet die detaillierte Verbrauchsaufschlüsselung für jeden Tag des Jahres.
    
    Args:
        consumption_series: Verbrauchsdaten mit 15-Minuten-Intervallen
        annual_consumption_kwh: Jährlicher Verbrauch in kWh
        number_of_persons: Anzahl der Haushalte
        bundesland_code: Bundesland-Code
        selected_year: Jahr
        file_path: Pfad zur Excel-Datei (für direkte Rohwert-Extraktion)
        profile_type: Profiltyp (H25, G25, etc.)
        
    Returns:
        Dictionary mit täglicher, monatlicher und jährlicher Verbrauchsaufschlüsselung
    """
    try:
        # Lade Wochentag-Sheet für Tagestyp-Bestimmung
        import pandas as pd
        from data_import import create_weekday_sheet
        
        # Erstelle Wochentag-Sheet für das Jahr
        weekday_df = create_weekday_sheet(selected_year)
        
        # Gruppiere 15-Minuten-Daten zu Tagesdaten
        daily_consumption = consumption_series.resample('D').sum()
        
        # Erstelle DataFrame für die Aufschlüsselung
        breakdown_data = []
        
        # Lade Rohwerte direkt aus Excel basierend auf Bundesland und Tagestyp
        raw_daily_values = {}
        if file_path:
            try:
                # Versuche verschiedene Sheet-Namen für das Lastprofil
                df_profile_raw = None
                for sheet_name in ['Verbrauchsdaten', 'Lastprofil', 'Custom', 'H25', 'G25']:
                    try:
                        df_profile_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                        print(f"Lastprofil-Sheet '{sheet_name}' erfolgreich geladen")
                        break
                    except:
                        continue
                
                if df_profile_raw is None:
                    print("Kein Lastprofil-Sheet gefunden, verwende Fallback-Berechnung")
                else:
                    # Finde die Header-Zeile mit WT/SA/FT
                    header_row = None
                    for i in range(min(10, len(df_profile_raw))):
                        row_str = str(df_profile_raw.iloc[i])
                        if 'WT' in row_str and 'SA' in row_str and 'FT' in row_str:
                            header_row = i
                            print(f"Header-Zeile gefunden in Zeile {i}")
                            break
                    
                    if header_row is not None:
                        # Extrahiere Spaltenindizes für jeden Monat und Tagestyp
                        header_values = df_profile_raw.iloc[header_row]
                        month_cols = {}
                        
                        # Finde Spalten für jeden Monat (Januar = 1, Februar = 2, etc.)
                        for month in range(1, 13):
                            month_cols[month] = {}
                            # Suche nach SA, FT, WT Spalten für diesen Monat
                            for col_idx, value in enumerate(header_values):
                                if str(value) == 'SA':
                                    month_cols[month]['SA'] = col_idx
                                elif str(value) == 'FT':
                                    month_cols[month]['FT'] = col_idx
                                elif str(value) == 'WT':
                                    month_cols[month]['WT'] = col_idx
                        
                        print(f"Gefundene Spalten: {month_cols}")
                        
                        # Berechne tägliche Rohwerte basierend auf Bundesland und Tagestyp
                        for date, daily_kwh in daily_consumption.items():
                            day_of_year = date.timetuple().tm_yday
                            if day_of_year <= len(weekday_df):
                                day_row = weekday_df.iloc[day_of_year - 1]
                                day_type = day_row[f'{bundesland_code}_Tagestyp']
                                month = date.month
                                
                                if month in month_cols and day_type in month_cols[month]:
                                    col_idx = month_cols[month][day_type]
                                    # Summiere alle 15-Minuten-Werte für diesen Tag (96 Werte)
                                    daily_raw_sum = 0
                                    for row_idx in range(header_row + 1, min(header_row + 97, len(df_profile_raw))):
                                        try:
                                            value = df_profile_raw.iloc[row_idx, col_idx]
                                            if pd.notna(value) and isinstance(value, (int, float)):
                                                daily_raw_sum += value
                                        except:
                                            continue
                                    raw_daily_values[date] = daily_raw_sum
                                    print(f"Datum {date.strftime('%d.%m.%Y')}: Tagestyp {day_type}, Spalte {col_idx}, Rohwert {daily_raw_sum:.2f}")
                                else:
                                    raw_daily_values[date] = 0
                                    print(f"Datum {date.strftime('%d.%m.%Y')}: Keine Spalte gefunden für Monat {month}, Tagestyp {day_type}")
                            else:
                                raw_daily_values[date] = 0
                                
            except Exception as e:
                print(f"Fehler beim Laden der Rohwerte aus Excel: {e}")
                raw_daily_values = {}
        
        # Berechne Monats- und Jahresstatistiken
        monthly_consumption = consumption_series.resample('ME').sum()
        annual_total = consumption_series.sum()
        
        for date, daily_kwh in daily_consumption.items():
            # Finde den Tagestyp für dieses Datum
            day_of_year = date.timetuple().tm_yday
            if day_of_year <= len(weekday_df):
                day_row = weekday_df.iloc[day_of_year - 1]
                day_type = day_row[f'{bundesland_code}_Tagestyp']
                weekday_name = day_row['Wochentag_Deutsch']
            else:
                day_type = 'WT'
                weekday_name = date.strftime('%A')
            
            # Berechne verschiedene Kennzahlen für den Tag
            daily_peak = consumption_series[date.strftime('%Y-%m-%d')].max() if len(consumption_series[date.strftime('%Y-%m-%d')]) > 0 else 0
            daily_min = consumption_series[date.strftime('%Y-%m-%d')].min() if len(consumption_series[date.strftime('%Y-%m-%d')]) > 0 else 0
            
            # Zähle 15-Minuten-Intervalle mit Verbrauch
            intervals_with_consumption = len(consumption_series[date.strftime('%Y-%m-%d')][consumption_series[date.strftime('%Y-%m-%d')] > 0])
            daily_hours_with_consumption = intervals_with_consumption / 4.0  # 4 15-min-Intervalle pro Stunde
            
            # Berechne durchschnittliche Leistung
            avg_power = daily_kwh / 24.0  # kWh / 24h = kW
            
            # Berechne Verbrauch pro Person (falls mehrere Personen)
            consumption_per_person = daily_kwh / number_of_persons if number_of_persons > 0 else daily_kwh
            
            # Verwende die echten Rohwerte aus der Excel-Datei oder Fallback
            if date in raw_daily_values and raw_daily_values[date] > 0:
                # Verwende den direkt aus Excel extrahierten Rohwert
                raw_value_1mio = raw_daily_values[date]
            else:
                # Fallback: Berechne aus den tatsächlichen Verbrauchsdaten
                if annual_consumption_kwh > 0:
                    raw_value_1mio = daily_kwh * 1_000_000.0 / annual_consumption_kwh
                else:
                    raw_value_1mio = daily_kwh
            
            # 2. Normierter Wert (geteilt durch 1.000.000)
            normalized_value = raw_value_1mio / 1_000_000.0
            
            # 3. Skalierter Wert für einen Haushalt (sollte gleich daily_kwh sein)
            # annual_consumption_kwh ist bereits der Gesamtverbrauch für alle Haushalte
            # Daher teilen wir durch number_of_persons um den Verbrauch pro Haushalt zu erhalten
            annual_consumption_per_household = annual_consumption_kwh / number_of_persons if number_of_persons > 0 else annual_consumption_kwh
            household_value = normalized_value * annual_consumption_per_household
            
            # 4. Skalierter Wert für alle Haushalte
            total_value = household_value * number_of_persons
            
            breakdown_data.append({
                'Datum': date.strftime('%d.%m.%Y'),
                'Wochentag': weekday_name,
                'Tagestyp': day_type,
                'Tagesverbrauch_kWh': daily_kwh,
                'Spitzenleistung_kW': daily_peak,
                'Minimalleistung_kW': daily_min,
                'Durchschnittsleistung_kW': avg_power,
                'Stunden_mit_Verbrauch': daily_hours_with_consumption,
                'Verbrauch_pro_Person_kWh': consumption_per_person,
                'Rohwert_1Mio_kWh': raw_value_1mio,
                'Normiert_1Mio': normalized_value,
                'Alle_Haushalte_kWh': total_value,
                'Anteil_am_Jahresverbrauch_%': (daily_kwh / annual_total * 100) if annual_total > 0 else 0
            })
        
        # Erstelle Monatsaufschlüsselung
        monthly_breakdown = []
        for date, monthly_kwh in monthly_consumption.items():
            monthly_peak = consumption_series[date.strftime('%Y-%m')].max() if len(consumption_series[date.strftime('%Y-%m')]) > 0 else 0
            monthly_min = consumption_series[date.strftime('%Y-%m')].min() if len(consumption_series[date.strftime('%Y-%m')]) > 0 else 0
            monthly_avg_daily = monthly_kwh / len(daily_consumption[date.strftime('%Y-%m')]) if len(daily_consumption[date.strftime('%Y-%m')]) > 0 else 0
            
            monthly_breakdown.append({
                'Monat': date.strftime('%B %Y'),
                'Monatsverbrauch_kWh': monthly_kwh,
                'Spitzenleistung_kW': monthly_peak,
                'Minimalleistung_kW': monthly_min,
                'Durchschnittlicher_Tagesverbrauch_kWh': monthly_avg_daily,
                'Verbrauch_pro_Person_kWh': monthly_kwh / number_of_persons if number_of_persons > 0 else monthly_kwh,
                'Anteil_am_Jahresverbrauch_%': (monthly_kwh / annual_total * 100) if annual_total > 0 else 0
            })
        
        return {
            'daily_breakdown': pd.DataFrame(breakdown_data),
            'monthly_breakdown': pd.DataFrame(monthly_breakdown),
            'annual_stats': {
                'total_energy_kwh': annual_total,
                'avg_daily_energy_kwh': annual_total / len(daily_consumption),
                'peak_power_kw': consumption_series.max(),
                'min_power_kw': consumption_series.min(),
                'avg_power_kw': consumption_series.mean(),
                'consumption_per_person_kwh': annual_total / number_of_persons if number_of_persons > 0 else annual_total,
                'days_with_consumption': len(daily_consumption[daily_consumption > 0])
            }
        }
        
    except Exception as e:
        print(f"Fehler bei der Verbrauchsaufschlüsselung: {e}")
        return {
            'daily_breakdown': pd.DataFrame(),
            'monthly_breakdown': pd.DataFrame(),
            'annual_stats': {}
        }

def create_pv_generation_breakdown(pv_generation_series, pv_system_size_kwp, selected_year=2024):
    """
    Erstellt eine detaillierte Aufschlüsselung der PV-Erzeugungsdaten
    """
    try:
        if pv_generation_series is None or len(pv_generation_series) == 0:
            return pd.DataFrame()
        
        breakdown_data = []
        
        # Gruppiere nach Tagen
        daily_pv = pv_generation_series.resample('D').sum()
        
        # Berechne Monats- und Jahresstatistiken
        monthly_pv = pv_generation_series.resample('ME').sum()
        annual_total = pv_generation_series.sum()
        
        # Erstelle tägliche Aufschlüsselung
        for date, daily_kwh in daily_pv.items():
            # Berechne verschiedene Kennzahlen
            # Verwende die ursprüngliche Spitzenleistung, falls verfügbar (für stündliche Daten)
            if hasattr(pv_generation_series, 'attrs') and 'original_peak_power' in pv_generation_series.attrs:
                # Für stündliche Daten: Verwende die ursprüngliche Spitzenleistung
                # Die interpolierten 15-Min-Werte sind durch 4 geteilt worden
                daily_peak = pv_generation_series[date.strftime('%Y-%m-%d')].max() * 4.0 if len(pv_generation_series[date.strftime('%Y-%m-%d')]) > 0 else 0
            else:
                # Fallback: Verwende die aktuellen Werte
                daily_peak = pv_generation_series[date.strftime('%Y-%m-%d')].max() if len(pv_generation_series[date.strftime('%Y-%m-%d')]) > 0 else 0
            
            # Zähle 15-Minuten-Intervalle mit Erzeugung und konvertiere zu Stunden
            intervals_with_generation = len(pv_generation_series[date.strftime('%Y-%m-%d')][pv_generation_series[date.strftime('%Y-%m-%d')] > 0])
            daily_hours_with_generation = intervals_with_generation / 4.0  # 4 15-min-Intervalle pro Stunde
            
            # Berechne spezifischen Ertrag (kWh/kWp)
            specific_yield = daily_kwh / pv_system_size_kwp if pv_system_size_kwp > 0 else 0
            
            # Berechne durchschnittliche Leistung
            avg_power = daily_kwh / 24.0  # kWh / 24h = kW
            
            breakdown_data.append({
                'Datum': date.strftime('%d.%m.%Y'),
                'Wochentag': date.strftime('%A'),
                'Tagesertrag_kWh': daily_kwh,
                'Spitzenleistung_kW': daily_peak,
                'Durchschnittsleistung_kW': avg_power,
                'Stunden_mit_Erzeugung': daily_hours_with_generation,
                'Spezifischer_Ertrag_kWh_kWp': specific_yield,
                'Anteil_am_Jahresertrag_%': (daily_kwh / annual_total * 100) if annual_total > 0 else 0
            })
        
        # Erstelle Monatsaufschlüsselung
        monthly_breakdown = []
        for date, monthly_kwh in monthly_pv.items():
            # Verwende die ursprüngliche Spitzenleistung, falls verfügbar (für stündliche Daten)
            if hasattr(pv_generation_series, 'attrs') and 'original_peak_power' in pv_generation_series.attrs:
                # Für stündliche Daten: Verwende die ursprüngliche Spitzenleistung
                # Die interpolierten 15-Min-Werte sind durch 4 geteilt worden
                monthly_peak = pv_generation_series[date.strftime('%Y-%m')].max() * 4.0 if len(pv_generation_series[date.strftime('%Y-%m')]) > 0 else 0
            else:
                # Fallback: Verwende die aktuellen Werte
                monthly_peak = pv_generation_series[date.strftime('%Y-%m')].max() if len(pv_generation_series[date.strftime('%Y-%m')]) > 0 else 0
            
            monthly_avg_daily = monthly_kwh / len(daily_pv[date.strftime('%Y-%m')]) if len(daily_pv[date.strftime('%Y-%m')]) > 0 else 0
            
            monthly_breakdown.append({
                'Monat': date.strftime('%B %Y'),
                'Monatsertrag_kWh': monthly_kwh,
                'Spitzenleistung_kW': monthly_peak,
                'Durchschnittlicher_Tagesertrag_kWh': monthly_avg_daily,
                'Spezifischer_Ertrag_kWh_kWp': monthly_kwh / pv_system_size_kwp if pv_system_size_kwp > 0 else 0,
                'Anteil_am_Jahresertrag_%': (monthly_kwh / annual_total * 100) if annual_total > 0 else 0
            })
        
        return {
            'daily_breakdown': pd.DataFrame(breakdown_data),
            'monthly_breakdown': pd.DataFrame(monthly_breakdown),
            'annual_stats': {
                'total_energy_kwh': annual_total,
                'specific_yield_kwh_kwp': annual_total / pv_system_size_kwp if pv_system_size_kwp > 0 else 0,
                'peak_power_kw': pv_generation_series.max() * 4.0 if hasattr(pv_generation_series, 'attrs') and 'original_peak_power' in pv_generation_series.attrs else pv_generation_series.max(),
                'avg_daily_energy_kwh': annual_total / len(daily_pv),
                'days_with_generation': len(daily_pv[daily_pv > 0])
            }
        }
        
    except Exception as e:
        print(f"Fehler bei der PV-Erzeugungsaufschlüsselung: {e}")
        return {
            'daily_breakdown': pd.DataFrame(),
            'monthly_breakdown': pd.DataFrame(),
            'annual_stats': {}
        }

def create_amortisationszeit_sheet(wb, optimization_results, optimal_sim_result, cost_comparison_details, current_settings, battery_cost_curve=None, results_summary=None):
    """
    Erstellt ein detailliertes Sheet mit der Amortisationszeit-Berechnung Jahr für Jahr
    """
    try:
        ws_amort = wb.create_sheet("Amortisationszeit")
        
        # Überschrift
        ws_amort['A1'] = "AMORTISATIONSZEIT - DETAILLIERTE BERECHNUNG"
        ws_amort['A1'].font = Font(size=16, bold=True, color="FF6B35")
        ws_amort.merge_cells('A1:G1')
        ws_amort['A1'].alignment = Alignment(horizontal='center')
        
        # Farben für das Styling
        orange_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
        light_orange_fill = PatternFill(start_color="FFB366", end_color="FFB366", fill_type="solid")
        black_font = Font(color="000000", bold=True)
        white_font = Font(color="FFFFFF", bold=True)
        header_font = Font(color="000000", bold=True, size=12)
        
        # Hole die optimale Lösung
        if not optimization_results:
            ws_amort['A3'] = "Keine Optimierungsergebnisse verfügbar"
            return
        
        df_results = pd.DataFrame(optimization_results)
        optimization_criterion = current_settings.get('optimization_criterion', 'Deckungsbeitrag III gesamt (Barwert)')
        
        if optimization_criterion == "Deckungsbeitrag III gesamt (Nominal)":
            best_idx = df_results['total_db3_nominal'].idxmax()
        else:
            best_idx = df_results['total_db3_present_value'].idxmax()
        
        best_option = df_results.loc[best_idx]
        
        # Hole Parameter
        project_lifetime = int(current_settings.get('project_lifetime_years', 15) or 15)
        annual_capacity_loss = float(current_settings.get('annual_capacity_loss_percent', 2.0) or 2.0)
        
        # Hole Investitionskosten und jährliche Ersparnisse
        if results_summary:
            investment_cost = results_summary.get('investment_cost', best_option.get('investment_cost', 0))
            annual_savings = results_summary.get('annual_savings', best_option.get('annual_savings', 0))
            payback_period = results_summary.get('payback_period_years', best_option.get('payback_period_years', 999))
        else:
            investment_cost = best_option.get('investment_cost', 0)
            annual_savings = best_option.get('annual_savings', 0)
            payback_period = best_option.get('payback_period_years', 999)
        
        # Übersicht
        row = 3
        ws_amort[f'A{row}'] = "ÜBERSICHT"
        ws_amort[f'A{row}'].font = header_font
        ws_amort[f'A{row}'].fill = orange_fill
        ws_amort[f'A{row}'].font = white_font
        ws_amort.merge_cells(f'A{row}:G{row}')
        row += 1
        
        overview_data = [
            ("Investitionskosten", f"{investment_cost:.2f} €"),
            ("Jährliche Ersparnis (Jahr 1)", f"{annual_savings:.2f} €"),
            ("Jährlicher Kapazitätsverlust", f"{annual_capacity_loss:.2f} %"),
            ("Projektlaufzeit", f"{project_lifetime} Jahre"),
            ("Amortisationszeit", f"{payback_period:.2f} Jahre" if payback_period < 999 else "Nicht erreicht")
        ]
        
        for label, value in overview_data:
            ws_amort[f'A{row}'] = label
            ws_amort[f'B{row}'] = value
            ws_amort[f'A{row}'].font = black_font
            ws_amort[f'B{row}'].font = Font(bold=True)
            row += 1
        
        row += 2
        
        # Detaillierte Berechnung Jahr für Jahr
        ws_amort[f'A{row}'] = "DETAILLIERTE BERECHNUNG JAHR FÜR JAHR"
        ws_amort[f'A{row}'].font = header_font
        ws_amort[f'A{row}'].fill = orange_fill
        ws_amort[f'A{row}'].font = white_font
        ws_amort.merge_cells(f'A{row}:G{row}')
        row += 1
        
        # Tabellen-Header
        headers = ["Jahr", "Degradationsfaktor", "Jährliche Ersparnis (degradiert)", "Kumulierte Ersparnis", "Verbleibende Investition", "Status", "Hinweis"]
        for col, header in enumerate(headers, 1):
            cell = ws_amort.cell(row=row, column=col)
            cell.value = header
            cell.font = white_font
            cell.fill = light_orange_fill
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        # Berechne Jahr für Jahr
        cumulative_savings = 0
        break_even_found = False
        
        for year in range(1, project_lifetime + 1):
            # Degradation der Ersparnisse
            degradation_factor = (1.0 - annual_capacity_loss / 100.0) ** (year - 1)
            year_savings = annual_savings * degradation_factor
            cumulative_savings += year_savings
            remaining_investment = investment_cost - cumulative_savings
            
            # Status bestimmen
            if cumulative_savings >= investment_cost and not break_even_found:
                status = "✓ Amortisiert"
                break_even_found = True
                if year == 1:
                    hint = f"Amortisation bereits im ersten Jahr erreicht"
                else:
                    previous_cumulative = cumulative_savings - year_savings
                    remaining_before = investment_cost - previous_cumulative
                    fraction_of_year = remaining_before / year_savings if year_savings > 0 else 0
                    actual_payback = (year - 1) + fraction_of_year
                    hint = f"Amortisation nach {actual_payback:.2f} Jahren erreicht"
            elif cumulative_savings >= investment_cost:
                status = "Weiterhin amortisiert"
                hint = ""
            else:
                status = "Noch nicht amortisiert"
                hint = f"Verbleibend: {remaining_investment:.2f} €"
            
            # Schreibe Daten
            ws_amort[f'A{row}'] = year
            ws_amort[f'B{row}'] = f"{degradation_factor:.4f}"
            ws_amort[f'C{row}'] = f"{year_savings:.2f} €"
            ws_amort[f'D{row}'] = f"{cumulative_savings:.2f} €"
            ws_amort[f'E{row}'] = f"{remaining_investment:.2f} €" if remaining_investment > 0 else "0.00 €"
            ws_amort[f'F{row}'] = status
            ws_amort[f'G{row}'] = hint
            
            # Formatierung
            ws_amort[f'A{row}'].alignment = Alignment(horizontal='center')
            ws_amort[f'B{row}'].number_format = '0.0000'
            ws_amort[f'C{row}'].number_format = '#,##0.00 "€"'
            ws_amort[f'D{row}'].number_format = '#,##0.00 "€"'
            ws_amort[f'E{row}'].number_format = '#,##0.00 "€"'
            
            # Hervorhebung bei Amortisation
            if break_even_found and cumulative_savings >= investment_cost and year <= payback_period + 1:
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                    ws_amort[f'{col}{row}'].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    ws_amort[f'{col}{row}'].font = Font(bold=True)
            
            row += 1
        
        # Zusammenfassung am Ende
        row += 1
        ws_amort[f'A{row}'] = "ZUSAMMENFASSUNG"
        ws_amort[f'A{row}'].font = header_font
        ws_amort[f'A{row}'].fill = orange_fill
        ws_amort[f'A{row}'].font = white_font
        ws_amort.merge_cells(f'A{row}:G{row}')
        row += 1
        
        if payback_period < 999:
            ws_amort[f'A{row}'] = f"Die Amortisationszeit beträgt {payback_period:.2f} Jahre."
            ws_amort[f'A{row}'].font = Font(bold=True, size=11)
            ws_amort.merge_cells(f'A{row}:G{row}')
            row += 1
            
            ws_amort[f'A{row}'] = f"Das bedeutet, dass nach {payback_period:.2f} Jahren die kumulierten Ersparnisse die Anfangsinvestition decken."
            ws_amort[f'A{row}'].font = Font(size=10)
            ws_amort.merge_cells(f'A{row}:G{row}')
        else:
            ws_amort[f'A{row}'] = "Die Amortisationszeit wird innerhalb der Projektlaufzeit nicht erreicht."
            ws_amort[f'A{row}'].font = Font(bold=True, size=11, color="FF0000")
            ws_amort.merge_cells(f'A{row}:G{row}')
            row += 1
            
            ws_amort[f'A{row}'] = f"Nach {project_lifetime} Jahren beträgt die kumulierte Ersparnis {cumulative_savings:.2f} €, die Investitionskosten betragen {investment_cost:.2f} €."
            ws_amort[f'A{row}'].font = Font(size=10)
            ws_amort.merge_cells(f'A{row}:G{row}')
        
        row += 1
        ws_amort[f'A{row}'] = "Hinweis: Die jährliche Ersparnis wird durch die Kapazitätsalterung der Batterie reduziert."
        ws_amort[f'A{row}'].font = Font(size=9, italic=True, color="666666")
        ws_amort.merge_cells(f'A{row}:G{row}')
        row += 1
        
        ws_amort[f'A{row}'] = f"Degradationsfaktor = (1 - {annual_capacity_loss}%)^(Jahr-1)"
        ws_amort[f'A{row}'].font = Font(size=9, italic=True, color="666666")
        ws_amort.merge_cells(f'A{row}:G{row}')
        
        # Spaltenbreiten anpassen
        ws_amort.column_dimensions['A'].width = 8
        ws_amort.column_dimensions['B'].width = 18
        ws_amort.column_dimensions['C'].width = 25
        ws_amort.column_dimensions['D'].width = 20
        ws_amort.column_dimensions['E'].width = 20
        ws_amort.column_dimensions['F'].width = 20
        ws_amort.column_dimensions['G'].width = 35
        
    except Exception as e:
        st.error(f"Fehler beim Erstellen des Amortisationszeit-Sheets: {str(e)}")
        import traceback
        st.error(f"Traceback: {traceback.format_exc()}")

def create_deckungsbeitragsrechnung_sheet(wb, optimization_results, optimal_sim_result, cost_comparison_details, current_settings, battery_cost_curve=None, results_summary=None):
    """
    Erstellt ein detailliertes Sheet mit mehrstufiger Deckungsbeitragsrechnung
    basierend auf der ursprünglichen Excel-Logik für jede Speichergröße
    """
    try:
        ws_db = wb.create_sheet("Deckungsbeitragsrechnung")
        
        # Hole relevante Daten
        df_results = pd.DataFrame(optimization_results)
        
        # Dynamische Optimierungslogik basierend auf ausgewähltem Kriterium
        optimization_criterion = current_settings.get('optimization_criterion', 'Deckungsbeitrag III gesamt (Barwert)')
        
        # Nur DB Barwert und DB Nominal unterstützen
        if optimization_criterion == "Deckungsbeitrag III gesamt (Nominal)":
            best_idx = df_results['total_db3_nominal'].idxmax()
            criterion_name = "DB III gesamt (Nominal)"
        else:  # Default: Deckungsbeitrag III gesamt (Barwert)
            best_idx = df_results['total_db3_present_value'].idxmax()
            criterion_name = "DB III gesamt (Barwert)"
        
        best_option = df_results.loc[best_idx]
        battery_capacity = best_option['battery_capacity_kwh']
        
        # Titel
        ws_db['A1'] = f"DECKUNGSBEITRAGSRECHNUNG - BATTERIESPEICHER ({criterion_name})"
        ws_db['A1'].font = Font(size=16, bold=True, color="FF6B35")
        ws_db.merge_cells('A1:H1')
        ws_db['A1'].alignment = Alignment(horizontal='center')
        
        ws_db['A2'] = f"Optimiert nach: {optimization_criterion} - Kapazität: {battery_capacity:.1f} kWh"
        ws_db['A2'].font = Font(size=10, italic=True)
        ws_db.merge_cells('A2:H2')
        
        # Styling
        header_font = Font(size=14, bold=True, color="FFFFFF")
        orange_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
        black_font = Font(color="000000", bold=True)
        
        # === 1. ÜBERSICHT DER OPTIMALEN LÖSUNG ===
        row = 4
        
        ws_db[f'A{row}'] = "Optimale Lösung - Übersicht"
        ws_db[f'A{row}'].font = header_font
        ws_db[f'A{row}'].fill = orange_fill
        ws_db.merge_cells(f'A{row}:H{row}')
        row += 1
        
        # Benutzerfreundliche Anzeige der Amortisationszeit (999 Jahre = "Nicht erreicht")
        payback_display = "Nicht erreicht" if best_option['payback_period_years'] >= 999 else f"{best_option['payback_period_years']:.1f} Jahre"
        
        # Verwende results_summary wenn verfügbar (konsistent mit UI)
        if results_summary:
            investment_cost_db = results_summary.get('investment_cost', best_option['investment_cost'])
            roi_percentage_db = results_summary.get('roi_percentage')
            if roi_percentage_db is None:
                # Fallback-Berechnung
                annual_savings_db = results_summary.get('annual_savings', best_option.get('annual_savings', 0))
                project_lifetime_db = results_summary.get('project_lifetime_years', current_settings.get('project_lifetime_years', 20))
                roi_percentage_db = (annual_savings_db * project_lifetime_db / investment_cost_db - 1) * 100 if investment_cost_db else 0
            payback_period_db = results_summary.get('payback_period_years', best_option.get('payback_period_years', 999))
            payback_display_db = "Nicht erreicht" if payback_period_db >= 999 else f"{payback_period_db:.1f} Jahre"
            optimal_value_db = results_summary.get('criterion_value', best_option.get('total_db3_present_value', best_option.get('npv', 0)))
        else:
            investment_cost_db = best_option['investment_cost']
            roi_percentage_db = best_option.get('roi_percentage', 0)
            payback_display_db = payback_display
            optimal_value_db = best_option.get('total_db3_present_value', best_option.get('npv', 0))
        
        # Übersichtsdaten aus der optimalen Lösung
        overview_data = [
            ("Batteriekapazität", f"{battery_capacity:.1f} kWh"),
            ("Investitionskosten", f"{investment_cost_db:.0f} €"),
            ("Optimierungskriterium", optimization_criterion),
            ("Optimaler Wert", f"{optimal_value_db:.2f}"),
            ("Amortisationszeit", payback_display_db),
            ("ROI", f"{roi_percentage_db:.1f}%")
        ]
        
        for label, value in overview_data:
            ws_db[f'A{row}'] = label
            ws_db[f'B{row}'] = value
            ws_db[f'A{row}'].font = black_font
            ws_db[f'B{row}'].font = Font(bold=True)
            row += 1
        
        row += 1
        
        # === 2. JAHRESWERTE TABELLE (wie in ursprünglicher Excel-Vorlage) ===
        ws_db[f'A{row}'] = "Jahreswerte - nominal & Barwert"
        ws_db[f'A{row}'].font = header_font
        ws_db[f'A{row}'].fill = orange_fill
        ws_db.merge_cells(f'A{row}:O{row}')
        row += 1
        
        # Spalten-Header
        headers = [
            "Jahr", "Umsatz_nominal_€", "Kosten_vor_DB1_nominal_€", "DB1_nominal_€", 
            "Abschreibung_nominal_€", "DB2_nominal_€", "Zinsen_nominal_€", "DB3_nominal_€",
            "Umsatz_Barwert_€", "Kosten_vor_DB1_Barwert_€", "DB1_Barwert_€", 
            "Abschreibung_Barwert_€", "DB2_Barwert_€", "Zinsen_Barwert_€", "DB3_Barwert_€"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws_db.cell(row=row, column=col)
            cell.value = header
            cell.font = black_font
            cell.fill = PatternFill(start_color="FFB366", end_color="FFB366", fill_type="solid")
        
        row += 1
        
        # Berechne Werte für jedes Jahr mit korrekter Methodik
        project_lifetime = int(current_settings.get('project_lifetime_years', 15) or 15)
        discount_rate = float(current_settings.get('discount_rate', 0.02) or 0.0)
        degradation_rate_percent = float(current_settings.get('annual_capacity_loss_percent', 1.0) or 0.0)
        degradation_rate = max(degradation_rate_percent, 0.0) / 100.0
        
        # Basis-Werte aus der optimalen Lösung
        base_revenue = best_option['annual_revenue']
        base_costs = best_option['annual_variable_costs']
        base_db1 = best_option['contribution_margin_1']
        base_depreciation = best_option['annual_depreciation']
        base_db2 = best_option['contribution_margin_2']
        base_interest = best_option['annual_interest']
        base_db3 = best_option['contribution_margin_3']
        
        # Summen für Gesamtübersicht
        total_nominal_revenue = 0
        total_nominal_costs = 0
        total_nominal_db1 = 0
        total_nominal_depreciation = 0
        total_nominal_db2 = 0
        total_nominal_interest = 0
        total_nominal_db3 = 0
        
        total_pv_revenue = 0
        total_pv_costs = 0
        total_pv_db1 = 0
        total_pv_depreciation = 0
        total_pv_db2 = 0
        total_pv_interest = 0
        total_pv_db3 = 0
        
        for year in range(1, project_lifetime + 1):
            # Degradationsfaktor für dieses Jahr
            degradation_factor = (1 - degradation_rate) ** (year - 1)
            
            # Nominal-Werte (mit Degradation)
            nominal_revenue = base_revenue * degradation_factor
            nominal_costs = base_costs * degradation_factor
            nominal_db1 = base_db1 * degradation_factor
            nominal_depreciation = base_depreciation  # Konstante Abschreibung
            nominal_db2 = base_db2 * degradation_factor
            nominal_interest = base_interest  # Konstante Zinsen
            nominal_db3 = base_db3 * degradation_factor
            
            # Barwert-Werte (mit Degradation und Abzinsung)
            discount_factor = 1 / ((1 + discount_rate) ** year) if (1 + discount_rate) != 0 else 1.0
            pv_revenue = nominal_revenue * discount_factor
            pv_costs = nominal_costs * discount_factor
            pv_db1 = nominal_db1 * discount_factor
            pv_depreciation = nominal_depreciation * discount_factor
            pv_db2 = nominal_db2 * discount_factor
            pv_interest = nominal_interest * discount_factor
            pv_db3 = nominal_db3 * discount_factor
            
            # Schreibe Zeile
            values = [
                year, nominal_revenue, nominal_costs, nominal_db1,
                nominal_depreciation, nominal_db2, nominal_interest, nominal_db3,
                pv_revenue, pv_costs, pv_db1, pv_depreciation, pv_db2, pv_interest, pv_db3
            ]
            
            for col, value in enumerate(values, 1):
                cell = ws_db.cell(row=row, column=col)
                cell.value = value
                if col > 1:  # Alle Spalten außer Jahr
                    cell.number_format = '#,##0.00'
            
            # Summen für Gesamtübersicht
            total_nominal_revenue += nominal_revenue
            total_nominal_costs += nominal_costs
            total_nominal_db1 += nominal_db1
            total_nominal_depreciation += nominal_depreciation
            total_nominal_db2 += nominal_db2
            total_nominal_interest += nominal_interest
            total_nominal_db3 += nominal_db3
            
            total_pv_revenue += pv_revenue
            total_pv_costs += pv_costs
            total_pv_db1 += pv_db1
            total_pv_depreciation += pv_depreciation
            total_pv_db2 += pv_db2
            total_pv_interest += pv_interest
            total_pv_db3 += pv_db3
            
            row += 1
        
        # Summen-Zeile
        ws_db[f'A{row}'] = "Summe"
        ws_db[f'A{row}'].font = black_font
        ws_db[f'A{row}'].fill = PatternFill(start_color="FFB366", end_color="FFB366", fill_type="solid")
        
        sum_values = [
            total_nominal_revenue, total_nominal_costs, total_nominal_db1,
            total_nominal_depreciation, total_nominal_db2, total_nominal_interest, total_nominal_db3,
            total_pv_revenue, total_pv_costs, total_pv_db1, total_pv_depreciation, total_pv_db2, total_pv_interest, total_pv_db3
        ]
        
        for col, value in enumerate(sum_values, 2):
            cell = ws_db.cell(row=row, column=col)
            cell.value = value
            cell.number_format = '#,##0.00'
            cell.font = black_font
            cell.fill = PatternFill(start_color="FFB366", end_color="FFB366", fill_type="solid")
        
        row += 3
        
        # === 3. ÜBERSICHT (wie in ursprünglicher Excel-Vorlage) ===
        ws_db[f'A{row}'] = f"Übersicht - Deckungsbeiträge (Summe über {project_lifetime} Jahre)"
        ws_db[f'A{row}'].font = header_font
        ws_db[f'A{row}'].fill = orange_fill
        ws_db.merge_cells(f'A{row}:C{row}')
        row += 1
        
        # Header für nominale Übersicht
        ws_db[f'A{row}'] = "Position"
        ws_db[f'B{row}'] = "Summe_15J_nominal_€"
        ws_db[f'C{row}'] = "Quote_von_Umsatz_%"
        for col in ['A', 'B', 'C']:
            ws_db[f'{col}{row}'].font = black_font
            ws_db[f'{col}{row}'].fill = PatternFill(start_color="FFB366", end_color="FFB366", fill_type="solid")
        
        # Header für Barwert-Übersicht (rechts daneben)
        ws_db[f'E{row}'] = "Position"
        ws_db[f'F{row}'] = "Summe_15J_Barwert_€"
        ws_db[f'G{row}'] = "Quote_von_Umsatz_%"
        for col in ['E', 'F', 'G']:
            ws_db[f'{col}{row}'].font = black_font
            ws_db[f'{col}{row}'].fill = PatternFill(start_color="FFB366", end_color="FFB366", fill_type="solid")
        
        row += 1
        
        # Nominal-Werte
        nominal_data = [
            ("Gesamtumsatz", total_nominal_revenue, 100.0),
            ("Kosten vor DB I", total_nominal_costs, (total_nominal_costs / total_nominal_revenue * 100) if total_nominal_revenue > 0 else 0),
            ("DB I", total_nominal_db1, (total_nominal_db1 / total_nominal_revenue * 100) if total_nominal_revenue > 0 else 0),
            ("Abschreibung", total_nominal_depreciation, (total_nominal_depreciation / total_nominal_revenue * 100) if total_nominal_revenue > 0 else 0),
            ("DB II", total_nominal_db2, (total_nominal_db2 / total_nominal_revenue * 100) if total_nominal_revenue > 0 else 0),
            ("Zinsen", total_nominal_interest, (total_nominal_interest / total_nominal_revenue * 100) if total_nominal_revenue > 0 else 0),
            ("DB III", total_nominal_db3, (total_nominal_db3 / total_nominal_revenue * 100) if total_nominal_revenue > 0 else 0)
        ]
        
        for idx, (position, value, quote) in enumerate(nominal_data):
            current_row = row + idx
            ws_db[f'A{current_row}'] = position
            ws_db[f'B{current_row}'] = value
            ws_db[f'C{current_row}'] = quote
            ws_db[f'B{current_row}'].number_format = '#,##0.00'
            ws_db[f'C{current_row}'].number_format = '0.0'
        
        # Barwert-Werte (dynamisch berechnet) – in separater Tabelle
        pv_data = [
            ("Gesamtumsatz", total_pv_revenue, 100.0),
            ("Kosten vor DB I", total_pv_costs, (total_pv_costs / total_pv_revenue * 100) if total_pv_revenue > 0 else 0),
            ("DB I", total_pv_db1, (total_pv_db1 / total_pv_revenue * 100) if total_pv_revenue > 0 else 0),
            ("Abschreibung", total_pv_depreciation, (total_pv_depreciation / total_pv_revenue * 100) if total_pv_revenue > 0 else 0),
            ("DB II", total_pv_db2, (total_pv_db2 / total_pv_revenue * 100) if total_pv_revenue > 0 else 0),
            ("Zinsen", total_pv_interest, (total_pv_interest / total_pv_revenue * 100) if total_pv_revenue > 0 else 0),
            ("DB III", total_pv_db3, (total_pv_db3 / total_pv_revenue * 100) if total_pv_revenue > 0 else 0)
        ]
        
        for idx, (position, value, quote) in enumerate(pv_data):
            current_row = row + idx
            ws_db[f'E{current_row}'] = position
            ws_db[f'F{current_row}'] = value
            ws_db[f'G{current_row}'] = quote
            ws_db[f'F{current_row}'].number_format = '#,##0.00'
            ws_db[f'G{current_row}'].number_format = '0.0'
        
        row += len(nominal_data) + 2
        
        # === 4. GESAMTWERTE ÜBER PROJEKTLAUFZEIT ===
        ws_db[f'A{row}'] = f"Gesamtwerte über Projektlaufzeit ({project_lifetime} Jahre)"
        ws_db[f'A{row}'].font = header_font
        ws_db[f'A{row}'].fill = orange_fill
        ws_db.merge_cells(f'A{row}:H{row}')
        row += 1
        
        # Gesamtwerte
        total_data = [
            ("DB III gesamt (Nominal)", f"{best_option['total_db3_nominal']:.0f} €"),
            ("DB III gesamt (Barwert)", f"{best_option['total_db3_present_value']:.0f} €"),
        ]
        
        for label, value in total_data:
            ws_db[f'A{row}'] = label
            ws_db[f'B{row}'] = value
            ws_db[f'A{row}'].font = black_font
            ws_db[f'B{row}'].font = Font(bold=True)
            row += 1
        
        row += 1
        
        # === 4. DETAILLIERTE ERLÄUTERUNGEN ===
        ws_db[f'A{row}'] = "Erläuterungen - Herkunft der Werte"
        ws_db[f'A{row}'].font = header_font
        ws_db[f'A{row}'].fill = orange_fill
        ws_db.merge_cells(f'A{row}:C{row}')
        row += 1
        
        # Spalten-Header für Erklärungen
        ws_db[f'A{row}'] = "Begriff"
        ws_db[f'B{row}'] = "Erklärung/Formel"
        ws_db[f'C{row}'] = "Berechnung mit Werten"
        for col in ['A', 'B', 'C']:
            ws_db[f'{col}{row}'].font = Font(size=10, bold=True)
            ws_db[f'{col}{row}'].fill = PatternFill(start_color="FFB366", end_color="FFB366", fill_type="solid")
        row += 1
        
        # Hole die Parameter und tatsächlichen Werte für die Erklärungen
        price_grid = current_settings.get('price_grid_per_kwh', 0.30)
        price_feed_in = current_settings.get('price_feed_in_per_kwh', 0.08)
        discount_rate = current_settings.get('discount_rate', 0.02)
        interest_rate_db = current_settings.get('project_interest_rate_db', 0.03)
        
        # Berechne tatsächliche Werte für Beispielrechnungen
        
        try:
            # Nutze .get() für Series-Zugriff (funktioniert auch wenn Spalte fehlt)
            investment_cost = float(best_option.get('investment_cost', 0))
            annual_revenue = float(best_option.get('annual_revenue', best_option.get('annual_savings', 0)))
            annual_costs = float(best_option.get('annual_variable_costs', 0))
            db1_value = float(best_option.get('contribution_margin_1', 0))
            annual_depreciation = float(best_option.get('annual_depreciation', 0))
            db2_value = float(best_option.get('contribution_margin_2', 0))
            annual_interest = float(best_option.get('annual_interest', 0))
            db3_value = float(best_option.get('contribution_margin_3', 0))
            total_db3_nominal = float(best_option.get('total_db3_nominal', 0))
            total_db3_pv = float(best_option.get('total_db3_present_value', 0))
            db3_per_kwh = float(best_option.get('contribution_margin_3_per_kwh', 0))
            
            # Hol Ersparnis-Details
            savings_import = float(best_option.get('savings_from_reduced_import', 0))
            loss_export = float(best_option.get('loss_from_reduced_export', 0))
            
            # Falls nicht in best_option, versuche aus cost_comparison_details zu holen
            if savings_import == 0 and cost_comparison_details:
                savings_import = float(cost_comparison_details.get('savings_from_reduced_import', 0))
                loss_export = float(cost_comparison_details.get('loss_from_reduced_export', 0))
            
            # Wenn annual_revenue 0 ist, aber savings/loss vorhanden sind, berechne
            if annual_revenue == 0 and (savings_import > 0 or loss_export > 0):
                annual_revenue = savings_import - loss_export
            
            # Wenn DB-Werte 0 sind, berechne sie
            if db1_value == 0 and annual_revenue > 0:
                db1_value = annual_revenue - annual_costs
            
            if annual_depreciation == 0 and investment_cost > 0:
                annual_depreciation = investment_cost / project_lifetime
            
            if db2_value == 0 and db1_value > 0:
                db2_value = db1_value - annual_depreciation
            
            if annual_interest == 0 and investment_cost > 0:
                annual_interest = (investment_cost / 2) * interest_rate_db
            
            if db3_value == 0:
                db3_value = db2_value - annual_interest
            
        except Exception as e:
            print(f"[Deckungsbeitrag] FEHLER beim Laden der Werte für Erläuterungen: {e}")
            import traceback
            traceback.print_exc()
            # Fallback-Werte
            investment_cost = 0
            annual_revenue = 0
            annual_costs = 0
            db1_value = 0
            annual_depreciation = 0
            db2_value = 0
            annual_interest = 0
            db3_value = 0
            total_db3_nominal = 0
            total_db3_pv = 0
            db3_per_kwh = 0
            savings_import = 0
            loss_export = 0
        
        # Erweiterte Erklärungen: (Begriff, Erklärung, Konkrete Berechnung)
        explanations = [
            ("GRUNDLEGENDE BEGRIFFE:", "", ""),
            ("Umsatz (Erlöse)", "Jährliche Netto-Kostenreduktion", f"Ersparnis Netzbezug - Verlust Einspeisung"),
            ("  Ersparnis Netzbezug", f"Differenz × {price_grid:.2f} €/kWh", f"{savings_import:.2f} €"),
            ("  Verlust Einspeisung", f"Differenz × {price_feed_in:.2f} €/kWh", f"{loss_export:.2f} €"),
            ("  → Umsatz Jahr 1", "Ersparnis - Verlust", f"{savings_import:.2f} € - {loss_export:.2f} € = {annual_revenue:.2f} €"),
            ("", "", ""),
            ("Variable Kosten", "Keine variablen Kosten (nur Investition)", f"{annual_costs:.2f} €"),
            ("", "", ""),
            ("DECKUNGSBEITRÄGE:", "", ""),
            ("DB I (nominal)", "Umsatz - Variable Kosten", f"{annual_revenue:.2f} € - {annual_costs:.2f} € = {db1_value:.2f} €"),
            ("", "", ""),
            ("Abschreibung", f"Linear über {project_lifetime} Jahre", f"{investment_cost:.0f} € ÷ {project_lifetime} = {annual_depreciation:.2f} €/Jahr"),
            ("", "", ""),
            ("DB II (nominal)", "DB I - Abschreibung", f"{db1_value:.2f} € - {annual_depreciation:.2f} € = {db2_value:.2f} €"),
            ("", "", ""),
            ("Zinsen", f"Kalk. Zinsen {interest_rate_db*100:.1f}% auf ∅ Kapital", f"({investment_cost:.0f} € ÷ 2) × {interest_rate_db*100:.1f}% = {annual_interest:.2f} €/Jahr"),
            ("", "", ""),
            ("DB III (nominal)", "DB II - Zinsen", f"{db2_value:.2f} € - {annual_interest:.2f} € = {db3_value:.2f} €"),
            ("", "", ""),
            ("BARWERT-BERECHNUNG:", "", ""),
            ("Diskontierungssatz", f"{discount_rate*100:.1f}%", "Aus Eingabeparametern"),
            ("Barwertfaktor Jahr t", f"1 / (1 + {discount_rate:.3f})^t", f"Jahr 1: {1/(1+discount_rate):.4f}, Jahr 5: {1/((1+discount_rate)**5):.4f}"),
            ("Barwert Jahr t", "Nominalwert × Barwertfaktor", f"Jahr 1: {db3_value:.2f} € × {1/(1+discount_rate):.4f} = {db3_value * (1/(1+discount_rate)):.2f} €"),
            ("", "", ""),
            ("DEGRADATION:", "", ""),
            ("Kapazitätsverlust", f"{current_settings.get('annual_capacity_loss_percent', 1.0):.1f}% pro Jahr", ""),
            ("Degradationsfaktor", f"(1 - {degradation_rate:.4f})^(t-1)", f"Jahr 1: 1.0000, Jahr 5: {(1-degradation_rate)**4:.4f}"),
            ("Umsatz mit Degradation", "Basis-Umsatz × Degradationsfaktor", f"Jahr 1: {annual_revenue:.2f} €, Jahr 5: {annual_revenue * ((1-degradation_rate)**4):.2f} €"),
            ("", "", ""),
            ("GESAMTWERTE:", "", ""),
            ("DB III gesamt (Nominal)", f"Summe über {project_lifetime} Jahre", f"{total_db3_nominal:.2f} €"),
            ("DB III gesamt (Barwert)", f"Summe über {project_lifetime} Jahre (abgezinst)", f"{total_db3_pv:.2f} €"),
        ]
        
        for label, explanation, calculation in explanations:
            if label == "":  # Leerzeile
                row += 1
                continue
            
            # Setze Label in Spalte A
            cell_a = ws_db.cell(row=row, column=1)
            cell_a.value = label
            
            # Formatierung und Merging
            is_section_header = label.endswith(":") and not label.startswith("  ")
            
            if is_section_header:
                # Nur Sections-Header (GRUNDLEGENDE BEGRIFFE:, DECKUNGSBEITRÄGE:, etc.) werden gemerged
                cell_a.font = Font(size=11, bold=True, color="FF6B35")
                ws_db.merge_cells(f'A{row}:C{row}')
            else:
                # Setze Erklärung in Spalte B (explizit als Text, um #NAME? Fehler zu vermeiden)
                cell_b = ws_db.cell(row=row, column=2)
                cell_b.value = str(explanation) if explanation else ""
                cell_b.data_type = 's'  # 's' = string
                
                # Setze Berechnung in Spalte C
                cell_c = ws_db.cell(row=row, column=3)
                cell_c.value = str(calculation) if calculation else ""
                cell_c.data_type = 's'  # 's' = string
                
                if label.startswith("  "):
                    cell_a.font = Font(size=9, italic=True)
                    cell_b.font = Font(size=9, italic=True)
                    cell_c.font = Font(size=9, italic=True)
                else:
                    cell_a.font = Font(size=10)
                    cell_b.font = Font(size=10)
                    cell_c.font = Font(size=10, bold=True, color="008000")  # Berechnungen in Grün und Fett
            
            row += 1
        
        # Spaltenbreiten anpassen
        ws_db.column_dimensions['A'].width = 30
        ws_db.column_dimensions['B'].width = 50
        ws_db.column_dimensions['C'].width = 50
        for col in ['D', 'E', 'F', 'G', 'H']:
            ws_db.column_dimensions[col].width = 15
        
    except Exception as e:
        print(f"Fehler beim Erstellen der Deckungsbeitragsrechnung: {e}")
        import traceback
        traceback.print_exc()

# Erste create_comprehensive_excel Funktion wurde entfernt (war Duplikat und unvollständig)
# Die korrekte Funktion befindet sich weiter unten

# Erste create_comprehensive_excel Funktion wurde entfernt (war Duplikat und unvollständig)
# Die korrekte Funktion befindet sich weiter unten

# Erste create_comprehensive_excel Funktion wurde entfernt (war Duplikat und unvollständig)
# Die korrekte Funktion befindet sich weiter unten

# Erste create_comprehensive_excel Funktion wurde entfernt (war Duplikat und unvollständig)
# Die korrekte Funktion befindet sich weiter unten

# Erste create_comprehensive_excel Funktion wurde entfernt (war Duplikat und unvollständig)
# Die korrekte Funktion befindet sich weiter unten

# Erste create_comprehensive_excel Funktion wurde entfernt (war Duplikat und unvollständig)
# Die korrekte Funktion befindet sich weiter unten

# Erste create_comprehensive_excel Funktion wurde entfernt (war Duplikat und unvollständig)
# Die korrekte Funktion befindet sich weiter unten

# Erste create_comprehensive_excel Funktion wurde entfernt (war Duplikat und unvollständig)
# Die korrekte Funktion befindet sich weiter unten

def create_comprehensive_excel(optimization_results, optimal_sim_result, variable_tariff_result, 
                              current_settings, calculation_details, cost_comparison_details,
                              consumption_series=None, pv_generation_series=None, battery_cost_curve=None,
                              results_summary=None):
    """
    Erstellt eine umfassende Excel-Datei mit allen Berechnungen und Grafiken
    """
    if not optimization_results:
        st.error("❌ Keine Daten verfügbar")
        return None
    
    # Leere den Cache um sicherzustellen, dass alle Grafiken neu generiert werden
    clear_plot_cache()
    
    try:
        # Excel erstellen
        wb = Workbook()
        
        # Farben für das Styling
        orange_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
        light_orange_fill = PatternFill(start_color="FFB366", end_color="FFB366", fill_type="solid")
        gray_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)
        black_font = Font(color="000000", bold=True)
        header_font = Font(color="000000", bold=True, size=12)
        
        # 1. TITELSEITE
        ws_title = wb.active
        ws_title.title = "Titelseite"
        
        # Titel
        ws_title['A1'] = "BATTERIESPEICHER ANALYSE"
        ws_title['A1'].font = Font(size=20, bold=True, color="FF6B35")
        ws_title.merge_cells('A1:H1')
        ws_title['A1'].alignment = Alignment(horizontal='center')
        
        # Untertitel
        ws_title['A2'] = "Umfassender Wirtschaftlichkeitsbericht"
        ws_title['A2'].font = Font(size=14, italic=True)
        ws_title.merge_cells('A2:H2')
        ws_title['A2'].alignment = Alignment(horizontal='center')
        
        # Erstellungsdatum
        ws_title['A4'] = f"Erstellt am: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws_title['A4'].font = Font(size=10, italic=True)
        
        # Inhaltsverzeichnis
        ws_title['A6'] = "INHALTSVERZEICHNIS"
        ws_title['A6'].font = header_font
        ws_title['A6'].fill = orange_fill
        ws_title['A6'].font = white_font
        
        toc_items = [
            ["1.", "Zusammenfassung", "Hauptkennzahlen und Empfehlungen"],
            ["2.", "Eingabeparameter", "Alle verwendeten Simulationsparameter"],
            ["3.", "Optimierungsergebnisse", "Vollständige Optimierungsanalyse"],
            ["4.", "Energiebilanz", "Detaillierte Energieflüsse"],
            ["5.", "Verbrauchsaufschlüsselung", "Detaillierte Verbrauchsanalyse"],
            ["6.", "Tagesbilanz", "Tägliche Energiebilanz + 15-Minuten-Aufschlüsselung"],
            ["7.", "Kostenvergleich", "Wirtschaftlichkeitsanalyse"],
            ["8.", "Amortisationszeit", "Detaillierte Berechnung der Amortisationszeit Jahr für Jahr"],
            ["9.", "Deckungsbeitragsrechnung", "Wissenschaftliche mehrstufige Deckungsbeitragsrechnung"],
            ["10.", "PV-Erzeugungsdaten", "Detaillierte Aufschlüsselung der PV-Daten"],
            ["11.", "Kapazitätsalterung", "Entwicklung der Batteriekapazität über die Jahre"],
            ["12.", "Grafiken", "Visualisierungen der Ergebnisse"]
        ]
        
        for i, (num, title, desc) in enumerate(toc_items, 7):
            ws_title[f'A{i}'] = num
            ws_title[f'B{i}'] = title
            ws_title[f'C{i}'] = desc
            ws_title[f'A{i}'].font = Font(bold=True)
            ws_title[f'B{i}'].font = Font(bold=True)
        
        # 2. ZUSAMMENFASSUNG
        ws_summary = wb.create_sheet("Zusammenfassung")
        
        ws_summary['A1'] = "ZUSAMMENFASSUNG"
        ws_summary['A1'].font = Font(size=16, bold=True, color="FF6B35")
        ws_summary.merge_cells('A1:H1')
        ws_summary['A1'].alignment = Alignment(horizontal='center')
        
        # Optimale Batteriegröße
        if optimization_results:
            df_results = pd.DataFrame(optimization_results)
            # Verwende das gleiche Optimierungskriterium wie in der UI
            optimization_criterion = current_settings.get('optimization_criterion', 'Deckungsbeitrag III gesamt (Barwert)')
            # Nur DB Barwert und DB Nominal unterstützen
            if optimization_criterion == "Deckungsbeitrag III gesamt (Nominal)":
                best_idx = df_results['total_db3_nominal'].idxmax()
                criterion_name = "DB III gesamt (Nominal)"
            else:  # Default: Deckungsbeitrag III gesamt (Barwert)
                best_idx = df_results['total_db3_present_value'].idxmax()
                criterion_name = "DB III gesamt (Barwert)"
            best_option = df_results.loc[best_idx]
            
            ws_summary['A3'] = "OPTIMALE BATTERIEGRÖSSE"
            ws_summary['A3'].font = header_font
            ws_summary['A3'].fill = orange_fill
            ws_summary['A3'].font = white_font
            
            # Hole die tatsächliche Projektlaufzeit aus den Einstellungen
            project_lifetime = current_settings.get('project_lifetime_years', 20)
            
            # Benutzerfreundliche Anzeige der Amortisationszeit (999 Jahre = "Nicht erreicht")
            payback_display = "Nicht erreicht" if best_option['payback_period_years'] >= 999 else f"{best_option['payback_period_years']:.1f} Jahre"
            
            # Bestimme den optimalen Wert basierend auf dem Kriterium
            if optimization_criterion == "Deckungsbeitrag III gesamt (Nominal)":
                optimal_value = best_option.get('total_db3_nominal', 0)
                optimal_value_label = "DB III gesamt (Nominal) (€)"
            else:  # Default: Deckungsbeitrag III gesamt (Barwert)
                optimal_value = best_option.get('total_db3_present_value', 0)
                optimal_value_label = "DB III gesamt (Barwert) (€)"
            
            # Verwende results_summary wenn verfügbar (konsistent mit UI), sonst best_option
            if results_summary:
                optimal_capacity_display = results_summary.get('optimal_capacity', best_option['battery_capacity_kwh'])
                investment_cost_display = results_summary.get('investment_cost', best_option['investment_cost'])
                roi_percentage_display = results_summary.get('roi_percentage')
                payback_period_display = results_summary.get('payback_period_years')
                irr_display = results_summary.get('irr_percentage')
                criterion_value_display = results_summary.get('criterion_value', optimal_value)
                criterion_name_display = results_summary.get('criterion_name', criterion_name)
                
                # Berechne ROI falls nicht in results_summary
                if roi_percentage_display is None:
                    annual_savings_display = results_summary.get('annual_savings', best_option.get('annual_savings', 0))
                    roi_percentage_display = (annual_savings_display * project_lifetime / investment_cost_display - 1) * 100 if investment_cost_display else 0
                
                # Verwende payback_period aus results_summary oder best_option
                if payback_period_display is None:
                    payback_period_display = best_option.get('payback_period_years', 999)
                payback_display = "Nicht erreicht" if payback_period_display >= 999 else f"{payback_period_display:.1f} Jahre"
                
                # Verwende IRR aus results_summary oder best_option
                if irr_display is None:
                    irr_display = best_option.get('irr_percentage', 'N/A')
            else:
                # Fallback auf best_option wenn results_summary nicht verfügbar
                optimal_capacity_display = best_option['battery_capacity_kwh']
                investment_cost_display = best_option['investment_cost']
                criterion_value_display = optimal_value
                criterion_name_display = criterion_name
                annual_savings_display = best_option.get('annual_savings', 0)
                roi_percentage_display = (annual_savings_display * project_lifetime / investment_cost_display - 1) * 100 if investment_cost_display else 0
                payback_period_display = best_option.get('payback_period_years', 999)
                payback_display = "Nicht erreicht" if payback_period_display >= 999 else f"{payback_period_display:.1f} Jahre"
                irr_display = best_option.get('irr_percentage', 'N/A')
            
            summary_data = [
                ["Kennzahl", "Wert", "Einheit", "Bedeutung"],
                ["Optimale Kapazität", optimal_capacity_display, "kWh", "Wirtschaftlichste Batteriegröße"],
                ["Optimierungskriterium", criterion_name_display, "-", "Verwendetes Optimierungskriterium"],
                ["Optimaler Wert", criterion_value_display, "€" if "€" in optimal_value_label else ("Jahre" if "Jahre" in optimal_value_label else "%"), optimal_value_label],
                ["Amortisationszeit", payback_display, "", "Zeit bis zur Kostendeckung"],
                ["Anfangsinvestition", investment_cost_display, "€", "Gesamtinvestition in Batteriespeicher"],
                [f"ROI über {project_lifetime} Jahre", roi_percentage_display, "%", "Return on Investment"],
                ["IRR (Interner Zinsfuß)", irr_display, "%", "Jährliche Rendite der Investitation"]
            ]
            
            for i, row in enumerate(summary_data):
                for j, value in enumerate(row):
                    cell = ws_summary.cell(row=4+i, column=1+j)
                    cell.value = value
                    if i == 0:  # Header
                        cell.font = black_font
                        cell.fill = light_orange_fill
                    else:
                        cell.font = Font(size=11)
                    # B7 (Zeile 7, Spalte B) rechtsbündig ausrichten
                    if 4+i == 7 and 1+j == 2:  # Zeile 7, Spalte B
                        cell.alignment = Alignment(horizontal='right')
            
            # Füge erweiterte Aufschlüsselung der jährlichen Ersparnis hinzu
            current_row = 4 + len(summary_data) + 1
            
            ws_summary[f'A{current_row}'] = "JÄHRLICHE NETTO-KOSTENREDUKTION (Details)"
            ws_summary[f'A{current_row}'].font = header_font
            ws_summary[f'A{current_row}'].fill = orange_fill
            ws_summary[f'A{current_row}'].font = white_font
            current_row += 1
            
            # Header für die Aufschlüsselung
            savings_breakdown_data = [
                ["Komponente", "Wert", "Einheit", "Erklärung"],
            ]
            
            # Prüfe ob die detaillierten Komponenten verfügbar sind
            if best_option.get('savings_from_reduced_import') is not None and best_option.get('loss_from_reduced_export') is not None:
                savings_breakdown_data.extend([
                    ["Ersparnis durch reduzierten Netzbezug", best_option['savings_from_reduced_import'], "€/Jahr", "Weniger Strom vom Netz bezogen"],
                    ["Verlust durch reduzierte Einspeisung", -best_option['loss_from_reduced_export'], "€/Jahr", "Weniger PV-Strom ins Netz eingespeist"],
                    ["Netto-Kostenreduktion (Summe)", best_option['annual_savings'], "€/Jahr", "Gesamtersparnis für IRR/NPV-Berechnung"]
                ])
            else:
                # Fallback wenn keine Details verfügbar
                savings_breakdown_data.extend([
                    ["Netto-Kostenreduktion", best_option['annual_savings'], "€/Jahr", "Jährliche Reduktion der Nettokosten (Bezug - Einspeisung)"]
                ])
            
            for i, row in enumerate(savings_breakdown_data):
                for j, value in enumerate(row):
                    cell = ws_summary.cell(row=current_row+i, column=1+j)
                    cell.value = value
                    if i == 0:  # Header
                        cell.font = black_font
                        cell.fill = light_orange_fill
                    else:
                        cell.font = Font(size=11)
                        # Formatiere negative Werte rot
                        if j == 1 and isinstance(value, (int, float)) and value < 0:
                            cell.font = Font(size=11, color="FF0000")
            
            # Füge eine erklärende Notiz hinzu
            note_row = current_row + len(savings_breakdown_data) + 1
            ws_summary[f'A{note_row}'] = "Hinweis:"
            ws_summary[f'A{note_row}'].font = Font(size=10, bold=True, color="FF6B35")
            note_row += 1
            ws_summary[f'A{note_row}'] = "Die Netto-Kostenreduktion berücksichtigt sowohl die Einsparungen durch reduzierten Netzbezug als auch"
            ws_summary[f'A{note_row}'].font = Font(size=9, italic=True)
            ws_summary.merge_cells(f'A{note_row}:D{note_row}')
            note_row += 1
            ws_summary[f'A{note_row}'] = "den Verlust durch reduzierte Einspeisevergütung (da mehr Strom selbst verbraucht wird)."
            ws_summary[f'A{note_row}'].font = Font(size=9, italic=True)
            ws_summary.merge_cells(f'A{note_row}:D{note_row}')
        
        # 3. EINGABEPARAMETER
        ws_params = wb.create_sheet("Eingabeparameter")
        
        ws_params['A1'] = "EINGABEPARAMETER"
        ws_params['A1'].font = Font(size=16, bold=True, color="FF6B35")
        ws_params.merge_cells('A1:H1')
        ws_params['A1'].alignment = Alignment(horizontal='center')
        
        # Versuche, Max Lade-/Entladeleistung für die optimale Kapazität zu ermitteln
        battery_max_charge_kw_param = current_settings.get('battery_max_charge_kw')
        battery_max_discharge_kw_param = current_settings.get('battery_max_discharge_kw')
        try:
            if optimization_results:
                df_results_params = pd.DataFrame(optimization_results)
                # Verwende das gleiche Optimierungskriterium wie in der UI
                optimization_criterion_params = current_settings.get('optimization_criterion', 'Deckungsbeitrag III gesamt (Barwert)')
                # Nur DB Barwert und DB Nominal unterstützen
                if optimization_criterion_params == "Deckungsbeitrag III gesamt (Nominal)":
                    best_idx_params = df_results_params['total_db3_nominal'].idxmax()
                else:  # Default: Deckungsbeitrag III gesamt (Barwert)
                    best_idx_params = df_results_params['total_db3_present_value'].idxmax()
                best_option_params = df_results_params.loc[best_idx_params]
                optimal_capacity_kwh_params = best_option_params['battery_capacity_kwh']
                
                from data_import import load_battery_tech_params
                from config import BATTERY_COST_EXCEL_PATH
                tech_params = load_battery_tech_params(BATTERY_COST_EXCEL_PATH)
                if tech_params:
                    cap_int = int(round(optimal_capacity_kwh_params))
                    if cap_int in tech_params:
                        tech = tech_params[cap_int]
                        if 'max_charge_kw' in tech:
                            battery_max_charge_kw_param = tech['max_charge_kw']
                        if 'max_discharge_kw' in tech:
                            battery_max_discharge_kw_param = tech['max_discharge_kw']
        except Exception as e:
            print(f"Warnung: Konnte technische Batterie-Parameter für Eingabeparameter nicht laden: {e}")
        
        param_data = [
            ["Kategorie", "Parameter", "Wert", "Einheit", "Beschreibung"],
            ["Lastprofil", "Profiltyp", current_settings.get('load_profile_type', 'N/A'), "-", "Verwendetes Standard-Lastprofil"],
            ["Lastprofil", "Jährlicher Verbrauch", current_settings.get('annual_consumption_kwh', 'N/A'), "kWh", "Jährlicher Stromverbrauch"],
            ["Allgemein", "Anzahl Haushalte", current_settings.get('number_of_persons', 'N/A'), "Haushalte", "Anzahl der Haushalte (nicht Personen)"],
            ["Allgemein", "Bundesland", current_settings.get('bundesland', 'N/A'), "-", "Bundesland für Lastprofil"],
            ["Allgemein", "Bundesland-Code", current_settings.get('bundesland_code', 'N/A'), "-", "Bundesland-Code für Lastprofil"],
            ["Allgemein", "Simulationsjahr", current_settings.get('selected_year', 'N/A'), "Jahr", "Jahr für die Simulation"],
            ["Allgemein", "PV-Anlagengröße", current_settings.get('pv_system_size_kwp', 'N/A'), "kWp", "Nennleistung der PV-Anlage"],
            ["Batterie", "Kosten", "-", "-", "Kosten basieren auf Kostenkurve (battery_cost_curve_offers.xlsx)"],
            ["Batterie", "Ladeeffizienz", current_settings.get('battery_efficiency_charge', 'N/A'), "%", "Ladeeffizienz des Batteriespeichers"],
            ["Batterie", "Entladeeffizienz", current_settings.get('battery_efficiency_discharge', 'N/A'), "%", "Entladeeffizienz des Batteriespeichers"],
            [
                "Batterie",
                "Max. Ladeleistung",
                battery_max_charge_kw_param if battery_max_charge_kw_param is not None else current_settings.get('battery_max_charge_kw', 'N/A'),
                "kW",
                "Maximale Ladeleistung (aus C-Rate/Tech-Parametern berechnet, falls verfügbar)"
            ],
            [
                "Batterie",
                "Max. Entladeleistung",
                battery_max_discharge_kw_param if battery_max_discharge_kw_param is not None else current_settings.get('battery_max_discharge_kw', 'N/A'),
                "kW",
                "Maximale Entladeleistung (aus C-Rate/Tech-Parametern berechnet, falls verfügbar)"
            ],
            ["Batterie", "Anfangs-SOC", current_settings.get('initial_soc_percent', 'N/A'), "%", "Anfänglicher Batterieladezustand"],
            ["Batterie", "Min. SOC", current_settings.get('min_soc_percent', 'N/A'), "%", "Minimaler Batterieladezustand"],
            ["Batterie", "Max. SOC", current_settings.get('max_soc_percent', 'N/A'), "%", "Maximaler Batterieladezustand"],
            ["Batterie", "Jährlicher Kapazitätsverlust", current_settings.get('annual_capacity_loss_percent', 'N/A'), "%", "Jährlicher Batteriekapazitätsverlust"],
            ["Optimierung", "Min. Batteriekapazität", current_settings.get('min_battery_capacity', 'N/A'), "kWh", "Minimale untersuchte Batteriekapazität"],
            ["Optimierung", "Max. Batteriekapazität", current_settings.get('max_battery_capacity', 'N/A'), "kWh", "Maximale untersuchte Batteriekapazität"],
            ["Optimierung", "Schrittweite", current_settings.get('battery_step_size', 'N/A'), "kWh", "Schrittweite der Optimierung"],
            ["Strompreise", "Strombezugspreis", current_settings.get('price_grid_per_kwh', 'N/A'), "€/kWh", "Preis für Strombezug aus dem Netz"],
            ["Strompreise", "Einspeisevergütung", current_settings.get('price_feed_in_per_kwh', 'N/A'), "€/kWh", "Vergütung für Netzeinspeisung"],
            ["Finanzen", "Projektlaufzeit", current_settings.get('project_lifetime_years', 'N/A'), "Jahre", "Betrachtungszeitraum"],
            ["Finanzen", "Diskontierungsrate", current_settings.get('discount_rate', 'N/A'), "%", "Diskontierungsrate für NPV-Berechnung"],
            ["Finanzen", "Projektzins (DB)", current_settings.get('project_interest_rate_db', 'N/A'), "%", "Kalkulatorischer Zinssatz für DB II/III"],
            ["Dateien", "Verbrauchsdatei", current_settings.get('consumption_file', 'N/A'), "-", "Hochgeladene Verbrauchsdatei"],
            ["Dateien", "PV-Datei", current_settings.get('pv_file', 'N/A'), "-", "Hochgeladene PV-Datei"]
        ]
        
        for i, row in enumerate(param_data):
            for j, value in enumerate(row):
                cell = ws_params.cell(row=3+i, column=1+j)
                cell.value = value
                if i == 0:  # Header
                    cell.font = black_font
                    cell.fill = light_orange_fill
                else:
                    cell.font = Font(size=10)
        
        # 4. OPTIMIERUNGSERGEBNISSE
        ws_optimization = wb.create_sheet("Optimierungsergebnisse")
        
        ws_optimization['A1'] = "OPTIMIERUNGSERGEBNISSE"
        ws_optimization['A1'].font = Font(size=16, bold=True, color="FF6B35")
        ws_optimization.merge_cells('A1:H1')
        ws_optimization['A1'].alignment = Alignment(horizontal='center')
        
        if optimization_results:
            # DataFrame zu Excel
            df_results = pd.DataFrame(optimization_results)
            
            # Spaltenreihenfolge anpassen: 
            # - pv_generation_kwh und total_consumption_kwh neben self_consumption_rate
            # - grid_import_kwh neben grid_import_cost
            # - grid_export_kwh neben grid_export_revenue
            desired_order = [
                'battery_capacity_kwh',
                'autarky_rate',
                'self_consumption_rate',
                'self_consumption_kwh',
                'pv_generation_kwh',
                'total_consumption_kwh',
                'effective_annual_energy_cost',
                'grid_import_cost',
                'grid_import_kwh',
                'grid_export_revenue',
                'grid_export_kwh',
                'battery_losses_kwh',
                'investment_cost',
                'contribution_margin_1',
                'contribution_margin_2',
                'contribution_margin_3',
                'contribution_margin_3_npv',
                'payback_period_years',
                'roi_percentage',
                'npv',
                'irr_percentage'
            ]
            
            # Nur vorhandene Spalten in gewünschter Reihenfolge
            available_cols = [col for col in desired_order if col in df_results.columns]
            # Füge eventuell übrige Spalten am Ende hinzu
            remaining_cols = [col for col in df_results.columns if col not in available_cols]
            final_cols = available_cols + remaining_cols
            df_results = df_results[final_cols]
            
            for r in dataframe_to_rows(df_results.round(2), index=False, header=True):
                ws_optimization.append(r)
            
            # Header stylen
            for cell in ws_optimization[2]:
                cell.font = black_font
                cell.fill = light_orange_fill
            
            # Farbliche Gestaltung der Daten: Ergebniswerte dunkelgrün, Berechnungswerte hellgrün
            # Definiere welche Spalten welche Farbe bekommen sollen
            result_columns = ['effective_annual_energy_cost', 'autarky_rate', 'self_consumption_rate']
            calculation_columns = ['grid_import_cost', 'grid_import_kwh', 'grid_export_revenue', 'grid_export_kwh', 'self_consumption_kwh', 'pv_generation_kwh', 'total_consumption_kwh']
            
            # Finde die Spaltennummern für jede Spalte
            header_row = 2
            column_mapping = {}
            for col_idx, cell in enumerate(ws_optimization[header_row], start=1):
                if cell.value:
                    column_mapping[cell.value] = col_idx
            
            # Definiere Farben
            dark_green_fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")  # Dunkelgrün
            light_green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Hellgrün
            
            # Färbe die Datenzellen (ab Zeile 3 bis last_row)
            for row_idx in range(3, len(df_results) + 3):
                for col_name, col_idx in column_mapping.items():
                    cell = ws_optimization.cell(row=row_idx, column=col_idx)
                    if col_name in result_columns:
                        cell.fill = dark_green_fill
                    elif col_name in calculation_columns:
                        cell.fill = light_green_fill
            
            # Finde die letzte Zeile der Tabelle
            last_row = len(df_results) + 2  # +2 für Header und Titel
            
            # DETAILLIERTE BERECHNUNGSERKLÄRUNGEN
            explanation_start_row = last_row + 3
            
            # Überschrift für Erklärungen
            ws_optimization[f'A{explanation_start_row}'] = "Berechnungsmethoden"
            ws_optimization[f'A{explanation_start_row}'].font = Font(size=14, bold=True, color="FF6B35")
            ws_optimization.merge_cells(f'A{explanation_start_row}:H{explanation_start_row}')
            ws_optimization[f'A{explanation_start_row}'].alignment = Alignment(horizontal='center')
            
            # Dynamische Rahmenbedingungen für Beschreibungen
            annual_capacity_loss = current_settings.get('annual_capacity_loss_percent', 1.0)
            project_lifetime = current_settings.get('project_lifetime_years', 15)
            discount_rate = current_settings.get('discount_rate', 0.02) or 0.02
            discount_rate_pct = discount_rate * 100
            degradation_formula = f"(1 - {annual_capacity_loss/100:.4f})^(Jahr-1)"
            project_interest_rate_db = float(current_settings.get('project_interest_rate_db', 0.03) or 0.0)
            project_interest_rate_pct = project_interest_rate_db * 100
            
            def meta(label, method, source, description):
                return {
                    "label": label,
                    "method": method,
                    "source": source,
                    "description": description
                }
            
            explanation_lookup = {
                "battery_capacity_kwh": meta(
                    "Batteriekapazität",
                    "Optimierungsschleife",
                    "UI-Einstellungen (min/max/Schritt)",
                    "Kapazität der jeweils simulierten Batterie – iteriert im definierten Bereich."
                ),
                "autarky_rate": meta(
                    "Autarkiegrad",
                    "Simulation (15‑Minuten)",
                    "Zeitreihen Consumption/Grid",
                    "(Gesamtverbrauch − Netzbezug) / Gesamtverbrauch, Ergebnis als Anteil (0-1)."
                ),
                "self_consumption_rate": meta(
                    "Eigenverbrauchsquote",
                    "Simulation (15‑Minuten)",
                    "Zeitreihen PV/Verbrauch",
                    "(Direkter Eigenverbrauch + Batterieentladung) / PV-Erzeugung."
                ),
                "annual_energy_cost": meta(
                    "Jährliche Energiekosten",
                    "Simulation (Kostenbilanz)",
                    "Zeitreihen + Strompreise",
                    "(Grid_Import_kWh × Bezugspreis) − (Grid_Export_kWh × Einspeisetarif)."
                ),
                "self_consumption_kwh": meta(
                    "Eigenverbrauch (kWh)",
                    "Simulation (Aggregation)",
                    "Direct_Self_Consumption + Battery_Discharge",
                    "Direkt genutzte PV-Energie zuzüglich abgegebener Batterieenergie (netto)."
                ),
                "grid_export_kwh": meta(
                    "Netzeinspeisung (kWh)",
                    "Simulation (Aggregation)",
                    "Grid_Export_kWh",
                    "Summe aller Einspeisemengen der simulierten Zeitreihe."
                ),
                "pv_generation_kwh": meta(
                    "PV-Erzeugung (kWh)",
                    "Simulation (Aggregation)",
                    "PV_Generation_kWh",
                    "Summe der erzeugten PV-Energie über das Jahr."
                ),
                "total_consumption_kwh": meta(
                    "Gesamtverbrauch (kWh)",
                    "Simulation (Aggregation)",
                    "Consumption_kWh",
                    "Summe des Verbrauchs über alle Zeitschritte."
                ),
                "grid_import_kwh": meta(
                    "Netzbezug (kWh)",
                    "Simulation (Aggregation)",
                    "Grid_Import_kWh",
                    "Summe aller bezogenen Energiemengen aus dem Netz."
                ),
                "battery_losses_kwh": meta(
                    "Batterieverluste (kWh)",
                    "Simulation (Aggregation)",
                    "Charge-/Discharge-Losses",
                    "Summe der Lade- und Entladeverluste der Batterie."
                ),
                "investment_cost": meta(
                    "Investitionskosten",
                    "Kostenfunktion",
                    "Batteriespeicherkosten.xlsm",
                    "get_battery_cost() weist jeder Kapazität Kosten gemäß Excel-Kurve zu (inkl. Installationsanteil)."
                ),
                "annual_savings": meta(
                    "Jährliche Ersparnis",
                    "Differenzkosten",
                    "Simulation mit/ohne Batterie",
                    "(Reduzierter Netzbezug × Strompreis) − (reduzierte Einspeisung × Einspeisetarif)."
                ),
                "savings_from_reduced_import": meta(
                    "Ersparnis Netzbezug",
                    "Differenzkosten",
                    "Simulation mit/ohne Batterie",
                    "Reduzierter Netzbezug × Strompreis (positiver Beitrag)."
                ),
                "loss_from_reduced_export": meta(
                    "Verlust Einspeisung",
                    "Differenzkosten",
                    "Simulation mit/ohne Batterie",
                    "Reduzierte Einspeisung × Einspeisetarif (negativer Beitrag)."
                ),
                "payback_period_years": meta(
                    "Amortisationszeit",
                    "Kumulative Cashflows",
                    "Ersparnisse + Investition",
                    "Interpolierter Break-even sobald kumulierte Jahresersparnisse (mit Faktor "
                    f"{(1 - annual_capacity_loss/100):.4f}) die Investition decken."
                ),
                "npv": meta(
                    "Net Present Value",
                    "Diskontierte Cashflows",
                    "Jahresersparnisse",
                    f"Barwert der degradierenden Ersparnisse (Diskontsatz {discount_rate_pct:.1f}%) minus Investition."
                ),
                "irr_percentage": meta(
                    "Interner Zinsfuß",
                    "IRR auf Cashflows",
                    "Jahresersparnisse",
                    "Effektiver Zinssatz, bei dem der NPV der Cashflows = 0 wird."
                ),
                "annual_revenue": meta(
                    "Umsatz (DB I)",
                    "Deckungsbeitragsrechnung",
                    "Simulation mit/ohne Batterie",
                    "Reduzierter Netzbezug × Strompreis (Mehrerlös durch Speicher)."
                ),
                "annual_variable_costs": meta(
                    "Variable Kosten",
                    "Deckungsbeitragsrechnung",
                    "Simulation mit/ohne Batterie",
                    "Reduzierte Einspeisung × Einspeisetarif (entgangene Vergütung)."
                ),
                "contribution_margin_1": meta(
                    "DB I",
                    "Deckungsbeitragsrechnung",
                    "annual_revenue & annual_variable_costs",
                    "DB I = Umsatz − variable Kosten."
                ),
                "annual_depreciation": meta(
                    "Abschreibung",
                    "Lineare Abschreibung",
                    "Investition + Projektlaufzeit",
                    f"investment_cost / {project_lifetime} Jahre (linear)."
                ),
                "contribution_margin_2": meta(
                    "DB II",
                    "Deckungsbeitragsrechnung",
                    "DB I & Abschreibung",
                    "DB II = DB I − Abschreibung."
                ),
                "annual_interest": meta(
                    "Zinsen",
                    "Kalkulatorischer Zins",
                    "Investition",
                    f"(investment_cost / 2) × {project_interest_rate_pct:.1f}%."
                ),
                "contribution_margin_3": meta(
                    "DB III",
                    "Deckungsbeitragsrechnung",
                    "DB II & Zinsen",
                    "DB III = DB II − kalkulatorische Zinsen."
                ),
                "total_db3_nominal": meta(
                    "DB III gesamt (Nominal)",
                    "Summation mit Degradation",
                    "DB III + Degradationsfaktor",
                    f"Summe über {project_lifetime} Jahre: DB III × {degradation_formula}."
                ),
                "total_db3_present_value": meta(
                    "DB III gesamt (Barwert)",
                    "Summation mit Degradation & Diskontierung",
                    "DB III + Diskontsatz",
                    f"Wie nominal, zusätzlich diskontiert mit 1/(1+{discount_rate_pct:.1f}% )^Jahr."
                ),
                "roi_percentage": meta(
                    "ROI",
                    "Deckungsbeitrag",
                    "DB III & Investition",
                    "Durchschnittlicher DB III (nominal) / Investitionskosten × 100."
                )
            }
            
            explanation_table = [["Kennzahl", "Berechnungsmethode", "Datenquelle", "Beschreibung"]]
            for column in df_results.columns:
                meta_info = explanation_lookup.get(
                    column,
                    meta(
                        column,
                        "Simulation/Berechnung",
                        "siehe Datenbasis",
                        f"Beschreibung für '{column}' bitte ergänzen – aktuell generischer Wert."
                    )
                )
                label_text = f"{meta_info['label']} [{column}]" if meta_info['label'] != column else column
                explanation_table.append([
                    label_text,
                    meta_info['method'],
                    meta_info['source'],
                    meta_info['description']
                ])
            
            # Erklärungen in Excel einfügen
            for i, row_data in enumerate(explanation_table):
                target_row = explanation_start_row + 1 + i
                for j, value in enumerate(row_data):
                    cell = ws_optimization.cell(row=target_row, column=1 + j)
                    cell.value = value
                    if i == 0:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = orange_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.font = Font(size=10)
                        if j == 3:
                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                        else:
                            cell.alignment = Alignment(vertical='top')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
            
            ws_optimization.column_dimensions['A'].width = 28
            ws_optimization.column_dimensions['B'].width = 22
            ws_optimization.column_dimensions['C'].width = 26
            ws_optimization.column_dimensions['D'].width = 65
            
            # Hinweisblock unter der Tabelle
            hint_start_row = explanation_start_row + len(explanation_table) + 3
            ws_optimization.merge_cells(f'A{hint_start_row}:D{hint_start_row}')
            ws_optimization[f'A{hint_start_row}'] = "HINWEISE"
            ws_optimization[f'A{hint_start_row}'].font = Font(bold=True, color="FF6B35")
            
            hints = [
                "• Simulation: Alle Energie- und Kostenwerte stammen aus der 15-Minuten-Simulation (simulate_one_year).",
                f"• Degradation: Einsparnisse und DB-Werte werden jährlich mit {annual_capacity_loss:.1f}% Kapazitätsverlust via {degradation_formula} reduziert.",
                f"• Break-even & NPV: Cashflows werden diskontiert mit {discount_rate_pct:.1f}% und zwischen Jahren interpoliert.",
                "• Referenzfall: Für jede Kennzahl wird ein vollständiges Referenzszenario ohne Batterie simuliert (battery_capacity = 0 kWh)."
            ]
            
            for idx, text in enumerate(hints, start=hint_start_row + 1):
                ws_optimization.merge_cells(f'A{idx}:D{idx}')
                ws_optimization[f'A{idx}'] = text
                ws_optimization[f'A{idx}'].font = Font(size=10, italic=True)
        
        # 5. ENERGIEBILANZ
        ws_energy = wb.create_sheet("Energiebilanz")
        
        ws_energy['A1'] = "ENERGIEBILANZ"
        ws_energy['A1'].font = Font(size=16, bold=True, color="FF6B35")
        ws_energy.merge_cells('A1:H1')
        ws_energy['A1'].alignment = Alignment(horizontal='center')
        
        if calculation_details:
            energy_data = [
                ["Kategorie", "Kennzahl", "Wert", "Einheit", "Beschreibung"],
                ["Verbrauchsquellen", "Gesamtverbrauch", calculation_details['total_consumption_kwh'], "kWh", "Gesamter Energieverbrauch"],
                ["Verbrauchsquellen", "Direkter Eigenverbrauch", calculation_details['total_direct_self_consumption_kwh'], "kWh", "Direkt verbrauchte PV-Energie"],
                ["Verbrauchsquellen", "Batterieentladung", calculation_details['total_battery_discharge_kwh'], "kWh", "Aus Batterie entladene Energie"],
                ["Verbrauchsquellen", "Netzbezug", calculation_details['total_grid_import_kwh'], "kWh", "Aus dem Netz bezogene Energie"],
                ["PV-Energieverteilung", "Gesamterzeugung", calculation_details['total_pv_generation_kwh'], "kWh", "Gesamte PV-Erzeugung"],
                ["PV-Energieverteilung", "Direkter Eigenverbrauch", calculation_details['total_direct_self_consumption_kwh'], "kWh", "Direkt verbrauchte PV-Energie"],
                ["PV-Energieverteilung", "Batterieladung", calculation_details['total_battery_charge_kwh'], "kWh", "In Batterie geladene Energie"],
                ["PV-Energieverteilung", "Netzeinspeisung", calculation_details['total_grid_export_kwh'], "kWh", "Ins Netz eingespeiste Energie"],
                ["Batteriespeicher", "Batterieladung", calculation_details['total_battery_charge_kwh'], "kWh", "Gesamte Batterieladung"],
                ["Batteriespeicher", "Batterieentladung", calculation_details['total_battery_discharge_kwh'], "kWh", "Gesamte Batterieentladung"],
                ["Batteriespeicher", "Batterieverluste", calculation_details['total_battery_losses_kwh'], "kWh", "Batterieverluste durch Effizienz"],
                ["Kennzahlen", "Autarkiegrad", calculation_details['autarky_rate_percent'], "%", "Anteil selbst erzeugter Energie"],
                ["Kennzahlen", "Eigenverbrauchsquote", calculation_details['self_consumption_rate_percent'], "%", "Anteil verbrauchter PV-Energie"],
                ["Kennzahlen", "Gesamter Eigenverbrauch", calculation_details['self_consumption_kwh'], "kWh", "Gesamter Eigenverbrauch"]
            ]
            
            for i, row in enumerate(energy_data):
                for j, value in enumerate(row):
                    cell = ws_energy.cell(row=3+i, column=1+j)
                    cell.value = value
                    if i == 0:  # Header
                        cell.font = black_font
                        cell.fill = light_orange_fill
                    else:
                        cell.font = Font(size=10)
        
        # 6. VERBRAUCHSAUFSCHLÜSSELUNG
        ws_breakdown = wb.create_sheet("Verbrauchsaufschlüsselung")
        
        ws_breakdown['A1'] = "TÄGLICHE VERBRAUCHSAUFSCHLÜSSELUNG"
        ws_breakdown['A1'].font = Font(size=16, bold=True, color="FF6B35")
        ws_breakdown.merge_cells('A1:H1')
        ws_breakdown['A1'].alignment = Alignment(horizontal='center')
        
        # Erstelle Verbrauchsaufschlüsselung falls Daten verfügbar
        if consumption_series is not None:
            try:
                # Hole Parameter aus current_settings
                annual_consumption_kwh = current_settings.get('annual_consumption_kwh', 0)
                number_of_persons = current_settings.get('number_of_persons', 1)
                bundesland_code = current_settings.get('bundesland_code', 'BW')
                selected_year = current_settings.get('selected_year', 2024)
                
                # Verwende den eingegebenen Jahresverbrauch
                if annual_consumption_kwh == 0:
                    # Fallback: Berechne aus den geladenen Daten
                    annual_consumption_kwh = consumption_series.sum() if consumption_series is not None else 3500
                    print(f"Jahresverbrauch aus Daten berechnet (Fallback): {annual_consumption_kwh:.2f} kWh")
                else:
                    print(f"Jahresverbrauch verwendet: {annual_consumption_kwh:.2f} kWh")
                
                # Berechne tägliche Aufschlüsselung
                # Die Verbrauchsdaten sind bereits für 1 Haushalt deskaliert
                original_consumption_series = consumption_series
                
                breakdown_result = calculate_daily_consumption_breakdown(
                    original_consumption_series, 
                    annual_consumption_kwh, 
                    number_of_persons, 
                    bundesland_code, 
                    selected_year,
                    file_path=current_settings.get('consumption_file', None),  # Verwende die hochgeladene Verbrauchsdatei
                    profile_type=current_settings.get('load_profile_type', 'Custom')
                )
                
                breakdown_df = breakdown_result['daily_breakdown']
                monthly_breakdown_df = breakdown_result['monthly_breakdown']
                annual_stats = breakdown_result['annual_stats']
                
                if not breakdown_df.empty:
                    # Header
                    ws_breakdown['A3'] = "Detaillierte Verbrauchsaufschlüsselung für jeden Tag des Jahres"
                    ws_breakdown['A3'].font = Font(size=12, bold=True)
                    ws_breakdown.merge_cells('A3:H3')
                    
                    # Erklärung
                    explanation = [
                        "Erklärung der Spalten:",
                        "• Rohwert_1Mio_kWh: Wert aus der Excel-Datei (normiert auf 1 Mio kWh)",
                        "• Normiert_1Mio: Rohwert geteilt durch 1.000.000",
                        "• Haushalt_kWh: Normierter Wert × (jährlicher Gesamtverbrauch ÷ Anzahl Haushalte)",
                        "• Alle_Haushalte_kWh: Haushaltswert × Anzahl Haushalte"
                    ]
                    
                    # Zusätzliche Informationen
                    annual_consumption_per_household = annual_consumption_kwh / number_of_persons if number_of_persons > 0 else annual_consumption_kwh
                    info_text = [
                        f"",
                        f"📊 **Berechnungsgrundlagen:**",
                        f"• Jährlicher Gesamtverbrauch: {annual_consumption_kwh:,.0f} kWh",
                        f"• Jährlicher Verbrauch pro Haushalt: {annual_consumption_per_household:,.0f} kWh",
                        f"• Anzahl Haushalte: {number_of_persons}",
                        f"• Bundesland: {bundesland_code}",
                        f"• Jahr: {selected_year}",
                        f"• Skalierungsfaktor: {annual_consumption_per_household / 1_000_000.0:.6f}"
                    ]
                    
                    for i, text in enumerate(explanation, 5):
                        ws_breakdown[f'A{i}'] = text
                        ws_breakdown[f'A{i}'].font = Font(size=10, italic=True)
                    
                    # Zusätzliche Informationen hinzufügen
                    for i, text in enumerate(info_text, 10):
                        ws_breakdown[f'A{i}'] = text
                        if "📊" in text:
                            ws_breakdown[f'A{i}'].font = Font(size=11, bold=True)
                        else:
                            ws_breakdown[f'A{i}'].font = Font(size=10)
                    
                    # Daten-Tabelle
                    start_row = 16
                    
                    # Erweiterte Informationen hinzufügen
                    extended_info = [
                        f"",
                        f"📈 **Jährliche Statistiken:**",
                        f"• Gesamtverbrauch: {annual_stats['total_energy_kwh']:,.0f} kWh",
                        f"• Durchschnittlicher Tagesverbrauch: {annual_stats['avg_daily_energy_kwh']:.1f} kWh",
                        f"• Spitzenleistung: {annual_stats['peak_power_kw']:.2f} kW",
                        f"• Minimalleistung: {annual_stats['min_power_kw']:.2f} kW",
                        f"• Durchschnittsleistung: {annual_stats['avg_power_kw']:.2f} kW",
                        f"• Verbrauch pro Person: {annual_stats['consumption_per_person_kwh']:,.0f} kWh",
                        f"• Tage mit Verbrauch: {annual_stats['days_with_consumption']}"
                    ]
                    
                    for i, text in enumerate(extended_info, 20):
                        ws_breakdown[f'A{i}'] = text
                        if "📈" in text:
                            ws_breakdown[f'A{i}'].font = Font(size=11, bold=True)
                        else:
                            ws_breakdown[f'A{i}'].font = Font(size=10)
                    
                    # Header der Tabelle
                    headers = ['Datum', 'Wochentag', 'Tagestyp', 'Tagesverbrauch_kWh', 'Spitzenleistung_kW', 'Minimalleistung_kW', 
                             'Durchschnittsleistung_kW', 'Stunden_mit_Verbrauch', 'Verbrauch_pro_Haushalt_kWh', 
                             'Rohwert_1Mio_kWh', 'Normiert_1Mio', 'Alle_Haushalte_kWh', 'Anteil_am_Jahresverbrauch_%']
                    for col, header in enumerate(headers, 1):
                        cell = ws_breakdown.cell(row=start_row, column=col)
                        cell.value = header
                        cell.font = black_font
                        cell.fill = light_orange_fill
                    
                    # Daten
                    for row_idx, (_, data_row) in enumerate(breakdown_df.iterrows(), start_row + 1):
                        for col_idx, value in enumerate(data_row, 1):
                            cell = ws_breakdown.cell(row=row_idx, column=col_idx)
                            if isinstance(value, (int, float)):
                                if 'Anteil' in headers[col_idx-1]:
                                    cell.value = round(value, 2)  # Prozent mit 2 Dezimalstellen
                                else:
                                    cell.value = round(value, 4)
                            else:
                                cell.value = value
                            cell.font = Font(size=9)
                    
                    # Summenzeile
                    sum_row = start_row + len(breakdown_df) + 1
                    ws_breakdown[f'A{sum_row}'] = "SUMME"
                    ws_breakdown[f'A{sum_row}'].font = Font(bold=True)
                    ws_breakdown[f'A{sum_row}'].fill = gray_fill
                    
                    # Summen berechnen (nur für relevante Spalten)
                    sum_columns = ['D', 'I', 'J', 'L']  # Tagesverbrauch, Verbrauch_pro_Haushalt, Rohwert, Alle_Haushalte
                    for col in sum_columns:
                        if col in [chr(65 + i) for i in range(len(headers))]:
                            ws_breakdown[f'{col}{sum_row}'] = f"=SUM({col}{start_row+1}:{col}{sum_row-1})"
                            cell = ws_breakdown[f'{col}{sum_row}']
                            cell.font = Font(bold=True)
                            cell.fill = gray_fill
                    
                    # Spaltenbreiten anpassen
                    column_widths = [12, 12, 10, 15, 15, 15, 18, 18, 18, 18, 15, 15, 18]
                    for col, width in enumerate(column_widths, 1):
                        ws_breakdown.column_dimensions[chr(64 + col)].width = width
                    
                    # MONATSAUFSCHLÜSSELUNG hinzufügen
                    if not monthly_breakdown_df.empty:
                        # Abstand
                        monthly_start_row = sum_row + 5
                        
                        # Titel für Monatsaufschlüsselung
                        ws_breakdown[f'A{monthly_start_row}'] = "MONATSAUFSCHLÜSSELUNG"
                        ws_breakdown[f'A{monthly_start_row}'].font = Font(size=14, bold=True, color="FF6B35")
                        ws_breakdown.merge_cells(f'A{monthly_start_row}:H{monthly_start_row}')
                        ws_breakdown[f'A{monthly_start_row}'].alignment = Alignment(horizontal='center')
                        
                        # Header der Monatsaufschlüsselung
                        monthly_headers = ['Monat', 'Monatsverbrauch_kWh', 'Spitzenleistung_kW', 'Minimalleistung_kW', 
                                         'Durchschnittlicher_Tagesverbrauch_kWh', 'Verbrauch_pro_Person_kWh', 'Anteil_am_Jahresverbrauch_%']
                        for col, header in enumerate(monthly_headers, 1):
                            cell = ws_breakdown.cell(row=monthly_start_row + 2, column=col)
                            cell.value = header
                            cell.font = black_font
                            cell.fill = light_orange_fill
                        
                        # Monatsdaten
                        for row_idx, (_, data_row) in enumerate(monthly_breakdown_df.iterrows(), monthly_start_row + 3):
                            for col_idx, value in enumerate(data_row, 1):
                                cell = ws_breakdown.cell(row=row_idx, column=col_idx)
                                if isinstance(value, (int, float)):
                                    if 'Anteil' in monthly_headers[col_idx-1]:
                                        cell.value = round(value, 2)  # Prozent mit 2 Dezimalstellen
                                    else:
                                        cell.value = round(value, 4)
                                else:
                                    cell.value = value
                                cell.font = Font(size=9)
                        
                        # Monatssummenzeile
                        monthly_sum_row = monthly_start_row + 3 + len(monthly_breakdown_df)
                        ws_breakdown[f'A{monthly_sum_row}'] = "MONATSSUMME"
                        ws_breakdown[f'A{monthly_sum_row}'].font = Font(bold=True)
                        ws_breakdown[f'A{monthly_sum_row}'].fill = gray_fill
                        
                        # Monatssummen berechnen
                        monthly_sum_columns = ['B', 'E', 'F']  # Monatsverbrauch, Durchschnittlicher_Tagesverbrauch, Verbrauch_pro_Person
                        for col in monthly_sum_columns:
                            if col in [chr(65 + i) for i in range(len(monthly_headers))]:
                                ws_breakdown[f'{col}{monthly_sum_row}'] = f"=SUM({col}{monthly_start_row+3}:{col}{monthly_sum_row-1})"
                                cell = ws_breakdown[f'{col}{monthly_sum_row}']
                                cell.font = Font(bold=True)
                                cell.fill = gray_fill
                    
                else:
                    ws_breakdown['A5'] = "Keine Verbrauchsdaten verfügbar für die Aufschlüsselung."
                    ws_breakdown['A5'].font = Font(size=12, italic=True)
                    
            except Exception as e:
                ws_breakdown['A5'] = f"Fehler bei der Verbrauchsaufschlüsselung: {str(e)}"
                ws_breakdown['A5'].font = Font(size=12, italic=True, color="FF0000")
        else:
            ws_breakdown['A5'] = "Keine Verbrauchsdaten verfügbar."
            ws_breakdown['A5'].font = Font(size=12, italic=True)
        
        # 7. TAGESBILANZ
        ws_daily_balance = wb.create_sheet("Tagesbilanz")
        
        ws_daily_balance['A1'] = "TAGESBILANZ - ENERGIEBILANZ FÜR JEDEN TAG"
        ws_daily_balance['A1'].font = Font(size=16, bold=True, color="FF6B35")
        ws_daily_balance.merge_cells('A1:O1')
        ws_daily_balance['A1'].alignment = Alignment(horizontal='center')
        
        # Erstelle Tagesbilanz falls Simulationsdaten verfügbar
        if optimal_sim_result is not None and 'time_series_data' in optimal_sim_result:
            try:
                time_series_data = optimal_sim_result['time_series_data']
                
                # Gruppiere Daten nach Tagen
                daily_balance_data = []
                
                # Erstelle Datumsbereich für das Jahr
                start_date = time_series_data.index.min().normalize()
                end_date = time_series_data.index.max().normalize()
                date_range = pd.date_range(start=start_date, end=end_date, freq='D')
                
                for date in date_range:
                    # Filtere Daten für diesen Tag (korrekte Datumsfilterung für stündliche Daten)
                    day_start = date
                    day_end = date + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
                    day_data = time_series_data[(time_series_data.index >= day_start) & (time_series_data.index <= day_end)]
                    
                    if len(day_data) > 0:
                        # Berechne tägliche Summen (korrekte Spaltennamen aus model.py)
                        daily_consumption = day_data['Consumption_kWh'].sum()
                        daily_pv_generation = day_data['PV_Generation_kWh'].sum()
                        daily_direct_self_consumption = day_data['Direct_Self_Consumption_kWh'].sum()
                        daily_battery_charge = day_data['Battery_Charge_kWh'].sum()
                        daily_battery_discharge = day_data['Battery_Discharge_kWh'].sum()
                        daily_grid_import = day_data['Grid_Import_kWh'].sum()
                        daily_grid_export = day_data['Grid_Export_kWh'].sum()
                        
                        # Berechne separate Lade- und Entladeverluste
                        daily_battery_charge_losses = day_data['Battery_Charge_Losses_kWh'].sum()
                        daily_battery_discharge_losses = day_data['Battery_Discharge_Losses_kWh'].sum()
                        total_battery_losses = daily_battery_charge_losses + daily_battery_discharge_losses
                        
                        # Berechne Eigenverbrauch aus Batterie (KORRIGIERT: Brutto-Ladung verwenden)
                        # Wichtig: Wir verwenden die Brutto-Ladung, da diese die tatsächlich verbrauchte PV-Energie ist
                        # Die Netto-Entladung verfälscht das Ergebnis, da Verluste "verschwinden"
                        battery_self_consumption = daily_battery_charge  # Brutto-Ladung statt Netto-Entladung
                        
                        # Berechne Gesamteigenverbrauch (korrigiert: direkter + Brutto-Batterieladung)
                        # Dies gibt die vollständige PV-Energie wieder, die selbst genutzt wird
                        total_self_consumption = daily_direct_self_consumption + battery_self_consumption
                        
                        # Berechne Autarkiegrad
                        autarky_rate = (total_self_consumption / daily_consumption * 100) if daily_consumption > 0 else 0
                        
                        # Berechne Eigenverbrauchsquote (KORRIGIERT: Brutto-Ladung für korrekte Berechnung)
                        # Eigenverbrauchsquote = (Direkter Eigenverbrauch + Brutto-Batterieladung) / PV-Erzeugung
                        # Dies ist die korrekte Berechnung, da alle PV-Energie erfasst wird, die selbst genutzt wird
                        # Keine Begrenzung auf 100% mehr nötig, da die Berechnung jetzt korrekt ist
                        self_consumption_rate = (total_self_consumption / daily_pv_generation * 100) if daily_pv_generation > 0 else 0
                        
                        # Berechne Batterieauslastung (korrigiert: Prozent der maximalen Ladeleistung)
                        # daily_battery_charge ist in kWh, battery_max_charge_kw ist in kW
                        # Maximal mögliche Ladung pro Tag = max_charge_kw * 24h
                        battery_max_charge_kw = current_settings.get('battery_max_charge_kw', 10.0)
                        max_daily_charge_kwh = battery_max_charge_kw * 24.0
                        battery_utilization = (daily_battery_charge / max_daily_charge_kwh * 100) if max_daily_charge_kwh > 0 else 0
                        
                        # Berechne Batterieladezustand zu bestimmten Uhrzeiten (KORRIGIERT: aktuelle Batteriekapazität aus Simulation)
                        # Verwende die aktuelle Batteriekapazität aus der Simulation (berücksichtigt Kapazitätsalterung)
                        battery_capacity_kwh = current_settings.get('battery_capacity_kwh', 10.0)
                        
                        # Versuche die aktuelle Kapazität aus verschiedenen Quellen zu holen
                        if 'current_capacity_kwh' in current_settings:
                            battery_capacity_kwh = current_settings['current_capacity_kwh']
                        elif 'optimal_sim_result' in locals() and optimal_sim_result and 'kpis' in optimal_sim_result:
                            # Hole aus den Simulationsergebnissen
                            battery_capacity_kwh = optimal_sim_result['kpis'].get('current_capacity_kwh', battery_capacity_kwh)
                        
                        # Debug-Ausgabe (kann später entfernt werden)
                        print(f"Verwendete Batteriekapazität für SOC-Berechnung: {battery_capacity_kwh} kWh")
                        
                        # SOC zu 6:00 Uhr (morgens, vor PV-Erzeugung)
                        soc_6h = 0.0
                        if len(day_data) > 0:
                            # Finde den Eintrag um 6:00 Uhr basierend auf dem tatsächlichen Zeitstempel
                            target_time_6h = date.replace(hour=6, minute=0, second=0, microsecond=0)
                            # Suche den nächstgelegenen Zeitstempel um 6:00 Uhr
                            closest_6h = None
                            min_diff_6h = float('inf')
                            for idx, row in day_data.iterrows():
                                time_diff = abs((idx - target_time_6h).total_seconds())
                                if time_diff < min_diff_6h:
                                    min_diff_6h = time_diff
                                    closest_6h = row
                            
                            if closest_6h is not None and 'SOC_kWh' in closest_6h:
                                soc_raw = closest_6h['SOC_kWh']
                                # SOC-Werte sind immer in kWh, konvertiere zu Prozent der aktuellen Batteriekapazität
                                soc_6h = (soc_raw / battery_capacity_kwh * 100) if battery_capacity_kwh > 0 else 0
                                # Begrenze auf 0-100%
                                soc_6h = max(0.0, min(100.0, soc_6h))
                                # Debug-Ausgabe
                                print(f"Tag {date.strftime('%d.%m.%Y')} - SOC 6:00: {soc_raw:.2f} kWh -> {soc_6h:.1f}%")
                        
                        # SOC zu 14:00 Uhr (mittags, maximale PV-Erzeugung)
                        soc_14h = 0.0
                        if len(day_data) > 0:
                            # Finde den Eintrag um 14:00 Uhr basierend auf dem tatsächlichen Zeitstempel
                            target_time_14h = date.replace(hour=14, minute=0, second=0, microsecond=0)
                            # Suche den nächstgelegenen Zeitstempel um 14:00 Uhr
                            closest_14h = None
                            min_diff_14h = float('inf')
                            for idx, row in day_data.iterrows():
                                time_diff = abs((idx - target_time_14h).total_seconds())
                                if time_diff < min_diff_14h:
                                    min_diff_14h = time_diff
                                    closest_14h = row
                            
                            if closest_14h is not None and 'SOC_kWh' in closest_14h:
                                soc_raw = closest_14h['SOC_kWh']
                                # SOC-Werte sind immer in kWh, konvertiere zu Prozent der aktuellen Batteriekapazität
                                soc_14h = (soc_raw / battery_capacity_kwh * 100) if battery_capacity_kwh > 0 else 0
                                # Begrenze auf 0-100%
                                soc_14h = max(0.0, min(100.0, soc_14h))
                        
                        # SOC zu 22:00 Uhr (abends, nach PV-Erzeugung)
                        soc_22h = 0.0
                        if len(day_data) > 0:
                            # Finde den Eintrag um 22:00 Uhr basierend auf dem tatsächlichen Zeitstempel
                            target_time_22h = date.replace(hour=22, minute=0, second=0, microsecond=0)
                            # Suche den nächstgelegenen Zeitstempel um 22:00 Uhr
                            closest_22h = None
                            min_diff_22h = float('inf')
                            for idx, row in day_data.iterrows():
                                time_diff = abs((idx - target_time_22h).total_seconds())
                                if time_diff < min_diff_22h:
                                    min_diff_22h = time_diff
                                    closest_22h = row
                            
                            if closest_22h is not None and 'SOC_kWh' in closest_22h:
                                soc_raw = closest_22h['SOC_kWh']
                                # SOC-Werte sind immer in kWh, konvertiere zu Prozent der aktuellen Batteriekapazität
                                soc_22h = (soc_raw / battery_capacity_kwh * 100) if battery_capacity_kwh > 0 else 0
                                # Begrenze auf 0-100%
                                soc_22h = max(0.0, min(100.0, soc_22h))
                        
                        daily_balance_data.append({
                            'Datum': date.strftime('%d.%m.%Y'),
                            'Wochentag': date.strftime('%A'),
                            'Verbrauch_kWh': daily_consumption,
                            'PV_Erzeugung_kWh': daily_pv_generation,
                            'Direkter_Eigenverbrauch_kWh': daily_direct_self_consumption,
                            'Batterieladung_kWh': daily_battery_charge,
                            'Batterieentladung_kWh': daily_battery_discharge,
                            'Ladeverluste_kWh': daily_battery_charge_losses,
                            'Entladeverluste_kWh': daily_battery_discharge_losses,
                            'Gesamtverluste_kWh': total_battery_losses,
                            'Netzbezug_kWh': daily_grid_import,
                            'Netzeinspeisung_kWh': daily_grid_export,
                            'Eigenverbrauch_aus_Batterie_kWh': battery_self_consumption,
                            'Gesamteigenverbrauch_kWh': total_self_consumption,
                            'Autarkiegrad_%': autarky_rate,
                            'Eigenverbrauchsquote_%': self_consumption_rate,
                            'Batterieauslastung_%': battery_utilization,
                            'SOC_6h_%': round(soc_6h, 1),
                            'SOC_14h_%': round(soc_14h, 1),
                            'SOC_22h_%': round(soc_22h, 1)
                        })
                
                # Erstelle DataFrame
                daily_balance_df = pd.DataFrame(daily_balance_data)
                
                if not daily_balance_df.empty:
                    # Erklärung
                    explanation = [
                        "Erklärung der Tagesbilanz:",
                        "• Direkter Eigenverbrauch: PV-Energie direkt an Last verbraucht",
                        "• Batterieladung: In Batterie geladene Energie (Brutto - inkl. Verluste)",
                        "• Batterieentladung: Aus Batterie entladene Energie (Netto - nach Verlusten)",
                        "• Ladeverluste: Energieverluste beim Laden der Batterie",
                        "• Entladeverluste: Energieverluste beim Entladen der Batterie",
                        "• Gesamtverluste: Summe aller Batterieverluste",
                        "• Eigenverbrauch aus Batterie: Batterieladung (Brutto - inkl. Verluste)",
                        "• Gesamteigenverbrauch: Direkter + Brutto-Batterieladung",
                        "• Autarkiegrad: Anteil selbst erzeugter Energie am Gesamtverbrauch",
                        "• Eigenverbrauchsquote: Anteil verbrauchter PV-Energie (korrekte Brutto-Berechnung)",
                        "• Batterieauslastung: Prozentuale Auslastung der Batterie",
                        "• SOC_6h: Batterieladezustand um 6:00 Uhr (morgens)",
                        "• SOC_14h: Batterieladezustand um 14:00 Uhr (mittags)",
                        "• SOC_22h: Batterieladezustand um 22:00 Uhr (abends)"
                    ]
                    
                    for i, text in enumerate(explanation, 3):
                        ws_daily_balance[f'A{i}'] = text
                        ws_daily_balance[f'A{i}'].font = Font(size=10, italic=True)
                    
                    # Header der Tabelle
                    start_row = 15
                    headers = ['Datum', 'Wochentag', 'Verbrauch_kWh', 'PV_Erzeugung_kWh', 'Direkter_Eigenverbrauch_kWh',
                              'Batterieladung_kWh', 'Batterieentladung_kWh', 'Ladeverluste_kWh', 'Entladeverluste_kWh',
                              'Gesamtverluste_kWh', 'Netzbezug_kWh', 'Netzeinspeisung_kWh', 'Eigenverbrauch_aus_Batterie_kWh',
                              'Gesamteigenverbrauch_kWh', 'Autarkiegrad_%', 'Eigenverbrauchsquote_%', 'Batterieauslastung_%',
                              'SOC_6h_%', 'SOC_14h_%', 'SOC_22h_%']
                    
                    for col, header in enumerate(headers, 1):
                        cell = ws_daily_balance.cell(row=start_row, column=col)
                        cell.value = header
                        cell.font = black_font
                        cell.fill = light_orange_fill
                    
                    # Daten
                    for row_idx, (_, data_row) in enumerate(daily_balance_df.iterrows(), start_row + 1):
                        for col_idx, value in enumerate(data_row, 1):
                            cell = ws_daily_balance.cell(row=row_idx, column=col_idx)
                            if isinstance(value, (int, float)):
                                if 'Autarkiegrad' in headers[col_idx-1] or 'Eigenverbrauchsquote' in headers[col_idx-1] or 'Batterieauslastung' in headers[col_idx-1]:
                                    cell.value = round(value, 2)  # Prozent mit 2 Dezimalstellen
                                else:
                                    cell.value = round(value, 4)
                            else:
                                cell.value = value
                            cell.font = Font(size=9)
                    
                    # Summenzeile
                    sum_row = start_row + len(daily_balance_df) + 1
                    ws_daily_balance[f'A{sum_row}'] = "JAHRESSUMME"
                    ws_daily_balance[f'A{sum_row}'].font = Font(bold=True)
                    ws_daily_balance[f'A{sum_row}'].fill = gray_fill
                    
                    # Summen berechnen (nur für relevante Spalten)
                    sum_columns = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']  # Alle kWh-Spalten + Prozent-Spalten
                    for col in sum_columns:
                        if col in [chr(65 + i) for i in range(len(headers))]:
                            ws_daily_balance[f'{col}{sum_row}'] = f"=SUM({col}{start_row+1}:{col}{sum_row-1})"
                            cell = ws_daily_balance[f'{col}{sum_row}']
                            cell.font = Font(bold=True)
                            cell.fill = gray_fill
                    
                    # Durchschnittswerte für Prozentangaben
                    avg_row = sum_row + 2
                    ws_daily_balance[f'A{avg_row}'] = "JAHRESDURCHSCHNITT"
                    ws_daily_balance[f'A{avg_row}'].font = Font(bold=True)
                    ws_daily_balance[f'A{avg_row}'].fill = gray_fill
                    
                    # Durchschnitt für Prozentangaben berechnen (ohne SOC-Spalten)
                    avg_columns = ['P', 'Q']  # Autarkiegrad, Eigenverbrauchsquote, Batterieauslastung (ohne SOC)
                    for col in avg_columns:
                        if col in [chr(65 + i) for i in range(len(headers))]:
                            ws_daily_balance[f'{col}{avg_row}'] = f"=AVERAGE({col}{start_row+1}:{col}{sum_row-1})"
                            cell = ws_daily_balance[f'{col}{avg_row}']
                            cell.font = Font(bold=True)
                            cell.fill = gray_fill
                    
                    # Spaltenbreiten anpassen
                    column_widths = [12, 12, 15, 15, 20, 15, 18, 15, 15, 15, 15, 15, 15, 22, 20, 15, 20, 18]
                    for col, width in enumerate(column_widths, 1):
                        ws_daily_balance.column_dimensions[chr(64 + col)].width = width
                    
                    # 15-MINUTEN-AUFSCHLÜSSELUNG FÜR DEN ERSTEN TAG
                    # Leerzeile nach der Tagesbilanz
                    first_day_start_row = avg_row + 5
                    
                    # Überschrift für 15-Minuten-Aufschlüsselung
                    ws_daily_balance[f'A{first_day_start_row}'] = "15-MINUTEN-AUFSCHLÜSSELUNG FÜR DEN ERSTEN TAG"
                    ws_daily_balance[f'A{first_day_start_row}'].font = Font(size=14, bold=True, color="FF6B35")
                    ws_daily_balance.merge_cells(f'A{first_day_start_row}:O{first_day_start_row}')
                    ws_daily_balance[f'A{first_day_start_row}'].alignment = Alignment(horizontal='center')
                    
                    # Erstelle 15-Minuten-Aufschlüsselung für den ersten Tag
                    try:
                        # Finde den ersten Tag
                        first_day = time_series_data.index.min().normalize()
                        day_start = first_day
                        day_end = first_day + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
                        
                        # Filtere Daten für den ersten Tag
                        first_day_data = time_series_data[(time_series_data.index >= day_start) & (time_series_data.index <= day_end)]
                        
                        if len(first_day_data) > 0:
                            # Erstelle 15-Minuten-Aufschlüsselung
                            first_day_15min_data = []
                            
                            for timestamp, row in first_day_data.iterrows():
                                # Berechne separate Lade- und Entladeverluste für diesen Zeitpunkt
                                battery_charge_losses = row['Battery_Charge_Losses_kWh']
                                battery_discharge_losses = row['Battery_Discharge_Losses_kWh']
                                total_battery_losses = battery_charge_losses + battery_discharge_losses
                                
                                # Berechne Eigenverbrauch aus Batterie
                                battery_self_consumption = row['Battery_Discharge_kWh']
                                
                                # Berechne Gesamteigenverbrauch
                                total_self_consumption = row['Direct_Self_Consumption_kWh'] + battery_self_consumption
                                
                                # Berechne Autarkiegrad
                                autarky_rate = (total_self_consumption / row['Consumption_kWh'] * 100) if row['Consumption_kWh'] > 0 else 0
                                
                                # Berechne Eigenverbrauchsquote
                                self_consumption_rate = (total_self_consumption / row['PV_Generation_kWh'] * 100) if row['PV_Generation_kWh'] > 0 else 0
                                
                                # Berechne Batterieauslastung (prozentual zur maximalen Ladeleistung)
                                battery_utilization = (row['Battery_Charge_kWh'] / 1.0) * 100  # Annahme: 1 kW maximale Ladeleistung
                                
                                first_day_15min_data.append({
                                    'Zeitstempel': timestamp.strftime('%d.%m.%Y %H:%M'),
                                    'Uhrzeit': timestamp.strftime('%H:%M'),
                                    'Verbrauch_kWh': row['Consumption_kWh'],
                                    'PV_Erzeugung_kWh': row['PV_Generation_kWh'],
                                    'Direkter_Eigenverbrauch_kWh': row['Direct_Self_Consumption_kWh'],
                                    'Batterieladung_kWh': row['Battery_Charge_kWh'],
                                    'Batterieentladung_kWh': row['Battery_Discharge_kWh'],
                                    'Ladeverluste_kWh': battery_charge_losses,
                                    'Entladeverluste_kWh': battery_discharge_losses,
                                    'Gesamtverluste_kWh': total_battery_losses,
                                    'Netzbezug_kWh': row['Grid_Import_kWh'],
                                    'Netzeinspeisung_kWh': row['Grid_Export_kWh'],
                                    'Eigenverbrauch_aus_Batterie_kWh': battery_self_consumption,
                                    'Gesamteigenverbrauch_kWh': total_self_consumption,
                                    'Autarkiegrad_%': autarky_rate,
                                    'Eigenverbrauchsquote_%': self_consumption_rate,
                                    'Batterieauslastung_%': battery_utilization
                                })
                            
                            # Erstelle DataFrame
                            first_day_15min_df = pd.DataFrame(first_day_15min_data)
                            
                            if not first_day_15min_df.empty:
                                # Erklärung
                                explanation = [
                                    f"15-Minuten-Aufschlüsselung für den ersten Tag: {first_day.strftime('%d.%m.%Y')}",
                                    "• Direkter Eigenverbrauch: PV-Energie direkt an Last verbraucht",
                                    "• Batterieladung: In Batterie geladene Energie (Brutto - inkl. Verluste)",
                                    "• Batterieentladung: Aus Batterie entladene Energie (Netto - nach Verlusten)",
                                    "• Ladeverluste: Energieverluste beim Laden der Batterie",
                                    "• Entladeverluste: Energieverluste beim Entladen der Batterie",
                                    "• Gesamtverluste: Summe aller Batterieverluste",
                                    "• Eigenverbrauch aus Batterie: Batterieentladung (Netto)",
                                    "• Gesamteigenverbrauch: Direkter + Batterie-Eigenverbrauch",
                                    "• Autarkiegrad: Anteil selbst erzeugter Energie am Gesamtverbrauch",
                                    "• Eigenverbrauchsquote: Anteil verbrauchter PV-Energie",
                                    "• Batterieauslastung: Prozentuale Auslastung der Batterie"
                                ]
                                
                                for i, text in enumerate(explanation, first_day_start_row + 2):
                                    ws_daily_balance[f'A{i}'] = text
                                    ws_daily_balance[f'A{i}'].font = Font(size=10, italic=True)
                                
                                # Header der 15-Minuten-Tabelle
                                start_row_15min = first_day_start_row + 13
                                headers_15min = ['Zeitstempel', 'Uhrzeit', 'Verbrauch_kWh', 'PV_Erzeugung_kWh', 'Direkter_Eigenverbrauch_kWh',
                                              'Batterieladung_kWh', 'Batterieentladung_kWh', 'Ladeverluste_kWh', 'Entladeverluste_kWh',
                                              'Gesamtverluste_kWh', 'Netzbezug_kWh', 'Netzeinspeisung_kWh', 'Eigenverbrauch_aus_Batterie_kWh',
                                              'Gesamteigenverbrauch_kWh', 'Autarkiegrad_%', 'Eigenverbrauchsquote_%', 'Batterieauslastung_%']
                                
                                for col, header in enumerate(headers_15min, 1):
                                    cell = ws_daily_balance.cell(row=start_row_15min, column=col)
                                    cell.value = header
                                    cell.font = black_font
                                    cell.fill = light_orange_fill
                                
                                # Daten der 15-Minuten-Tabelle
                                for row_idx, (_, data_row) in enumerate(first_day_15min_df.iterrows(), start_row_15min + 1):
                                    for col_idx, value in enumerate(data_row, 1):
                                        cell = ws_daily_balance.cell(row=row_idx, column=col_idx)
                                        if isinstance(value, (int, float)):
                                            if 'Autarkiegrad' in headers_15min[col_idx-1] or 'Eigenverbrauchsquote' in headers_15min[col_idx-1] or 'Batterieauslastung' in headers_15min[col_idx-1]:
                                                cell.value = round(value, 2)  # Prozent mit 2 Dezimalstellen
                                            else:
                                                cell.value = round(value, 4)
                                        else:
                                            cell.value = value
                                        cell.font = Font(size=9)
                                
                                # Tages-Summenzeile für 15-Minuten-Daten
                                sum_row_15min = start_row_15min + len(first_day_15min_df) + 1
                                ws_daily_balance[f'A{sum_row_15min}'] = "TAGESSUMME (15-MINUTEN)"
                                ws_daily_balance[f'A{sum_row_15min}'].font = Font(bold=True)
                                ws_daily_balance[f'A{sum_row_15min}'].fill = gray_fill
                                
                                # Summen berechnen (nur für relevante Spalten)
                                sum_columns_15min = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']  # Alle kWh-Spalten
                                for col in sum_columns_15min:
                                    if col in [chr(65 + i) for i in range(len(headers_15min))]:
                                        ws_daily_balance[f'{col}{sum_row_15min}'] = f"=SUM({col}{start_row_15min+1}:{col}{sum_row_15min-1})"
                                        cell = ws_daily_balance[f'{col}{sum_row_15min}']
                                        cell.font = Font(bold=True)
                                        cell.fill = gray_fill
                                
                                # Durchschnittswerte für Prozentangaben
                                avg_row_15min = sum_row_15min + 2
                                ws_daily_balance[f'A{avg_row_15min}'] = "TAGESDURCHSCHNITT (15-MINUTEN)"
                                ws_daily_balance[f'A{avg_row_15min}'].font = Font(bold=True)
                                ws_daily_balance[f'A{avg_row_15min}'].fill = gray_fill
                                
                                # Durchschnitt für Prozentangaben berechnen
                                avg_columns_15min = ['P', 'Q', 'R']  # Autarkiegrad, Eigenverbrauchsquote, Batterieauslastung
                                for col in avg_columns_15min:
                                    if col in [chr(65 + i) for i in range(len(headers_15min))]:
                                        ws_daily_balance[f'{col}{avg_row_15min}'] = f"=AVERAGE({col}{start_row_15min+1}:{col}{sum_row_15min-1})"
                                        cell = ws_daily_balance[f'{col}{avg_row_15min}']
                                        cell.font = Font(bold=True)
                                        cell.fill = gray_fill
                                
                                # Spaltenbreiten anpassen
                                column_widths_15min = [18, 10, 15, 15, 20, 15, 18, 15, 15, 15, 15, 15, 15, 15, 22, 20, 15, 20, 18]
                                for col, width in enumerate(column_widths_15min, 1):
                                    ws_daily_balance.column_dimensions[chr(64 + col)].width = width
                                
                            else:
                                ws_daily_balance[f'A{first_day_start_row + 2}'] = "Keine 15-Minuten-Daten für den ersten Tag verfügbar."
                                ws_daily_balance[f'A{first_day_start_row + 2}'].font = Font(size=12, italic=True)
                                
                        else:
                            ws_daily_balance[f'A{first_day_start_row + 2}'] = "Keine Daten für den ersten Tag verfügbar."
                            ws_daily_balance[f'A{first_day_start_row + 2}'].font = Font(size=12, italic=True)
                            
                    except Exception as e:
                        ws_daily_balance[f'A{first_day_start_row + 2}'] = f"Fehler bei der 15-Minuten-Aufschlüsselung: {str(e)}"
                        ws_daily_balance[f'A{first_day_start_row + 2}'].font = Font(size=12, italic=True, color="FF0000")
                    
                else:
                    ws_daily_balance['A5'] = "Keine Simulationsdaten verfügbar für die Tagesbilanz."
                    ws_daily_balance['A5'].font = Font(size=12, italic=True)
                    
            except Exception as e:
                ws_daily_balance['A5'] = f"Fehler bei der Tagesbilanz: {str(e)}"
                ws_daily_balance['A5'].font = Font(size=12, italic=True, color="FF0000")
        else:
            ws_daily_balance['A5'] = "Keine Simulationsdaten verfügbar."
            ws_daily_balance['A5'].font = Font(size=12, italic=True)
        

        
        # 8. KOSTENVERGLEICH
        ws_costs = wb.create_sheet("Kostenvergleich")
        
        ws_costs['A1'] = "KOSTENVERGLEICH: BATTERIESPEICHER ZU BESTEHENDER PV-ANLAGE"
        ws_costs['A1'].font = Font(size=16, bold=True, color="FF6B35")
        ws_costs.merge_cells('A1:H1')
        ws_costs['A1'].alignment = Alignment(horizontal='center')
        
        # Erklärung
        ws_costs['A2'] = "Bewertung der Wirtschaftlichkeit eines Batteriespeichers zu einer bestehenden PV-Anlage"
        ws_costs['A2'].font = Font(size=10, italic=True)
        ws_costs.merge_cells('A2:H2')
        
        if cost_comparison_details:
            # Neue Tabelle nach Ihrem Format
            ws_costs['A4'] = "KOSTEN-/ERLÖSBESTANDTEIL"
            ws_costs['B4'] = "OHNE SPEICHER (GESAMT)"
            ws_costs['C4'] = "MIT SPEICHER (GESAMT)"
            ws_costs['D4'] = "DIFFERENZ"
            
            # Header-Styling
            for col in ['A', 'B', 'C', 'D']:
                ws_costs[f'{col}4'].font = Font(bold=True, color="FFFFFF")
                ws_costs[f'{col}4'].fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
                ws_costs[f'{col}4'].alignment = Alignment(horizontal='center')
            
            # Datenzeilen
            row = 5
            
            # Investitionskosten
            ws_costs[f'A{row}'] = "Investitionskosten"
            ws_costs[f'B{row}'] = "0 €"
            ws_costs[f'C{row}'] = f"{cost_comparison_details['investment_cost']:.0f} €"
            ws_costs[f'D{row}'] = f"+{cost_comparison_details['investment_cost']:.0f} €"
            row += 1
            
            # Laufende Betriebskosten
            maintenance_without = cost_comparison_details.get('maintenance_cost_without_battery', 0)
            maintenance_with = cost_comparison_details.get('maintenance_cost_with_battery', 0)
            ws_costs[f'A{row}'] = "Laufende Betriebskosten"
            ws_costs[f'B{row}'] = f"{maintenance_without:.0f} €/Jahr"
            ws_costs[f'C{row}'] = f"{maintenance_with:.0f} €/Jahr"
            ws_costs[f'D{row}'] = f"{maintenance_with - maintenance_without:+.0f} €/Jahr"
            row += 1
            
            # Ersatzkosten (Rücklagen)
            replacement_cost = cost_comparison_details.get('battery_replacement_cost', 0)
            replacement_year = cost_comparison_details.get('battery_replacement_year', 10)
            ws_costs[f'A{row}'] = f"Ersatzkosten (Rücklagen)"
            ws_costs[f'B{row}'] = "0 €"
            ws_costs[f'C{row}'] = f"{replacement_cost:.0f} € (nach {replacement_year} Jahren)"
            ws_costs[f'D{row}'] = f"+{replacement_cost:.0f} €"
            row += 1
            
            # Berechne Strombezugskosten und Einspeisevergütung separat
            feed_in_without = cost_comparison_details.get('feed_in_revenue_without_battery', 0)
            feed_in_with = cost_comparison_details.get('feed_in_revenue_with_battery', 0)
            
            # Strombezugskosten (reine Bezugskosten vom Netz)
            # Da annual_cost = grid_import_cost - feed_in_revenue, gilt: grid_import_cost = annual_cost + feed_in_revenue
            grid_import_cost_without = cost_comparison_details['annual_cost_no_battery'] + feed_in_without
            grid_import_cost_with = cost_comparison_details['annual_cost_with_battery'] + feed_in_with
            
            ws_costs[f'A{row}'] = "Strombezugskosten (Netz)"
            ws_costs[f'B{row}'] = f"{grid_import_cost_without:.0f} €/Jahr"
            ws_costs[f'C{row}'] = f"{grid_import_cost_with:.0f} €/Jahr"
            ws_costs[f'D{row}'] = f"{grid_import_cost_with - grid_import_cost_without:+.0f} €/Jahr"
            row += 1
            
            # Einspeisevergütung (Einnahmen - als negative Kosten dargestellt)
            ws_costs[f'A{row}'] = "Einspeisevergütung (Einnahmen)"
            ws_costs[f'B{row}'] = f"-{feed_in_without:.0f} €/Jahr"
            ws_costs[f'C{row}'] = f"-{feed_in_with:.0f} €/Jahr"
            ws_costs[f'D{row}'] = f"{-feed_in_with + feed_in_without:+.0f} €/Jahr"
            row += 1
            
            # Trennlinie
            ws_costs[f'A{row}'] = "─" * 50
            ws_costs[f'B{row}'] = "─" * 30
            ws_costs[f'C{row}'] = "─" * 30
            ws_costs[f'D{row}'] = "─" * 20
            row += 1
            
            # Jährliche Netto-Stromkosten (Bezugskosten - Einspeisevergütung)
            # annual_cost_no_battery und annual_cost_with_battery sind bereits Nettokosten
            total_without = cost_comparison_details['annual_cost_no_battery']
            total_with = cost_comparison_details['annual_cost_with_battery'] + maintenance_with
            ws_costs[f'A{row}'] = "Jährliche Netto-Stromkosten"
            ws_costs[f'B{row}'] = f"{total_without:.0f} €/Jahr"
            ws_costs[f'C{row}'] = f"{total_with:.0f} €/Jahr"
            ws_costs[f'D{row}'] = f"{total_with - total_without:+.0f} €/Jahr"
            ws_costs[f'A{row}'].font = Font(bold=True)
            row += 1
            
            # Gesamtersparnis pro Jahr - berechnet aus den Nettokosten für Konsistenz
            # WICHTIG: Berechne aus den tatsächlichen Nettokosten, nicht aus cost_comparison_details
            actual_annual_savings = total_without - total_with
            
            ws_costs[f'A{row}'] = "Gesamtersparnis pro Jahr"
            ws_costs[f'B{row}'] = "–"
            ws_costs[f'C{row}'] = f"{actual_annual_savings:.0f} €/Jahr"
            ws_costs[f'D{row}'] = f"+{actual_annual_savings:.0f} €/Jahr"
            ws_costs[f'A{row}'].font = Font(bold=True)
            ws_costs[f'C{row}'].font = Font(bold=True)
            ws_costs[f'D{row}'].font = Font(bold=True)
            row += 1
            
            # Füge detaillierte Aufschlüsselung hinzu, wenn verfügbar
            if cost_comparison_details.get('savings_from_reduced_import') is not None and cost_comparison_details.get('loss_from_reduced_export') is not None:
                # Einrückung für Sub-Items
                ws_costs[f'A{row}'] = "  ├─ Ersparnis durch reduzierten Netzbezug"
                ws_costs[f'B{row}'] = "–"
                ws_costs[f'C{row}'] = f"+{cost_comparison_details['savings_from_reduced_import']:.0f} €"
                ws_costs[f'D{row}'] = f"+{cost_comparison_details['savings_from_reduced_import']:.0f} €"
                ws_costs[f'C{row}'].font = Font(color="008000")  # Grün für positive Ersparnis
                row += 1
                
                ws_costs[f'A{row}'] = "  └─ Verlust durch reduzierte Einspeisung"
                ws_costs[f'B{row}'] = "–"
                ws_costs[f'C{row}'] = f"-{cost_comparison_details['loss_from_reduced_export']:.0f} €"
                ws_costs[f'D{row}'] = f"-{cost_comparison_details['loss_from_reduced_export']:.0f} €"
                ws_costs[f'C{row}'].font = Font(color="FF0000")  # Rot für Verlust
                row += 1
                
                # Erklärung
                ws_costs[f'A{row}'] = "Hinweis: Netto-Kostenreduktion = Ersparnis - Verlust"
                ws_costs[f'A{row}'].font = Font(size=9, italic=True, color="666666")
                ws_costs.merge_cells(f'A{row}:D{row}')
                row += 1
            
            # Leerzeile
            row += 1
            
            # Kapitalwert (NPV) - mit Zeitraum-Information
            npv_without = cost_comparison_details.get('npv_without_battery', 0)
            npv_with = cost_comparison_details.get('npv_with_battery', 0)
            project_lifetime = current_settings.get('project_lifetime_years', 20)
            
            ws_costs[f'A{row}'] = f"Kapitalwert / NPV (über {project_lifetime} Jahre)"
            ws_costs[f'B{row}'] = f"{npv_without:.0f} €"
            ws_costs[f'C{row}'] = f"{npv_with:.0f} €"
            ws_costs[f'D{row}'] = f"{npv_with - npv_without:+.0f} €"
            ws_costs[f'A{row}'].font = Font(bold=True)
            row += 1
            
            # Erklärung zum NPV
            ws_costs[f'A{row}'] = "Hinweis: NPV ohne Speicher = 0 € (Referenzszenario). NPV mit Speicher zeigt den Barwert der Investition."
            ws_costs[f'A{row}'].font = Font(size=9, italic=True, color="666666")
            ws_costs.merge_cells(f'A{row}:D{row}')
            row += 1
            
            # Detaillierte NPV-Berechnung Jahr für Jahr
            if battery_capacity_kwh > 0 and actual_annual_savings > 0:
                row += 2  # Zwei Leerzeilen
                
                ws_costs[f'A{row}'] = "DETAILLIERTE NPV-BERECHNUNG (Jahr für Jahr)"
                ws_costs[f'A{row}'].font = Font(size=12, bold=True, color="FF6B35")
                ws_costs.merge_cells(f'A{row}:F{row}')
                row += 1
                
                # Hole Parameter
                project_lifetime = current_settings.get('project_lifetime_years', 20)
                discount_rate = current_settings.get('discount_rate', 0.05)
                annual_capacity_loss = current_settings.get('annual_capacity_loss_percent', 1.0)
                investment = cost_comparison_details.get('investment_cost', 0)
                
                # Tabellen-Header
                ws_costs[f'A{row}'] = "Jahr"
                ws_costs[f'B{row}'] = "Jährl. Ersparnis (degradiert)"
                ws_costs[f'C{row}'] = "Diskontfaktor"
                ws_costs[f'D{row}'] = "Diskontierter Wert"
                ws_costs[f'E{row}'] = "Kumulierter NPV"
                ws_costs[f'F{row}'] = "Hinweis"
                
                for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                    ws_costs[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
                    ws_costs[f'{col}{row}'].fill = PatternFill(start_color="FF8C42", end_color="FF8C42", fill_type="solid")
                    ws_costs[f'{col}{row}'].alignment = Alignment(horizontal='center')
                row += 1
                
                # Jahr 0: Investition
                ws_costs[f'A{row}'] = 0
                ws_costs[f'B{row}'] = f"-{investment:.2f} €"
                ws_costs[f'C{row}'] = 1.0
                ws_costs[f'D{row}'] = f"-{investment:.2f} €"
                ws_costs[f'E{row}'] = f"-{investment:.2f} €"
                ws_costs[f'F{row}'] = "Anfangsinvestition"
                ws_costs[f'B{row}'].font = Font(color="FF0000")
                ws_costs[f'D{row}'].font = Font(color="FF0000")
                ws_costs[f'E{row}'].font = Font(color="FF0000")
                row += 1
                
                # Berechne Jahr für Jahr
                cumulative_npv = -investment
                break_even_found = False
                
                for year in range(1, project_lifetime + 1):
                    # Degradation der Ersparnisse (Batteriekapazität nimmt ab)
                    degradation_factor = (1.0 - annual_capacity_loss / 100.0) ** (year - 1)
                    year_savings = actual_annual_savings * degradation_factor
                    
                    # Diskontfaktor: 1 / (1 + r)^t
                    discount_factor = 1 / ((1 + discount_rate) ** year)
                    
                    # Diskontierter Wert = Ersparnis × Diskontfaktor
                    discounted_value = year_savings * discount_factor
                    
                    # Kumulierter NPV
                    cumulative_npv += discounted_value
                    
                    # In Excel schreiben
                    ws_costs[f'A{row}'] = year
                    ws_costs[f'B{row}'] = f"{year_savings:.2f} €"
                    ws_costs[f'C{row}'] = f"{discount_factor:.4f}"
                    ws_costs[f'D{row}'] = f"{discounted_value:.2f} €"
                    ws_costs[f'E{row}'] = f"{cumulative_npv:.2f} €"
                    
                    # Erklärung für spezielle Jahre
                    if year == 1:
                        ws_costs[f'F{row}'] = "Erstes Betriebsjahr"
                    elif cumulative_npv >= 0 and not break_even_found:
                        ws_costs[f'F{row}'] = "✓ Break-even erreicht!"
                        ws_costs[f'E{row}'].font = Font(bold=True, color="008000")
                        break_even_found = True
                    elif year == project_lifetime:
                        ws_costs[f'F{row}'] = "Ende Projektlaufzeit"
                        ws_costs[f'E{row}'].font = Font(bold=True, color="FF6B35")
                    else:
                        ws_costs[f'F{row}'] = ""
                    
                    # Farbcodierung für kumulierten NPV
                    if cumulative_npv >= 0 and not (year == project_lifetime):
                        ws_costs[f'E{row}'].font = Font(color="008000")  # Grün für positive Werte
                    elif cumulative_npv < 0:
                        ws_costs[f'E{row}'].font = Font(color="FF0000")  # Rot für negative Werte
                    
                    row += 1
                
                # Zusammenfassung
                ws_costs[f'A{row}'] = "ENDERGEBNIS"
                ws_costs[f'A{row}'].font = Font(bold=True, color="FF6B35")
                ws_costs[f'E{row}'] = f"{cumulative_npv:.2f} €"
                ws_costs[f'E{row}'].font = Font(bold=True, size=12, color="FF6B35" if cumulative_npv > 0 else "FF0000")
                ws_costs[f'F{row}'] = f"Maximaler NPV über {project_lifetime} Jahre (sollte mit Zusammenfassung übereinstimmen)"
                ws_costs[f'F{row}'].font = Font(italic=True, size=9)
                row += 1
                
                # Formel-Erklärung
                row += 1
                ws_costs[f'A{row}'] = "Formel: NPV = -Investition + Σ(Jährliche Ersparnis × Degradationsfaktor × Diskontfaktor)"
                ws_costs[f'A{row}'].font = Font(size=9, italic=True, color="666666")
                ws_costs.merge_cells(f'A{row}:F{row}')
                row += 1
                
                ws_costs[f'A{row}'] = f"Degradationsfaktor = (1 - {annual_capacity_loss}%)^(Jahr-1),  Diskontfaktor = 1/(1+{discount_rate:.1%})^Jahr"
                ws_costs[f'A{row}'].font = Font(size=9, italic=True, color="666666")
                ws_costs.merge_cells(f'A{row}:F{row}')
                row += 1
                
                # Spaltenbreiten für NPV-Tabelle anpassen
                ws_costs.column_dimensions['B'].width = 28
                ws_costs.column_dimensions['C'].width = 18
                ws_costs.column_dimensions['D'].width = 22
                ws_costs.column_dimensions['E'].width = 22
                ws_costs.column_dimensions['F'].width = 40
            
            # Leerzeile
            row += 1
            
            # IRR - Verwende IRR aus cost_comparison_details oder aus optimization_results
            irr = cost_comparison_details.get('irr_percentage', np.nan)
            
            # Fallback: Suche IRR in optimization_results
            if np.isnan(irr) and optimization_results:
                try:
                    df_results = pd.DataFrame(optimization_results)
                    # Verwende das gleiche Optimierungskriterium wie in der UI
                    optimization_criterion_irr = current_settings.get('optimization_criterion', 'Deckungsbeitrag III gesamt (Barwert)')
                    # Nur DB Barwert und DB Nominal unterstützen
                    if optimization_criterion_irr == "Deckungsbeitrag III gesamt (Nominal)":
                        best_idx_irr = df_results['total_db3_nominal'].idxmax()
                    else:  # Default: Deckungsbeitrag III gesamt (Barwert)
                        best_idx_irr = df_results['total_db3_present_value'].idxmax()
                    irr = df_results.loc[best_idx_irr, 'irr_percentage']
                except:
                    irr = np.nan
            
            ws_costs[f'A{row}'] = "IRR (Interner Zinsfuß)"
            ws_costs[f'B{row}'] = "–"
            if not np.isnan(irr):
                ws_costs[f'C{row}'] = f"{irr:.2f} %"
                ws_costs[f'D{row}'] = f"+{irr:.2f} %"
            else:
                ws_costs[f'C{row}'] = "Nicht berechenbar"
                ws_costs[f'D{row}'] = "–"
            row += 1
            
            # Detaillierte IRR-Berechnung
            if not np.isnan(irr) and battery_capacity_kwh > 0 and actual_annual_savings > 0:
                row += 2
                
                ws_costs[f'A{row}'] = "DETAILLIERTE IRR-BERECHNUNG (Interner Zinsfuß)"
                ws_costs[f'A{row}'].font = Font(size=12, bold=True, color="FF6B35")
                ws_costs.merge_cells(f'A{row}:F{row}')
                row += 1
                
                # Erklärung
                ws_costs[f'A{row}'] = f"Der IRR ist der Zinssatz r, bei dem NPV = 0. Formel: NPV = -I₀ + Σ(CFₜ × (1-d)^(t-1) / (1+r)^t) = 0"
                ws_costs[f'A{row}'].font = Font(size=9, italic=True, color="666666")
                ws_costs.merge_cells(f'A{row}:F{row}')
                row += 1
                
                # Warum iterativ?
                ws_costs[f'A{row}'] = "Warum iterative Methode statt analytische Lösung?"
                ws_costs[f'A{row}'].font = Font(size=10, bold=True, color="FF6B35")
                ws_costs.merge_cells(f'A{row}:F{row}')
                row += 1
                
                ws_costs[f'A{row}'] = "• Analytische Lösung (PQ-Formel): Funktioniert nur bei 2-3 Jahren ohne Degradation"
                ws_costs[f'A{row}'].font = Font(size=9, italic=True, color="666666")
                ws_costs.merge_cells(f'A{row}:F{row}')
                row += 1
                
                ws_costs[f'A{row}'] = f"• Batteriespeicher: {project_lifetime} Jahre + jährliche Degradation → Gleichung zu komplex für PQ-Formel"
                ws_costs[f'A{row}'].font = Font(size=9, italic=True, color="666666")
                ws_costs.merge_cells(f'A{row}:F{row}')
                row += 1
                
                ws_costs[f'A{row}'] = "• Newton-Raphson Iteration: Standardverfahren für komplexe IRR-Berechnungen"
                ws_costs[f'A{row}'].font = Font(size=9, italic=True, color="666666")
                ws_costs.merge_cells(f'A{row}:F{row}')
                row += 1
                
                ws_costs[f'A{row}'] = "Iterative Annäherung durch systematisches Probieren verschiedener Zinssätze:"
                ws_costs[f'A{row}'].font = Font(size=9, italic=True)
                ws_costs.merge_cells(f'A{row}:F{row}')
                row += 1
                
                # Tabelle Header
                ws_costs[f'A{row}'] = "Iteration"
                ws_costs[f'B{row}'] = "Test-Zinssatz"
                ws_costs[f'C{row}'] = "NPV bei diesem Zinssatz"
                ws_costs[f'D{row}'] = "Abweichung von 0"
                ws_costs[f'E{row}'] = "Bewertung"
                ws_costs[f'F{row}'] = "Nächster Schritt"
                
                for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                    ws_costs[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
                    ws_costs[f'{col}{row}'].fill = PatternFill(start_color="FF8C42", end_color="FF8C42", fill_type="solid")
                    ws_costs[f'{col}{row}'].alignment = Alignment(horizontal='center')
                row += 1
                
                # Funktion zur NPV-Berechnung mit beliebigem Zinssatz
                def calculate_npv_for_rate(rate_percent):
                    rate = rate_percent / 100.0
                    npv = -investment
                    for year in range(1, project_lifetime + 1):
                        degradation_factor = (1.0 - annual_capacity_loss / 100.0) ** (year - 1)
                        year_savings = actual_annual_savings * degradation_factor
                        npv += year_savings / ((1 + rate) ** year)
                    return npv
                
                # Probiere verschiedene Zinssätze
                test_rates = [0, 2, 5, round(irr * 0.5, 2), round(irr * 0.9, 2), round(irr * 0.99, 2), irr, round(irr * 1.01, 2), round(irr * 1.5, 2), 10]
                test_rates = sorted(list(set([r for r in test_rates if r >= 0])))[:10]  # Max 10 Iterationen
                
                iteration = 1
                for test_rate in test_rates:
                    test_npv = calculate_npv_for_rate(test_rate)
                    abweichung = abs(test_npv)
                    
                    # Bewertung
                    if abs(test_npv) < 50:
                        bewertung = "✓ NPV ≈ 0"
                        color = "008000"
                        naechster_schritt = "IRR gefunden!"
                    elif test_npv > 0:
                        bewertung = "NPV > 0"
                        color = "FF6B35"
                        naechster_schritt = "Zinssatz erhöhen"
                    else:
                        bewertung = "NPV < 0"
                        color = "FF0000"
                        naechster_schritt = "Zinssatz senken"
                    
                    # In Excel schreiben
                    ws_costs[f'A{row}'] = iteration
                    ws_costs[f'B{row}'] = f"{test_rate:.2f}"
                    ws_costs[f'B{row}'].number_format = '0.00'
                    ws_costs[f'C{row}'] = test_npv
                    ws_costs[f'C{row}'].number_format = '#,##0.00'
                    ws_costs[f'D{row}'] = abweichung
                    ws_costs[f'D{row}'].number_format = '#,##0.00'
                    ws_costs[f'E{row}'] = bewertung
                    ws_costs[f'F{row}'] = naechster_schritt
                    
                    # Farbcodierung
                    ws_costs[f'C{row}'].font = Font(color=color)
                    ws_costs[f'E{row}'].font = Font(color=color)
                    if abs(test_npv) < 50:
                        ws_costs[f'B{row}'].font = Font(bold=True, color="008000")
                        ws_costs[f'F{row}'].font = Font(bold=True, color="008000")
                    
                    row += 1
                    iteration += 1
                
                # Endergebnis
                ws_costs[f'A{row}'] = "ERGEBNIS"
                ws_costs[f'A{row}'].font = Font(bold=True, color="FF6B35")
                ws_costs[f'B{row}'] = f"IRR = {irr:.2f} %"
                ws_costs[f'B{row}'].font = Font(bold=True, size=12, color="008000")
                ws_costs.merge_cells(f'B{row}:D{row}')
                ws_costs[f'F{row}'] = f"NPV = 0 bei diesem Zinssatz"
                ws_costs[f'F{row}'].font = Font(size=9, italic=True)
                row += 1
                
                # Interpretation
                row += 1
                ws_costs[f'A{row}'] = f"Interpretation: IRR = {irr:.2f}% bedeutet, dass die Investition eine jährliche Rendite von {irr:.2f}% erwirtschaftet."
                ws_costs[f'A{row}'].font = Font(size=9, italic=True, color="666666")
                ws_costs.merge_cells(f'A{row}:F{row}')
                row += 1
                
                vergleich_mit_diskont = "✓ liegt ÜBER" if irr > discount_rate * 100 else "✗ liegt UNTER"
                ws_costs[f'A{row}'] = f"{vergleich_mit_diskont} dem Diskontierungszinssatz ({discount_rate:.1%}) → {'Investition lohnenswert!' if irr > discount_rate * 100 else 'Investition fragwürdig.'}"
                ws_costs[f'A{row}'].font = Font(size=9, italic=True, color="008000" if irr > discount_rate * 100 else "FF0000")
                ws_costs.merge_cells(f'A{row}:F{row}')
                row += 1
                
                # Formel-Erklärung
                row += 1
                ws_costs[f'A{row}'] = "Berechnungsmethode: Newton-Raphson Iteration probiert systematisch verschiedene Zinssätze, bis NPV ≈ 0."
                ws_costs[f'A{row}'].font = Font(size=9, italic=True, color="666666")
                ws_costs.merge_cells(f'A{row}:F{row}')
                row += 1
                
                # Vergleich mit einfachem Fall
                ws_costs[f'A{row}'] = "Vergleich: Einfacher Fall (2 Jahre, keine Degradation) → PQ-Formel möglich"
                ws_costs[f'A{row}'].font = Font(size=9, italic=True, color="666666")
                ws_costs.merge_cells(f'A{row}:F{row}')
                row += 1
                
                ws_costs[f'A{row}'] = f"Unser Fall: {project_lifetime} Jahre + Degradation → Iterative Lösung erforderlich"
                ws_costs[f'A{row}'].font = Font(size=9, italic=True, color="666666")
                ws_costs.merge_cells(f'A{row}:F{row}')
                row += 1
                
                # IRR-Kontrolle: Kombinierte Darstellung wie im Bild
                row += 2
                ws_costs[f'A{row}'] = "DETAILLIERTE NPV-BERECHNUNG (Jahr für Jahr)"
                ws_costs[f'A{row}'].font = Font(size=12, bold=True, color="FF6B35")
                ws_costs.merge_cells(f'A{row}:G{row}')
                row += 1
                
                # Tabelle Header wie im Bild
                ws_costs[f'A{row}'] = "Jahr"
                ws_costs[f'B{row}'] = "Jährl. Ersparnis (degradiert)"
                ws_costs[f'C{row}'] = "Diskontfaktor"
                ws_costs[f'D{row}'] = "Diskontierter Wert"
                ws_costs[f'E{row}'] = "Kumulierter NPV"
                ws_costs[f'F{row}'] = "Berechnung"
                ws_costs[f'G{row}'] = "Hinweis"
                
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                    ws_costs[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
                    ws_costs[f'{col}{row}'].fill = PatternFill(start_color="FF8C42", end_color="FF8C42", fill_type="solid")
                    ws_costs[f'{col}{row}'].alignment = Alignment(horizontal='center')
                row += 1
                
                # Jahr 0: Anfangsinvestition
                ws_costs[f'A{row}'] = 0
                ws_costs[f'B{row}'] = -investment
                ws_costs[f'B{row}'].number_format = '#,##0.00'
                ws_costs[f'C{row}'] = 1.0000
                ws_costs[f'C{row}'].number_format = '0.0000'
                ws_costs[f'D{row}'] = -investment
                ws_costs[f'D{row}'].number_format = '#,##0.00'
                ws_costs[f'E{row}'] = -investment
                ws_costs[f'E{row}'].number_format = '#,##0.00'
                ws_costs[f'F{row}'] = f"1/(1+{irr:.2f}%)^0 = 1,0000"
                ws_costs[f'G{row}'] = "Anfangsinvestition"
                ws_costs[f'B{row}'].font = Font(color="FF0000")
                ws_costs[f'D{row}'].font = Font(color="FF0000")
                ws_costs[f'E{row}'].font = Font(color="FF0000")
                row += 1
                
                # Jahre 1-15: Ersparnisse mit Degradation und Diskontierung
                irr_rate = irr / 100.0
                cumulative_npv = -investment
                
                for year in range(1, project_lifetime + 1):
                    # Ersparnis mit Degradation
                    degradation_factor = (1.0 - annual_capacity_loss / 100.0) ** (year - 1)
                    degraded_savings = actual_annual_savings * degradation_factor
                    
                    # Diskontierungsfaktor
                    discount_factor = 1 / ((1 + irr_rate) ** year)
                    
                    # Diskontierter Wert
                    discounted_value = degraded_savings * discount_factor
                    cumulative_npv += discounted_value
                    
                    # Hinweis bestimmen
                    if year == 1:
                        hinweis = "Erstes Betriebsjahr"
                    elif cumulative_npv > 0 and year > 1:
                        hinweis = "✓ Break-even erreicht!"
                    elif year == project_lifetime:
                        hinweis = "Ende Projektlaufzeit"
                    else:
                        hinweis = ""
                    
                    # In Excel schreiben
                    ws_costs[f'A{row}'] = year
                    ws_costs[f'B{row}'] = degraded_savings
                    ws_costs[f'B{row}'].number_format = '#,##0.00'
                    ws_costs[f'C{row}'] = discount_factor
                    ws_costs[f'C{row}'].number_format = '0.0000'
                    ws_costs[f'D{row}'] = discounted_value
                    ws_costs[f'D{row}'].number_format = '#,##0.00'
                    ws_costs[f'E{row}'] = cumulative_npv
                    ws_costs[f'E{row}'].number_format = '#,##0.00'
                    ws_costs[f'F{row}'] = f"1/(1+{irr:.2f}%)^{year} = {discount_factor:.4f}"
                    ws_costs[f'G{row}'] = hinweis
                    
                    # Farbcodierung
                    if cumulative_npv < 0:
                        ws_costs[f'E{row}'].font = Font(color="FF0000")
                    else:
                        ws_costs[f'E{row}'].font = Font(color="008000")
                    
                    # Spezielle Markierung für Break-even und letztes Jahr
                    if "Break-even" in hinweis:
                        ws_costs[f'A{row}'].font = Font(bold=True, color="FF6B35")
                        ws_costs[f'E{row}'].font = Font(bold=True, color="008000")
                    elif year == project_lifetime:
                        ws_costs[f'A{row}'].font = Font(bold=True, color="FF6B35")
                        ws_costs[f'E{row}'].font = Font(bold=True, color="008000")
                    
                    row += 1
                
                # Endergebnis
                ws_costs[f'A{row}'] = "ENDERGEBNIS"
                ws_costs[f'A{row}'].font = Font(bold=True, color="FF6B35")
                ws_costs[f'E{row}'] = cumulative_npv
                ws_costs[f'E{row}'].number_format = '#,##0.00'
                ws_costs[f'E{row}'].font = Font(bold=True, color="008000" if abs(cumulative_npv) < 50 else "FF0000")
                ws_costs[f'F{row}'] = f"Finaler NPV = 0 (IRR-Kontrolle)"
                ws_costs[f'G{row}'] = f"Maximaler NPV über {project_lifetime} Jahre (sollte mit Zusammenfassung übereinstimmen)"
                row += 1
                
                # Formel-Erklärung
                ws_costs[f'A{row}'] = f"Formel: NPV = -Investition + Σ(Jährliche Ersparnis × Degradationsfaktor × Diskontfaktor)"
                ws_costs[f'A{row}'].font = Font(size=9, italic=True, color="666666")
                ws_costs.merge_cells(f'A{row}:G{row}')
                row += 1
                
                ws_costs[f'A{row}'] = f"Degradationsfaktor = (1 - {annual_capacity_loss}%)^(Jahr-1), Diskontfaktor = 1/(1+{irr:.2f}%)^Jahr"
                ws_costs[f'A{row}'].font = Font(size=9, italic=True, color="666666")
                ws_costs.merge_cells(f'A{row}:G{row}')
                row += 1
                
                # Erklärung der Berechnung
                ws_costs[f'A{row}'] = "Berechnung pro Zeile:"
                ws_costs[f'A{row}'].font = Font(size=10, bold=True, color="FF6B35")
                ws_costs.merge_cells(f'A{row}:G{row}')
                row += 1
                
                ws_costs[f'A{row}'] = "• Jahr 0: Anfangsinvestition (negativ)"
                ws_costs[f'A{row}'].font = Font(size=9, italic=True, color="666666")
                ws_costs.merge_cells(f'A{row}:G{row}')
                row += 1
                
                ws_costs[f'A{row}'] = f"• Jahr 1-{project_lifetime}: Jährl. Ersparnis × Degradationsfaktor × Diskontfaktor"
                ws_costs[f'A{row}'].font = Font(size=9, italic=True, color="666666")
                ws_costs.merge_cells(f'A{row}:G{row}')
                row += 1
                
                # Berechne den tatsächlichen Diskontfaktor für Jahr 1
                actual_discount_factor_year1 = 1 / ((1 + irr/100.0) ** 1)
                ws_costs[f'A{row}'] = f"• Beispiel Jahr 1: {actual_annual_savings:.0f} € × 1,0000 × {actual_discount_factor_year1:.4f} = {actual_annual_savings * actual_discount_factor_year1:.0f} €"
                ws_costs[f'A{row}'].font = Font(size=9, italic=True, color="666666")
                ws_costs.merge_cells(f'A{row}:G{row}')
                row += 1
                
                ws_costs[f'A{row}'] = "• Kumulierter NPV: Summe aller vorherigen Barwerte + aktueller Barwert"
                ws_costs[f'A{row}'].font = Font(size=9, italic=True, color="666666")
                ws_costs.merge_cells(f'A{row}:G{row}')
                row += 1
                
                ws_costs[f'A{row}'] = f"• Ziel: Bei IRR = {irr:.2f}% sollte der finale NPV ≈ 0 € sein (Kontrolle der IRR-Berechnung)"
                ws_costs[f'A{row}'].font = Font(size=9, italic=True, color="666666")
                ws_costs.merge_cells(f'A{row}:G{row}')
            
            # Spaltenbreiten anpassen
            ws_costs.column_dimensions['A'].width = 8
            ws_costs.column_dimensions['B'].width = 25
            ws_costs.column_dimensions['C'].width = 15
            ws_costs.column_dimensions['D'].width = 20
            ws_costs.column_dimensions['E'].width = 20
            ws_costs.column_dimensions['F'].width = 35
            ws_costs.column_dimensions['G'].width = 30
        
        # 8. AMORTISATIONSZEIT
        create_amortisationszeit_sheet(wb, optimization_results, optimal_sim_result, cost_comparison_details, current_settings, battery_cost_curve, results_summary)
        
        # 9. DECKUNGSBEITRAGSRECHNUNG
        create_deckungsbeitragsrechnung_sheet(wb, optimization_results, optimal_sim_result, cost_comparison_details, current_settings, battery_cost_curve, results_summary)
        
        # 10. PV-ERZEUGUNGSDATEN
        ws_pv = wb.create_sheet("PV-Erzeugungsdaten")
        
        ws_pv['A1'] = "PV-ERZEUGUNGSDATEN - DETAILLIERTE AUFSCHLÜSSELUNG"
        ws_pv['A1'].font = Font(size=16, bold=True, color="FF6B35")
        ws_pv.merge_cells('A1:H1')
        ws_pv['A1'].alignment = Alignment(horizontal='center')
        
        if pv_generation_series is not None and len(pv_generation_series) > 0:
            # Erstelle PV-Aufschlüsselung
            pv_breakdown = create_pv_generation_breakdown(
                pv_generation_series, 
                current_settings.get('pv_system_size_kwp', 10.0),
                current_settings.get('selected_year', 2024)
            )
            
            # Jahresstatistiken
            if pv_breakdown['annual_stats']:
                stats = pv_breakdown['annual_stats']
                ws_pv['A3'] = "JAHRESSTATISTIKEN"
                ws_pv['A3'].font = header_font
                ws_pv['A3'].fill = orange_fill
                ws_pv['A3'].font = white_font
                
                annual_data = [
                    ["Kennzahl", "Wert", "Einheit", "Beschreibung"],
                    ["Gesamter Jahresertrag", f"{stats.get('total_energy_kwh', 0):.2f}", "kWh", "Gesamte PV-Erzeugung im Jahr"],
                    ["Spezifischer Ertrag", f"{stats.get('specific_yield_kwh_kwp', 0):.2f}", "kWh/kWp", "Ertrag pro kWp installierte Leistung"],
                    ["Spitzenleistung", f"{stats.get('peak_power_kw', 0):.2f}", "kW", "Maximale PV-Leistung im Jahr"],
                    ["Durchschnittlicher Tagesertrag", f"{stats.get('avg_daily_energy_kwh', 0):.2f}", "kWh/Tag", "Durchschnittlicher Tagesertrag"],
                    ["Tage mit Erzeugung", f"{stats.get('days_with_generation', 0)}", "Tage", "Anzahl Tage mit PV-Erzeugung > 0"]
                ]
                
                for i, row in enumerate(annual_data):
                    for j, value in enumerate(row):
                        cell = ws_pv.cell(row=5+i, column=1+j)
                        cell.value = value
                        if i == 0:  # Header
                            cell.font = black_font
                            cell.fill = light_orange_fill
                        else:
                            cell.font = Font(size=10)
                
                current_row = 12
            else:
                current_row = 5
            
            # Monatsaufschlüsselung
            if not pv_breakdown['monthly_breakdown'].empty:
                ws_pv[f'A{current_row}'] = "MONATSAUFSCHLÜSSELUNG"
                ws_pv[f'A{current_row}'].font = header_font
                ws_pv[f'A{current_row}'].fill = orange_fill
                ws_pv[f'A{current_row}'].font = white_font
                current_row += 2
                
                # Header für Monatsdaten
                monthly_headers = list(pv_breakdown['monthly_breakdown'].columns)
                for j, header in enumerate(monthly_headers):
                    cell = ws_pv.cell(row=current_row, column=1+j)
                    cell.value = header
                    cell.font = black_font
                    cell.fill = light_orange_fill
                
                current_row += 1
                
                # Monatsdaten
                for _, row in pv_breakdown['monthly_breakdown'].iterrows():
                    for j, value in enumerate(row):
                        cell = ws_pv.cell(row=current_row, column=1+j)
                        cell.value = value
                        cell.font = Font(size=10)
                    current_row += 1
                
                current_row += 2
            
            # Tagesaufschlüsselung (alle Tage des Jahres)
            if not pv_breakdown['daily_breakdown'].empty:
                ws_pv[f'A{current_row}'] = "TAGESAUFSCHLÜSSELUNG (alle Tage des Jahres)"
                ws_pv[f'A{current_row}'].font = header_font
                ws_pv[f'A{current_row}'].fill = orange_fill
                ws_pv[f'A{current_row}'].font = white_font
                current_row += 2
                
                # Header für Tagesdaten
                daily_headers = list(pv_breakdown['daily_breakdown'].columns)
                for j, header in enumerate(daily_headers):
                    cell = ws_pv.cell(row=current_row, column=1+j)
                    cell.value = header
                    cell.font = black_font
                    cell.fill = light_orange_fill
                
                current_row += 1
                
                # Tagesdaten (alle Tage des Jahres)
                for _, row in pv_breakdown['daily_breakdown'].iterrows():
                    for j, value in enumerate(row):
                        cell = ws_pv.cell(row=current_row, column=1+j)
                        cell.value = value
                        cell.font = Font(size=10)
                    current_row += 1
        else:
            ws_pv['A3'] = "Keine PV-Erzeugungsdaten verfügbar"
            ws_pv['A3'].font = Font(size=12, italic=True)
        
        # 10. KAPAZITÄTSALTERUNG
        if current_settings and 'annual_capacity_loss_percent' in current_settings:
            try:
                # Finde die optimale Batteriekapazität basierend auf dem Optimierungskriterium
                if optimization_results:
                    df_results = pd.DataFrame(optimization_results)
                    # Verwende das gleiche Optimierungskriterium wie in der UI
                    optimization_criterion_cap = current_settings.get('optimization_criterion', 'Deckungsbeitrag III gesamt (Barwert)')
                    # Nur DB Barwert und DB Nominal unterstützen
                    if optimization_criterion_cap == "Deckungsbeitrag III gesamt (Nominal)":
                        best_idx_cap = df_results['total_db3_nominal'].idxmax()
                    else:  # Default: Deckungsbeitrag III gesamt (Barwert)
                        best_idx_cap = df_results['total_db3_present_value'].idxmax()
                    battery_capacity = df_results.loc[best_idx_cap, 'battery_capacity_kwh']
                else:
                    battery_capacity = 10.0  # Fallback
                
                # Nur Kapazitätsalterung anzeigen, wenn eine Batterie vorhanden ist
                if battery_capacity > 0:
                    annual_loss = current_settings.get('annual_capacity_loss_percent', 2.0)
                    project_lifetime = current_settings.get('project_lifetime_years', 20)
                    
                    create_capacity_aging_sheet(wb, battery_capacity, annual_loss, project_lifetime)
                else:
                    # Erstelle ein informatives Sheet für den Fall ohne Batterie
                    ws_capacity = wb.create_sheet("Kapazitätsalterung")
                    ws_capacity['A1'] = "Kapazitätsalterung - Keine Batterie"
                    ws_capacity['A1'].font = Font(size=16, bold=True, color="FF6B35")
                    ws_capacity.merge_cells('A1:D1')
                    ws_capacity['A1'].alignment = Alignment(horizontal='center')
                    
                    ws_capacity['A3'] = "Keine Batterie konfiguriert"
                    ws_capacity['A3'].font = Font(size=12, italic=True)
                    ws_capacity['A4'] = "Die Kapazitätsalterung wird nur für Systeme mit Batteriespeicher berechnet."
                    
            except Exception as e:
                st.error(f"Fehler beim Erstellen des Kapazitätsalterung-Sheets: {str(e)}")
        
        # 11. GRAFIKEN
        ws_charts = wb.create_sheet("Grafiken")
        
        ws_charts['A1'] = "GRAFIKEN UND VISUALISIERUNGEN"
        ws_charts['A1'].font = Font(size=16, bold=True, color="FF6B35")
        ws_charts.merge_cells('A1:H1')
        ws_charts['A1'].alignment = Alignment(horizontal='center')
        
        # Alle Grafiken verwenden jetzt PNG-Export-Funktionen für bessere Darstellung
        
        current_row = 3
        
        # 1. Optimierungskurve - Verwende PNG-Export für bessere Darstellung
        if optimization_results:
            try:
                # Verwende die neue PNG-Funktion für bessere Darstellung
                opt_img = create_optimization_curve_png(optimization_results, current_settings)
                if opt_img:
                    current_row = add_chart_to_excel(ws_charts, opt_img, current_row, 0, 
                                                   "Optimierungskurve - Batteriespeichergröße")
            except Exception as e:
                st.error(f"Fehler bei Optimierungskurve: {str(e)}")
        
        # 2. Energieflüsse über die Zeit - ENTFERNT (nicht mehr benötigt)
        # Diese Grafik wird nicht mehr in die Excel-Datei eingefügt
        
        # 2b. Monatsgrafiken (falls verfügbar) - Verwende plot_energy_flows_for_period wie in der UI
        if optimal_sim_result and "time_series_data" in optimal_sim_result:
            try:
                # Bestimme das Jahr aus den Daten
                year = optimal_sim_result["time_series_data"].index.year[0]
                
                # Verwende die tatsächliche optimale Batteriekapazität
                optimal_capacity = current_settings.get('optimal_battery_capacity_kwh', 10.0) if current_settings else 10.0
                
                # Aktualisiere current_settings mit der optimalen Kapazität
                settings_with_capacity = current_settings.copy() if current_settings else {}
                settings_with_capacity['optimal_battery_capacity_kwh'] = optimal_capacity
                
                # Monatsnamen für bessere Ausgabe
                import calendar
                month_names_german = {
                    1: 'Januar', 2: 'Februar', 3: 'März', 4: 'April',
                    5: 'Mai', 6: 'Juni', 7: 'Juli', 8: 'August',
                    9: 'September', 10: 'Oktober', 11: 'November', 12: 'Dezember'
                }
                
                # Erstelle für jeden Monat eine Grafik - GLEICHE Funktion wie in der UI
                for month in range(1, 13):
                    try:
                        month_img = create_monthly_energy_flow_png(
                            optimal_sim_result["time_series_data"], 
                            year, 
                            month, 
                            settings_with_capacity
                        )
                        
                        if month_img:
                            month_name = month_names_german[month]
                            current_row = add_chart_to_excel(ws_charts, month_img, current_row, 0, 
                                                           f"Energieflüsse - {month_name} {year}")
                        else:
                            print(f"Keine Bilddaten für Monat {month} erhalten")
                            
                    except Exception as img_error:
                        print(f"Fehler beim PNG-Export für Monat {month}: {img_error}")
                        import traceback
                        traceback.print_exc()
            except Exception as e:
                st.error(f"Fehler bei Monatsgrafiken: {str(e)}")
        
        # 3. Sankey-Diagramm (mit Batterie) - Verwende PNG-Export für bessere Darstellung
        if optimal_sim_result and "kpis" in optimal_sim_result:
            try:
                # Verwende die neue PNG-Funktion für bessere Darstellung
                # Sankey-Diagramm MIT Batterie
                sankey_with_img = create_sankey_diagram_png(optimal_sim_result["kpis"])
                
                # Sankey-Diagramm OHNE Batterie
                sankey_without_img = None
                if cost_comparison_details:
                    kpis_no_battery = {
                        "total_grid_import_kwh": cost_comparison_details.get('grid_import_no_battery', 0),
                        "total_grid_export_kwh": cost_comparison_details.get('grid_export_no_battery', 0),
                        "total_direct_self_consumption_kwh": cost_comparison_details.get('self_consumption_no_battery', 0)
                    }
                    total_consumption = optimal_sim_result["kpis"].get("total_consumption_kwh", 0)
                    total_pv_generation = optimal_sim_result["kpis"].get("total_pv_generation_kwh", 0)
                    sankey_without_img = create_sankey_diagram_no_battery_png(kpis_no_battery, total_consumption, total_pv_generation)
                
                # Füge beide Diagramme nebeneinander ein
                if sankey_without_img:
                    current_row = add_two_charts_side_by_side(ws_charts, sankey_without_img, sankey_with_img, 
                                                             current_row, "Ohne Batteriespeicher", "Mit Batteriespeicher")
                    
                    # Füge Vergleichstabelle unter den Sankey-Diagrammen hinzu
                    current_row = add_sankey_comparison_table(
                        ws_charts, 
                        current_row,
                        kpis_no_battery,
                        optimal_sim_result["kpis"],
                        total_consumption,
                        total_pv_generation
                    )
                elif sankey_with_img:
                    current_row = add_chart_to_excel(ws_charts, sankey_with_img, current_row, 0, 
                                                   "Jährliche Energieflüsse (Sankey-Diagramm - Mit Batterie)")
            except Exception as e:
                st.error(f"Fehler bei Sankey-Diagramm: {str(e)}")
        
        # 4. Kostenvergleich - Verwende PNG-Export für bessere Darstellung
        if cost_comparison_details and optimal_sim_result:
            try:
                # Verwende die neue PNG-Funktion für bessere Darstellung
                cost_img = create_cost_comparison_png(optimal_sim_result, current_settings, optimization_results, consumption_series, pv_generation_series, battery_cost_curve)
                if cost_img:
                    current_row = add_chart_to_excel(ws_charts, cost_img, current_row, 0, 
                                                   "Kostenvergleich: Mit vs. Ohne Batteriespeicher")
            except Exception as e:
                st.error(f"Fehler bei Kostenvergleich: {str(e)}")
        
        # Spaltenbreiten anpassen
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = None
                for cell in column:
                    try:
                        # Prüfe ob es eine merged cell ist
                        if hasattr(cell, 'column_letter'):
                            if column_letter is None:
                                column_letter = cell.column_letter
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                    except:
                        pass
                
                # Nur anpassen wenn column_letter gefunden wurde
                if column_letter:
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
        
        # Excel speichern
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Lösche alle temporären Dateien nach dem Speichern
        for ws in wb.worksheets:
            if hasattr(ws, '_temp_files'):
                for temp_file in ws._temp_files:
                    try:
                        if os.path.exists(temp_file):
                            os.unlink(temp_file)
                    except:
                        pass
        
        return buffer
        
    except Exception as e:
        st.error(f"❌ Fehler beim Erstellen der Excel: {str(e)}")
        return None

def excel_export_section(optimization_results, optimal_sim_result=None, variable_tariff_result=None, 
                        current_settings=None, calculation_details=None, cost_comparison_details=None,
                        consumption_series=None, pv_generation_series=None, battery_cost_curve=None,
                        results_summary=None):
    """
    Excel Export Sektion für die UI - EINFACH UND DIREKT
    """
    
    # Prüfe ob Daten verfügbar sind
    if not optimization_results:
        st.error("❌ Keine Daten verfügbar")
        st.info("Führen Sie zuerst eine vollständige Analyse durch.")
        return
    
    progress_placeholder = st.progress(0, text="Excel-Bericht wird vorbereitet …")
    
    if not st.session_state.get('excel_created', False):
        try:
            progress_placeholder.progress(30, text="Excel-Bericht wird erstellt …")
            buffer = create_comprehensive_excel(
                optimization_results=optimization_results,
                optimal_sim_result=optimal_sim_result,
                variable_tariff_result=variable_tariff_result,
                current_settings=current_settings,
                calculation_details=calculation_details,
                cost_comparison_details=cost_comparison_details,
                consumption_series=consumption_series,
                pv_generation_series=pv_generation_series,
                battery_cost_curve=battery_cost_curve,
                results_summary=results_summary
            )
            
            if buffer:
                st.session_state['excel_buffer'] = buffer.getvalue()
                st.session_state['excel_created'] = True
                st.session_state['excel_created_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                progress_placeholder.progress(100, text="Excel-Bericht ist bereit zum Download")
            else:
                progress_placeholder.empty()
                st.error("❌ Fehler beim automatischen Erstellen des Excel-Berichts")
                
        except Exception as e:
            progress_placeholder.empty()
            st.error(f"❌ **Fehler beim automatischen Erstellen des Excel-Berichts:** {str(e)}")
            st.error(f"**Fehlerdetails:** {type(e).__name__}: {str(e)}")
    else:
        progress_placeholder.progress(100, text="Excel-Bericht ist bereit zum Download")
    
    if st.session_state.get('excel_created', False):
        excel_data = st.session_state.get('excel_buffer')
        if excel_data:
            st.download_button(
                label="📥 Excel-Bericht herunterladen",
                data=excel_data,
                file_name=f"Batteriespeicher_Analyse_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="download_excel_stable"
            )
    
